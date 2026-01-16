"""
============================================================================
CONCILIACIÓN BANCARIA INTELIGENTE v6.5 DEFINITIVO
============================================================================
Proyecto: GREEN PARK - BANRESERBAS

MOTOR DEFINITIVO = v6.0 (completo) + v6.1 (estrategias 1.6 y 7) + OPTIMIZACIONES

✅ TODAS LAS ESTRATEGIAS (7 estrategias completas):
   1.   Monto Exacto (1:1)
   1.5  Transferencias + Comisión ($7)
   1.6  Comisiones Agrupadas Multi-fecha [v6.1]
   2.   Agrupaciones N→1 (4 métodos)
   3.   Agrupaciones 1→N
   4.   Agrupaciones N↔M
   5.   Impuestos DGII (0.15%)
   6.   Segunda Pasada Flexible
   7.   Búsqueda Exhaustiva Final [v6.1]

✅ OPTIMIZACIONES v6.5:
   - Lectura rápida: solo filas con Fecha Y Valor válidos
   - Límites por estrategia para evitar timeouts
   - Alias TC/LEGAL mejorados: TC LEGAL ↔ TC Corporativa ↔ Legalizaciones ↔ IPI
   - Descripción vacía funcional (crea columna si no existe)
   - Score combinado para desempate (Concepto_Norm del v5 restaurado)

🎯 OBJETIVO: < 1 minuto de ejecución con máxima conciliación
============================================================================
"""

import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import os
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import re
from difflib import SequenceMatcher
import unicodedata
from itertools import combinations
import warnings
warnings.filterwarnings('ignore')

# ============================================================================
# ⚙️ CONFIGURACIÓN DE RUTAS
# ============================================================================

RUTA_SALIDA = r"...\Consiliacion\Conciliacion_Resultados.xlsx"

# ============================================================================
# ⚙️ PARÁMETROS AJUSTABLES
# ============================================================================

# 💰 TOLERANCIAS DE VALOR
TOLERANCIA_VALOR_EXACTA = 0.01
TOLERANCIA_VALOR_AGRUPACION = 1.00
TOLERANCIA_PORCENTAJE_PARCIAL = 0.02

# 📅 VENTANAS DE TIEMPO
VENTANA_DIAS_EXACTA = 10
VENTANA_DIAS_AGRUPACION = 20
VENTANA_DIAS_FLEXIBLE = 30
VENTANA_DIAS_COMISIONES = 45

# 🎯 UMBRALES DE SIMILITUD
UMBRAL_SIMILITUD_BAJA = 0.05
UMBRAL_SIMILITUD_MEDIA = 0.20
UMBRAL_SIMILITUD_ALTA = 0.40

# 🔧 CONFIGURACIÓN AVANZADA
PERMITIR_SOLO_MONTO = True
USAR_FECHAS_PARA_DESAMBIGUAR = True
DETECTAR_CASOS_ESPECIALES = True
APLICAR_FORMATO_PROFESIONAL = True
EJECUTAR_SEGUNDA_PASADA = True
EJECUTAR_BUSQUEDA_EXHAUSTIVA = True

# 💵 COMISIONES BANCARIAS
COMISION_TRANSFERENCIA_USD = 7.00
DETECTAR_COMISIONES = True

# ⚡ LÍMITES DE RENDIMIENTO (OPTIMIZACIÓN)
MAX_PARTIDAS_AGRUPACION = 30        # Reducido de 100 a 15 para velocidad
MAX_COMBINACIONES_POR_BUSQUEDA = 10000  # Límite por cada búsqueda individual
UMBRAL_PARTIDAS_EXHAUSTIVA = 25     # Aumentado de 20 a 25
MAX_COMBINACIONES_EXHAUSTIVA = 100000

# ============================================================================
# 📝 ALIAS DE EMPRESAS Y CONCEPTOS (AMPLIADO PARA TC/LEGAL)
# ============================================================================

ALIAS_EMPRESAS = {
    # RG-RB / CONCRESOL
    'RGRB': ['RG RB', 'RGRB', 'RG-RB', 'CONCRESOL', 'CONCRESOL SRL'],
    'CONCRESOL': ['RG RB', 'RGRB', 'RG-RB', 'CONCRESOL', 'CONCRESOL SRL'],
    
    # TC / TARJETA CORPORATIVA / LEGALIZACIONES [NUEVO v6.5]
    'TC_CORPORATIVA': [
        'TC LEGAL', 'TC CORPORATIVA', 'TARJETA CORPORATIVA', 'TC',
        'LEGAL', 'LEGALIZACIONES', 'LEGALIZACION', 'IPI',
        'CORTE TC', 'CORTE TC LEGAL', 'PAGO TC', 'REEMBOLSO TC',
        'REEMBOLSO CORTE TC', 'REEMBOLSO TC LEGAL', 'DEBITO PAGO',
        'DEBITO PAGO RD', 'PAGO RD TC', 'TC CORPORATIVA'
    ],
}

# Patrones específicos
PATRONES_TC = [
    'TC LEGAL', 'TC CORPORATIVA', 'TARJETA CORPORATIVA', 'CORTE TC',
    'REEMBOLSO TC', 'DEBITO PAGO', 'LEGALIZACIONES', 'IPI',
    'PAGO RD TC', 'REEMBOLSO CORTE'
]

PATRONES_IMPUESTO = ['IMP', 'IMPUESTO', 'DGII', '0.15', '0.15%', 'ITF', 'RETENCION']

PATRONES_COMISION = [
    'COM', 'COMISION', 'COM.', 'CARGO', 'COSTO', 'GASTO BANCARIO',
    'EXI', 'COMISION BANCARIA', 'COMISIONES', 'LBTR',
    'TRANS APROBA', 'DEBITO CUENTA', 'GASTO', 'SERVICIO'
]

PATRONES_TRANSFERENCIA = ['TRANSFER', 'TRANSF', 'ACH', 'CR TRANSFERENCIA', 'REALIT LLC','LLC','REALIT']

PALABRAS_COMUNES = {
    'DE', 'DEL', 'LA', 'EL', 'LOS', 'LAS', 'CON', 'SIN', 'PARA', 'POR', 'QUE',
    'COMO', 'DESDE', 'HASTA', 'ENTRE', 'ACH', 'TRANSFERENCIA', 'PAGO', 'TRANSF',
    'SAS', 'LTDA', 'SA', 'SRL', 'CIA', 'RECAUDO', 'INMOB', 'FIDEICOMISO',
    'DEPOSITO', 'BANCO', 'CTA', 'CTAS', 'FAC', 'FACT', 'FACTURA', 'NUM',
    'NUMERO', 'NO', 'Y', 'A', 'EN', 'AL', 'O', 'U', 'E', 'USD', 'DOP', 'RD'
}

# ============================================================================
# 🔧 FUNCIONES DE NORMALIZACIÓN
# ============================================================================

def quitar_acentos(texto):
    if pd.isna(texto) or texto is None:
        return ""
    texto = str(texto).strip()
    if not texto:
        return ""
    texto_nfd = unicodedata.normalize('NFD', texto)
    return ''.join(c for c in texto_nfd if unicodedata.category(c) != 'Mn')

def normalizar_texto(texto):
    if pd.isna(texto) or texto is None:
        return ""
    texto = quitar_acentos(str(texto))
    texto = texto.upper()
    texto = re.sub(r'[^A-Z0-9\s]', ' ', texto)
    texto = re.sub(r'\s+', ' ', texto)
    return texto.strip()

def es_patron_tc(texto):
    """Detecta si el texto corresponde a TC/Legalizaciones"""
    texto_norm = normalizar_texto(texto)
    return any(patron in texto_norm for patron in PATRONES_TC)

def normalizar_nombre_empresa(texto):
    """Normaliza nombre de empresa + detecta TC/LEGAL"""
    texto_norm = normalizar_texto(texto)
    if not texto_norm:
        return ""
    
    # PRIMERO: Verificar si es TC/LEGAL (prioridad alta)
    if es_patron_tc(texto):
        return 'TC_CORPORATIVA'
    
    # Eliminar sufijos legales
    sufijos = ['SRL', 'SA', 'SAS', 'LTDA', 'CIA', 'INC', 'LLC', 'CORP']
    for sufijo in sufijos:
        texto_norm = re.sub(rf'\b{sufijo}\b', '', texto_norm)
    
    texto_norm = re.sub(r'\s+', ' ', texto_norm).strip()
    texto_compacto = texto_norm.replace(' ', '')
    
    # Buscar en alias
    for clave, aliases in ALIAS_EMPRESAS.items():
        for alias in aliases:
            alias_compacto = normalizar_texto(alias).replace(' ', '')
            if alias_compacto and texto_compacto:
                if alias_compacto in texto_compacto or texto_compacto in alias_compacto:
                    return clave
    
    return texto_norm

def extraer_palabras_clave(texto):
    texto_norm = normalizar_texto(texto)
    if not texto_norm:
        return []
    palabras = [p for p in texto_norm.split() if len(p) >= 3 and p not in PALABRAS_COMUNES]
    numeros = re.findall(r'\d{5,}', texto_norm)
    return list(set(palabras + numeros))

def extraer_identificador_proveedor(texto):
    texto_norm = normalizar_texto(texto)
    if not texto_norm:
        return None
    match = re.search(r'^\s*(\d{8,})', texto_norm)
    return match.group(1) if match else None

def calcular_similitud(texto1, texto2):
    # Comparar nombres de empresa primero
    emp1 = normalizar_nombre_empresa(texto1)
    emp2 = normalizar_nombre_empresa(texto2)
    
    if emp1 and emp2 and emp1 == emp2:
        return 1.0
    
    palabras1 = set(extraer_palabras_clave(texto1))
    palabras2 = set(extraer_palabras_clave(texto2))
    
    if not palabras1 or not palabras2:
        return 0.0
    
    coincidencias = len(palabras1 & palabras2)
    
    # Coincidencias difusas
    for p1 in palabras1:
        for p2 in palabras2:
            if len(p1) >= 4 and len(p2) >= 4:
                if SequenceMatcher(None, p1, p2).ratio() > 0.85:
                    coincidencias += 0.5
                    break
    
    max_palabras = max(len(palabras1), len(palabras2))
    return coincidencias / max_palabras if max_palabras > 0 else 0.0

def es_patron_impuesto(texto):
    texto_norm = normalizar_texto(texto)
    return any(patron in texto_norm for patron in PATRONES_IMPUESTO)

def es_patron_comision(texto):
    texto_norm = normalizar_texto(texto)
    return any(patron in texto_norm for patron in PATRONES_COMISION)

def es_patron_transferencia(texto):
    texto_norm = normalizar_texto(texto)
    return any(patron in texto_norm for patron in PATRONES_TRANSFERENCIA)

# ============================================================================
# 📂 CARGA DE DATOS OPTIMIZADA (SOLO FILAS VÁLIDAS)
# ============================================================================

# Carga de Archivos BANRESERVAS
def cargar_banreservas(ruta, nombre=""):
    """Carga datos - OPTIMIZACIÓN: solo lee filas con Fecha Y Valor válidos"""
    print(f"\n  📂 Cargando: {nombre.upper() if nombre else ruta}")

    # Resolver ruta relativa: si el usuario pasa solo el nombre de archivo,
    # buscarlo en el mismo directorio del script y en el cwd.
    try:
        ruta_path = Path(ruta)
    except Exception:
        ruta_path = Path(str(ruta))

    if not ruta_path.is_absolute():
        script_dir = Path(__file__).resolve().parent
        candidates = [script_dir / ruta_path, Path.cwd() / ruta_path]

        # If no suffix provided, try common extensions too
        if not ruta_path.suffix:
            exts = ['.xlsx', '.xls', '.csv']
        else:
            exts = [ruta_path.suffix]

        found = None
        for base in candidates:
            for ext in exts:
                cand = base.with_suffix(ext) if not ruta_path.suffix else base
                if cand.exists():
                    found = cand
                    break
            if found:
                break

        if found:
            ruta = str(found)
        else:
            # if the literal provided (e.g., 'BANRESERVAS.xlsx') exists in script_dir
            alt = script_dir / ruta_path.name
            if alt.exists():
                ruta = str(alt)

    try:
        # Leer Excel
        df = pd.read_excel(ruta)

        # ⚡ OPTIMIZACIÓN 1: Eliminar filas completamente vacías PRIMERO
        df = df.dropna(how='all')

    except Exception as e:
        raise ValueError(f"❌ Error al cargar '{ruta}': {e}")
    
    print(f"  📋 Columnas encontradas: {list(df.columns)}")
    
    # Mapeo automático de columnas
    columnas_map = {}
    tiene_debito_credito = False
    for col in df.columns:
        col_upper = str(col).upper().strip()
        if 'FECHA' in col_upper and 'Fecha' not in columnas_map:
            columnas_map['Fecha'] = col
        elif 'CONCEPTO' in col_upper and 'Concepto' not in columnas_map:
            columnas_map['Concepto'] = col
        elif 'DEBITO' in col_upper and 'Debito' not in columnas_map:
            columnas_map['Debito'] = col
        elif 'CREDITO' in col_upper and 'Credito' not in columnas_map:
            columnas_map['Credito'] = col
        elif any(x in col_upper for x in ['VALOR', 'MONTO', 'IMPORTE']) and 'Valor' not in columnas_map:
            columnas_map['Valor'] = col
        elif any(x in col_upper for x in ['DESCRIP', 'DETALLE', 'OBSERV', 'REFERENCIA']) and 'Descripción' not in columnas_map:
            columnas_map['Descripción'] = col
    
    # Detectar si hay Debito y Credito
    tiene_debito_credito = 'Debito' in columnas_map and 'Credito' in columnas_map
    
    # Si no hay Descripción, crearla vacía
    if 'Descripción' not in columnas_map:
        df['Descripción'] = ''
        columnas_map['Descripción'] = 'Descripción'
        print("  ⚠️ Columna 'Descripción' no encontrada - creada vacía")
    
    # Verificar columnas requeridas
    if tiene_debito_credito:
        requeridas = ['Fecha', 'Concepto', 'Debito', 'Credito']
    else:
        requeridas = ['Fecha', 'Concepto', 'Valor']
    
    faltantes = [r for r in requeridas if r not in columnas_map]
    if faltantes:
        raise ValueError(f"❌ Faltan columnas requeridas: {faltantes}")
    
    # Renombrar columnas
    df = df.rename(columns={v: k for k, v in columnas_map.items()})
    
    # Convertir tipos y combinar Debito/Credito si es necesario
    df['Fecha'] = pd.to_datetime(df['Fecha'], errors='coerce')
    df['Concepto'] = df['Concepto'].fillna('').astype(str)
    df['Descripción'] = df['Descripción'].fillna('').astype(str)
    
    if tiene_debito_credito:
        # Convertir Debito y Credito a numéricos
        df['Debito'] = pd.to_numeric(df['Debito'], errors='coerce').fillna(0)
        df['Credito'] = pd.to_numeric(df['Credito'], errors='coerce').fillna(0)
        # Crear Valor: Credito positivo, Debito negativo
        df['Valor'] = df['Credito'] - df['Debito']
        # Eliminar columnas originales
        df = df.drop(columns=['Debito', 'Credito'])
    else:
        df['Valor'] = pd.to_numeric(df['Valor'], errors='coerce')
    
    # ⚡ OPTIMIZACIÓN 2: Filtrar SOLO filas con Fecha Y Valor válidos
    registros_antes = len(df)
    df = df.dropna(subset=['Fecha', 'Valor'])
    df = df[df['Valor'] != 0]  # Eliminar valores $0
    registros_despues = len(df)
    
    if registros_antes != registros_despues:
        print(f"  ⚠️ Filtradas {registros_antes - registros_despues} filas sin Fecha/Valor válidos")
    
    # Campos de búsqueda
    df['Texto_Busqueda'] = (df['Concepto'].astype(str) + ' ' + df['Descripción'].astype(str)).apply(normalizar_texto)
    df['Concepto_Norm'] = df['Concepto'].apply(normalizar_texto)  # ← RESTAURADO del v5
    df['Proveedor_ID'] = df['Descripción'].apply(extraer_identificador_proveedor)
    df['Empresa_Norm'] = df['Texto_Busqueda'].apply(normalizar_nombre_empresa)
    df['Es_Impuesto'] = df['Texto_Busqueda'].apply(es_patron_impuesto)
    df['Es_Comision'] = df['Texto_Busqueda'].apply(es_patron_comision)
    df['Es_TC'] = df['Texto_Busqueda'].apply(es_patron_tc)  # ← NUEVO v6.5
    
    # Control
    df['ID_Original'] = range(len(df))
    df['Conciliado'] = False
    
    print(f"  ✅ Cargados: {len(df)} registros válidos")
    return df

# Cargas de Archivos CONTABLE
def cargar_contable(ruta, nombre=""):
    """Carga datos - OPTIMIZACIÓN: solo lee filas con Fecha Y Valor válidos"""
    print(f"\n  📂 Cargando: {nombre.upper() if nombre else ruta}")

    # Resolver ruta relativa: si el usuario pasa solo el nombre de archivo,
    # buscarlo en el mismo directorio del script y en el cwd.
    try:
        ruta_path = Path(ruta)
    except Exception:
        ruta_path = Path(str(ruta))

    if not ruta_path.is_absolute():
        script_dir = Path(__file__).resolve().parent
        candidates = [script_dir / ruta_path, Path.cwd() / ruta_path]

        # If no suffix provided, try common extensions too
        if not ruta_path.suffix:
            exts = ['.xlsx', '.xls', '.csv']
        else:
            exts = [ruta_path.suffix]

        found = None
        for base in candidates:
            for ext in exts:
                cand = base.with_suffix(ext) if not ruta_path.suffix else base
                if cand.exists():
                    found = cand
                    break
            if found:
                break

        if found:
            ruta = str(found)
        else:
            # if the literal provided (e.g., 'BANRESERVAS.xlsx') exists in script_dir
            alt = script_dir / ruta_path.name
            if alt.exists():
                ruta = str(alt)

    try:
        # Leer Excel
        df = pd.read_excel(ruta)

        # ⚡ OPTIMIZACIÓN 1: Eliminar filas completamente vacías PRIMERO
        df = df.dropna(how='all')

    except Exception as e:
        raise ValueError(f"❌ Error al cargar '{ruta}': {e}")
    
    print(f"  📋 Columnas encontradas: {list(df.columns)}")
    
    # Mapeo automático de columnas
    columnas_map = {}
    tiene_debito_credito = False
    for col in df.columns:
        col_upper = str(col).upper().strip()
        if 'FECHA' in col_upper and 'Fecha' not in columnas_map:
            columnas_map['Fecha'] = col
        elif 'CONCEPTO' in col_upper and 'Concepto' not in columnas_map:
            columnas_map['Concepto'] = col
        elif 'DEBITO' in col_upper and 'Debito' not in columnas_map:
            columnas_map['Debito'] = col
        elif 'CREDITO' in col_upper and 'Credito' not in columnas_map:
            columnas_map['Credito'] = col
        elif any(x in col_upper for x in ['VALOR', 'MONTO', 'IMPORTE']) and 'Valor' not in columnas_map:
            columnas_map['Valor'] = col
        elif any(x in col_upper for x in ['DESCRIP', 'DETALLE', 'OBSERV', 'REFERENCIA']) and 'Descripción' not in columnas_map:
            columnas_map['Descripción'] = col
    
    # Detectar si hay Debito y Credito
    tiene_debito_credito = 'Debito' in columnas_map and 'Credito' in columnas_map
    
    # Si no hay Descripción, crearla vacía
    if 'Descripción' not in columnas_map:
        df['Descripción'] = ''
        columnas_map['Descripción'] = 'Descripción'
        print("  ⚠️ Columna 'Descripción' no encontrada - creada vacía")
    
    # Verificar columnas requeridas
    if tiene_debito_credito:
        requeridas = ['Fecha', 'Concepto', 'Debito', 'Credito']
    else:
        requeridas = ['Fecha', 'Concepto', 'Valor']
    
    faltantes = [r for r in requeridas if r not in columnas_map]
    if faltantes:
        raise ValueError(f"❌ Faltan columnas requeridas: {faltantes}")
    
    # Renombrar columnas
    df = df.rename(columns={v: k for k, v in columnas_map.items()})
    
    # Convertir tipos y combinar Debito/Credito si es necesario
    df['Fecha'] = pd.to_datetime(df['Fecha'], errors='coerce')
    df['Concepto'] = df['Concepto'].fillna('').astype(str)
    df['Descripción'] = df['Descripción'].fillna('').astype(str)
    
    if tiene_debito_credito:
        # Convertir Debito y Credito a numéricos
        df['Debito'] = pd.to_numeric(df['Debito'], errors='coerce').fillna(0)
        df['Credito'] = pd.to_numeric(df['Credito'], errors='coerce').fillna(0)
        # Crear Valor: Credito positivo, Debito negativo
        df['Valor'] = df['Credito'] - df['Debito']
        # Eliminar columnas originales
        df = df.drop(columns=['Debito', 'Credito'])
    else:
        df['Valor'] = pd.to_numeric(df['Valor'], errors='coerce')
    
    # ⚡ OPTIMIZACIÓN 2: Filtrar SOLO filas con Fecha Y Valor válidos
    registros_antes = len(df)
    df = df.dropna(subset=['Fecha', 'Valor'])
    df = df[df['Valor'] != 0]  # Eliminar valores $0
    registros_despues = len(df)
    
    if registros_antes != registros_despues:
        print(f"  ⚠️ Filtradas {registros_antes - registros_despues} filas sin Fecha/Valor válidos")
    
    # Campos de búsqueda
    df['Texto_Busqueda'] = (df['Concepto'].astype(str) + ' ' + df['Descripción'].astype(str)).apply(normalizar_texto)
    df['Concepto_Norm'] = df['Concepto'].apply(normalizar_texto)  # ← RESTAURADO del v5
    df['Proveedor_ID'] = df['Descripción'].apply(extraer_identificador_proveedor)
    df['Empresa_Norm'] = df['Texto_Busqueda'].apply(normalizar_nombre_empresa)
    df['Es_Impuesto'] = df['Texto_Busqueda'].apply(es_patron_impuesto)
    df['Es_Comision'] = df['Texto_Busqueda'].apply(es_patron_comision)
    df['Es_TC'] = df['Texto_Busqueda'].apply(es_patron_tc)  # ← NUEVO v6.5
    
    # Control
    df['ID_Original'] = range(len(df))
    df['Conciliado'] = False
    
    print(f"  ✅ Cargados: {len(df)} registros válidos")
    return df

# ============================================================================
# 🎯 ESTRATEGIA 1: MONTO EXACTO (1:1) - CON SCORE COMBINADO
# ============================================================================

def conciliacion_por_monto_exacto(banco, contable, conciliaciones):
    print("\n" + "="*70)
    print("🎯 ESTRATEGIA 1: MONTO EXACTO (1:1)")
    print("="*70)
    
    contador = 0
    id_conc = len(conciliaciones) + 1
    
    for idx_b, reg_b in banco[~banco['Conciliado']].iterrows():
        f_min = reg_b['Fecha'] - timedelta(days=VENTANA_DIAS_EXACTA)
        f_max = reg_b['Fecha'] + timedelta(days=VENTANA_DIAS_EXACTA)
        
        cands = contable[
            (~contable['Conciliado']) &
            (contable['Fecha'] >= f_min) &
            (contable['Fecha'] <= f_max) &
            (abs(contable['Valor'] - reg_b['Valor']) < TOLERANCIA_VALOR_EXACTA)
        ]
        
        if len(cands) == 0:
            continue
        
        if len(cands) == 1:
            reg_c = cands.iloc[0]
            sim = calcular_similitud(reg_b['Texto_Busqueda'], reg_c['Texto_Busqueda'])
            
            if PERMITIR_SOLO_MONTO or sim >= UMBRAL_SIMILITUD_BAJA:
                conciliaciones.append({
                    'ID_Conciliacion': id_conc,
                    'Tipo': '1. Monto Exacto (1:1)',
                    'Similitud': round(sim, 3),
                    'Fecha_Banco': reg_b['Fecha'],
                    'Concepto_Banco': reg_b['Concepto'],
                    'Valor_Banco': reg_b['Valor'],
                    'Descripcion_Banco': reg_b['Descripción'],
                    'Fecha_Contable': reg_c['Fecha'],
                    'Concepto_Contable': reg_c['Concepto'],
                    'Valor_Contable': reg_c['Valor'],
                    'Descripcion_Contable': reg_c['Descripción'],
                    'Diferencia': abs(reg_b['Valor'] - reg_c['Valor'])
                })
                banco.loc[idx_b, 'Conciliado'] = True
                contable.loc[reg_c.name, 'Conciliado'] = True
                contador += 1
                id_conc += 1
        else:
            # ⭐ SCORE COMBINADO para desempatar (restaurado del v5)
            cands = cands.copy()
            cands['Sim'] = cands['Texto_Busqueda'].apply(
                lambda x: calcular_similitud(reg_b['Texto_Busqueda'], x)
            )
            cands['Dias_Diff'] = abs((cands['Fecha'] - reg_b['Fecha']).dt.days)
            cands['Score'] = (cands['Sim'] * 0.7) + ((VENTANA_DIAS_EXACTA - cands['Dias_Diff']) / VENTANA_DIAS_EXACTA * 0.3)
            
            mejor_idx = cands['Score'].idxmax()
            mejor_sim = cands.loc[mejor_idx, 'Sim']
            
            if PERMITIR_SOLO_MONTO or mejor_sim >= UMBRAL_SIMILITUD_BAJA:
                reg_c = contable.loc[mejor_idx]
                conciliaciones.append({
                    'ID_Conciliacion': id_conc,
                    'Tipo': '1. Monto Exacto (1:1)',
                    'Similitud': round(mejor_sim, 3),
                    'Fecha_Banco': reg_b['Fecha'],
                    'Concepto_Banco': reg_b['Concepto'],
                    'Valor_Banco': reg_b['Valor'],
                    'Descripcion_Banco': reg_b['Descripción'],
                    'Fecha_Contable': reg_c['Fecha'],
                    'Concepto_Contable': reg_c['Concepto'],
                    'Valor_Contable': reg_c['Valor'],
                    'Descripcion_Contable': reg_c['Descripción'],
                    'Diferencia': abs(reg_b['Valor'] - reg_c['Valor'])
                })
                banco.loc[idx_b, 'Conciliado'] = True
                contable.loc[mejor_idx, 'Conciliado'] = True
                contador += 1
                id_conc += 1
    
    print(f"✓ Conciliaciones 1:1: {contador}")
    return contador

# ============================================================================
# 🎯 ESTRATEGIA 1.5: TRANSFERENCIAS CON COMISIÓN ($7)
# ============================================================================

def conciliacion_con_comisiones(banco, contable, conciliaciones):
    if not DETECTAR_COMISIONES:
        return 0
    
    print("\n" + "="*70)
    print("🎯 ESTRATEGIA 1.5: TRANSFERENCIAS CON COMISIÓN ($7)")
    print("="*70)
    
    contador = 0
    id_conc = len(conciliaciones) + 1 if conciliaciones else 1
    
    banco_no_conc = banco[~banco['Conciliado']].copy()
    
    for idx_com, reg_com in banco_no_conc[banco_no_conc['Es_Comision']].iterrows():
        if abs(abs(reg_com['Valor']) - COMISION_TRANSFERENCIA_USD) > 0.50:
            continue
        
        fecha_com = reg_com['Fecha']
        candidatos_transf = banco_no_conc[
            (~banco_no_conc['Conciliado']) &
            (banco_no_conc['Fecha'] == fecha_com) &
            (banco_no_conc['Valor'] > 0) &
            (banco_no_conc.index != idx_com)
        ]
        
        for idx_transf, reg_transf in candidatos_transf.iterrows():
            valor_neto = reg_transf['Valor'] - COMISION_TRANSFERENCIA_USD
            
            f_min = fecha_com - timedelta(days=VENTANA_DIAS_EXACTA)
            f_max = fecha_com + timedelta(days=VENTANA_DIAS_EXACTA)
            
            cands_cont = contable[
                (~contable['Conciliado']) &
                (contable['Fecha'] >= f_min) &
                (contable['Fecha'] <= f_max) &
                (abs(contable['Valor'] - valor_neto) < TOLERANCIA_VALOR_EXACTA)
            ]
            
            if len(cands_cont) >= 1:
                reg_c = cands_cont.iloc[0]
                sim = calcular_similitud(reg_transf['Texto_Busqueda'], reg_c['Texto_Busqueda'])
                
                conciliaciones.append({
                    'ID_Conciliacion': id_conc,
                    'Tipo': '1.5. Transf+Comisión (2→1)',
                    'Similitud': round(sim, 3),
                    'Fecha_Banco': reg_com['Fecha'],
                    'Concepto_Banco': reg_com['Concepto'],
                    'Valor_Banco': reg_com['Valor'],
                    'Descripcion_Banco': reg_com['Descripción'],
                    'Fecha_Contable': None,
                    'Concepto_Contable': '(Comisión bancaria)',
                    'Valor_Contable': None,
                    'Descripcion_Contable': '',
                    'Diferencia': None
                })
                
                conciliaciones.append({
                    'ID_Conciliacion': id_conc,
                    'Tipo': '1.5. Transf+Comisión (2→1)',
                    'Similitud': round(sim, 3),
                    'Fecha_Banco': reg_transf['Fecha'],
                    'Concepto_Banco': reg_transf['Concepto'],
                    'Valor_Banco': reg_transf['Valor'],
                    'Descripcion_Banco': reg_transf['Descripción'],
                    'Fecha_Contable': reg_c['Fecha'],
                    'Concepto_Contable': reg_c['Concepto'],
                    'Valor_Contable': reg_c['Valor'],
                    'Descripcion_Contable': reg_c['Descripción'],
                    'Diferencia': COMISION_TRANSFERENCIA_USD
                })
                
                banco.loc[idx_com, 'Conciliado'] = True
                banco.loc[idx_transf, 'Conciliado'] = True
                contable.loc[reg_c.name, 'Conciliado'] = True
                
                contador += 1
                id_conc += 1
                break
    
    print(f"✓ Transferencias con comisión: {contador}")
    return contador

# ============================================================================
# 🎯 ESTRATEGIA 1.6: COMISIONES AGRUPADAS MULTI-FECHA [v6.1]
# ============================================================================

def conciliacion_comisiones_agrupadas(banco, contable, conciliaciones):
    print("\n" + "="*70)
    print("🎯 ESTRATEGIA 1.6: COMISIONES AGRUPADAS (Multi-fecha)")
    print("="*70)
    
    contador = 0
    id_conc = len(conciliaciones) + 1 if conciliaciones else 1
    
    contable_com = contable[(~contable['Conciliado']) & (contable['Es_Comision'])].copy()
    
    if len(contable_com) == 0:
        print("⊘ No hay comisiones pendientes en contable")
        return 0
    
    banco_com = banco[(~banco['Conciliado']) & (banco['Es_Comision'])].copy()
    
    if len(banco_com) == 0:
        print("⊘ No hay comisiones pendientes en banco")
        return 0
    
    for idx_c, reg_c in contable_com.iterrows():
        if contable.loc[idx_c, 'Conciliado']:
            continue
        
        valor_objetivo = reg_c['Valor']
        
        banco_com_pend = banco_com[~banco_com.index.isin(banco[banco['Conciliado']].index)]
        if len(banco_com_pend) == 0:
            continue
        
        # Ventana amplia para comisiones dispersas
        mes_contable = reg_c['Fecha'].month if pd.notna(reg_c['Fecha']) else None
        if mes_contable:
            banco_com_mes = banco_com_pend[
                (banco_com_pend['Fecha'].dt.month >= mes_contable - 1) &
                (banco_com_pend['Fecha'].dt.month <= mes_contable + 1)
            ]
        else:
            banco_com_mes = banco_com_pend
        
        if len(banco_com_mes) == 0:
            continue
        
        # Búsqueda combinatoria LIMITADA
        mejor_combo, mejor_diff = None, float('inf')
        indices = banco_com_mes.index.tolist()[:MAX_PARTIDAS_AGRUPACION]
        
        total_combinaciones = 0
        for r in range(1, len(indices) + 1):
            if mejor_diff < TOLERANCIA_VALOR_AGRUPACION:
                break
            if total_combinaciones >= MAX_COMBINACIONES_POR_BUSQUEDA:
                break
            
            for combo in combinations(indices, r):
                total_combinaciones += 1
                if total_combinaciones > MAX_COMBINACIONES_POR_BUSQUEDA:
                    break
                
                suma = banco.loc[list(combo), 'Valor'].sum()
                diff = abs(suma - valor_objetivo)
                
                if diff < mejor_diff:
                    mejor_diff = diff
                    mejor_combo = combo
                
                if diff < TOLERANCIA_VALOR_AGRUPACION:
                    break
        
        if mejor_combo and mejor_diff < TOLERANCIA_VALOR_AGRUPACION:
            grupo = banco.loc[list(mejor_combo)]
            
            for i, (idx_b, reg_b) in enumerate(grupo.sort_values(['Fecha', 'Valor']).iterrows()):
                conciliaciones.append({
                    'ID_Conciliacion': id_conc,
                    'Tipo': f'1.6. Comisiones Agrupadas ({len(grupo)}→1)',
                    'Similitud': 1.0,
                    'Fecha_Banco': reg_b['Fecha'],
                    'Concepto_Banco': reg_b['Concepto'],
                    'Valor_Banco': reg_b['Valor'],
                    'Descripcion_Banco': reg_b['Descripción'],
                    'Fecha_Contable': reg_c['Fecha'] if i == 0 else None,
                    'Concepto_Contable': reg_c['Concepto'] if i == 0 else '',
                    'Valor_Contable': reg_c['Valor'] if i == 0 else None,
                    'Descripcion_Contable': reg_c['Descripción'] if i == 0 else '',
                    'Diferencia': mejor_diff if i == 0 else None
                })
            
            banco.loc[list(mejor_combo), 'Conciliado'] = True
            contable.loc[idx_c, 'Conciliado'] = True
            contador += 1
            id_conc += 1
    
    print(f"✓ Comisiones agrupadas: {contador}")
    return contador

# ============================================================================
# 🎯 ESTRATEGIA 2: N→1 INTELIGENTE (4 MÉTODOS) - RESTAURADO DEL v5
# ============================================================================

def conciliacion_n_a_1_inteligente(banco, contable, conciliaciones):
    print("\n" + "="*70)
    print("🎯 ESTRATEGIA 2: N→1 (Varios Contable → Uno Banco)")
    print("="*70)
    
    contador = 0
    id_conc = len(conciliaciones) + 1 if conciliaciones else 1
    
    for idx_b, reg_b in banco[~banco['Conciliado']].iterrows():
        if banco.loc[idx_b, 'Conciliado']:
            continue
            
        f_min = reg_b['Fecha'] - timedelta(days=VENTANA_DIAS_AGRUPACION)
        f_max = reg_b['Fecha'] + timedelta(days=VENTANA_DIAS_AGRUPACION)
        
        cands = contable[
            (~contable['Conciliado']) &
            (contable['Fecha'] >= f_min) &
            (contable['Fecha'] <= f_max)
        ].copy()
        
        if len(cands) < 2:
            continue
        
        encontrado = False
        
        # MÉTODO 1: Por Proveedor_ID con desambiguación por fechas
        if USAR_FECHAS_PARA_DESAMBIGUAR and 'Proveedor_ID' in cands.columns:
            for prov_id in cands['Proveedor_ID'].dropna().unique():
                if encontrado:
                    break
                    
                grupo_prov = cands[cands['Proveedor_ID'] == prov_id]
                if len(grupo_prov) < 2:
                    continue
                
                for fecha_grupo in grupo_prov['Fecha'].unique():
                    grupo = grupo_prov[
                        (grupo_prov['Fecha'] >= fecha_grupo - timedelta(days=3)) &
                        (grupo_prov['Fecha'] <= fecha_grupo + timedelta(days=3))
                    ]
                    
                    if len(grupo) < 2:
                        continue
                    
                    suma = grupo['Valor'].sum()
                    if abs(suma - reg_b['Valor']) < TOLERANCIA_VALOR_AGRUPACION:
                        sim_media = grupo['Texto_Busqueda'].apply(
                            lambda x: calcular_similitud(reg_b['Texto_Busqueda'], x)
                        ).mean()
                        
                        for i, (_, reg_c) in enumerate(grupo.sort_values(['Fecha', 'Valor']).iterrows()):
                            conciliaciones.append({
                                'ID_Conciliacion': id_conc,
                                'Tipo': f'2. Agrupación N→1 ({len(grupo)} reg)',
                                'Similitud': round(sim_media, 3),
                                'Fecha_Banco': reg_b['Fecha'] if i == 0 else None,
                                'Concepto_Banco': reg_b['Concepto'] if i == 0 else '',
                                'Valor_Banco': reg_b['Valor'] if i == 0 else None,
                                'Descripcion_Banco': reg_b['Descripción'] if i == 0 else '',
                                'Fecha_Contable': reg_c['Fecha'],
                                'Concepto_Contable': reg_c['Concepto'],
                                'Valor_Contable': reg_c['Valor'],
                                'Descripcion_Contable': reg_c['Descripción'],
                                'Diferencia': abs(reg_b['Valor'] - suma) if i == 0 else None
                            })
                        
                        banco.loc[idx_b, 'Conciliado'] = True
                        contable.loc[grupo.index, 'Conciliado'] = True
                        contador += 1
                        id_conc += 1
                        encontrado = True
                        break
        
        if encontrado:
            continue
        
        # ⭐ MÉTODO 2: Por Concepto_Norm (RESTAURADO del v5)
        for concepto in cands['Concepto_Norm'].unique():
            if encontrado or not concepto:
                break
            
            grupo = cands[cands['Concepto_Norm'] == concepto]
            if len(grupo) < 2:
                continue
            
            suma = grupo['Valor'].sum()
            if abs(suma - reg_b['Valor']) < TOLERANCIA_VALOR_AGRUPACION:
                sim_media = grupo['Texto_Busqueda'].apply(
                    lambda x: calcular_similitud(reg_b['Texto_Busqueda'], x)
                ).mean()
                
                for i, (_, reg_c) in enumerate(grupo.sort_values(['Fecha', 'Valor']).iterrows()):
                    conciliaciones.append({
                        'ID_Conciliacion': id_conc,
                        'Tipo': f'2. Agrupación N→1 ({len(grupo)} reg)',
                        'Similitud': round(sim_media, 3),
                        'Fecha_Banco': reg_b['Fecha'] if i == 0 else None,
                        'Concepto_Banco': reg_b['Concepto'] if i == 0 else '',
                        'Valor_Banco': reg_b['Valor'] if i == 0 else None,
                        'Descripcion_Banco': reg_b['Descripción'] if i == 0 else '',
                        'Fecha_Contable': reg_c['Fecha'],
                        'Concepto_Contable': reg_c['Concepto'],
                        'Valor_Contable': reg_c['Valor'],
                        'Descripcion_Contable': reg_c['Descripción'],
                        'Diferencia': abs(reg_b['Valor'] - suma) if i == 0 else None
                    })
                
                banco.loc[idx_b, 'Conciliado'] = True
                contable.loc[grupo.index, 'Conciliado'] = True
                contador += 1
                id_conc += 1
                encontrado = True
        
        if encontrado:
            continue
        
        # MÉTODO 3: Por empresa normalizada
        for emp_norm in cands['Empresa_Norm'].unique():
            if encontrado or not emp_norm:
                break
                
            grupo = cands[cands['Empresa_Norm'] == emp_norm]
            if len(grupo) < 2:
                continue
            
            suma = grupo['Valor'].sum()
            if abs(suma - reg_b['Valor']) < TOLERANCIA_VALOR_AGRUPACION:
                sim_media = grupo['Texto_Busqueda'].apply(
                    lambda x: calcular_similitud(reg_b['Texto_Busqueda'], x)
                ).mean()
                
                for i, (_, reg_c) in enumerate(grupo.sort_values(['Fecha', 'Valor']).iterrows()):
                    conciliaciones.append({
                        'ID_Conciliacion': id_conc,
                        'Tipo': f'2. Agrupación N→1 ({len(grupo)} reg)',
                        'Similitud': round(sim_media, 3),
                        'Fecha_Banco': reg_b['Fecha'] if i == 0 else None,
                        'Concepto_Banco': reg_b['Concepto'] if i == 0 else '',
                        'Valor_Banco': reg_b['Valor'] if i == 0 else None,
                        'Descripcion_Banco': reg_b['Descripción'] if i == 0 else '',
                        'Fecha_Contable': reg_c['Fecha'],
                        'Concepto_Contable': reg_c['Concepto'],
                        'Valor_Contable': reg_c['Valor'],
                        'Descripcion_Contable': reg_c['Descripción'],
                        'Diferencia': abs(reg_b['Valor'] - suma) if i == 0 else None
                    })
                
                banco.loc[idx_b, 'Conciliado'] = True
                contable.loc[grupo.index, 'Conciliado'] = True
                contador += 1
                id_conc += 1
                encontrado = True
        
        if encontrado:
            continue
        
        # MÉTODO 4: Por palabras clave
        cands['PalKey'] = cands['Texto_Busqueda'].apply(
            lambda x: ' '.join(sorted(extraer_palabras_clave(x)))
        )
        
        for key, grupo in cands.groupby('PalKey'):
            if len(grupo) < 2 or not key:
                continue
            
            suma = grupo['Valor'].sum()
            if abs(suma - reg_b['Valor']) < TOLERANCIA_VALOR_AGRUPACION:
                sim_media = grupo['Texto_Busqueda'].apply(
                    lambda x: calcular_similitud(reg_b['Texto_Busqueda'], x)
                ).mean()
                
                if sim_media >= UMBRAL_SIMILITUD_BAJA:
                    for i, (_, reg_c) in enumerate(grupo.sort_values(['Fecha', 'Valor']).iterrows()):
                        conciliaciones.append({
                            'ID_Conciliacion': id_conc,
                            'Tipo': f'2. Agrupación N→1 ({len(grupo)} reg)',
                            'Similitud': round(sim_media, 3),
                            'Fecha_Banco': reg_b['Fecha'] if i == 0 else None,
                            'Concepto_Banco': reg_b['Concepto'] if i == 0 else '',
                            'Valor_Banco': reg_b['Valor'] if i == 0 else None,
                            'Descripcion_Banco': reg_b['Descripción'] if i == 0 else '',
                            'Fecha_Contable': reg_c['Fecha'],
                            'Concepto_Contable': reg_c['Concepto'],
                            'Valor_Contable': reg_c['Valor'],
                            'Descripcion_Contable': reg_c['Descripción'],
                            'Diferencia': abs(reg_b['Valor'] - suma) if i == 0 else None
                        })
                    
                    banco.loc[idx_b, 'Conciliado'] = True
                    contable.loc[grupo.index, 'Conciliado'] = True
                    contador += 1
                    id_conc += 1
                    break
    
    print(f"✓ Agrupaciones N→1: {contador}")
    return contador

# ============================================================================
# 🎯 ESTRATEGIA 3: 1→N
# ============================================================================

def conciliacion_1_a_n_inteligente(banco, contable, conciliaciones):
    print("\n" + "="*70)
    print("🎯 ESTRATEGIA 3: 1→N (Uno Contable → Varios Banco)")
    print("="*70)
    
    contador = 0
    id_conc = len(conciliaciones) + 1 if conciliaciones else 1
    
    for idx_c, reg_c in contable[~contable['Conciliado']].iterrows():
        if contable.loc[idx_c, 'Conciliado']:
            continue
            
        f_min = reg_c['Fecha'] - timedelta(days=VENTANA_DIAS_AGRUPACION)
        f_max = reg_c['Fecha'] + timedelta(days=VENTANA_DIAS_AGRUPACION)
        
        cands = banco[
            (~banco['Conciliado']) &
            (banco['Fecha'] >= f_min) &
            (banco['Fecha'] <= f_max)
        ].copy()
        
        if len(cands) < 2:
            continue
        
        cands['PalKey'] = cands['Texto_Busqueda'].apply(
            lambda x: ' '.join(sorted(extraer_palabras_clave(x)))
        )
        
        for key, grupo in cands.groupby('PalKey'):
            if len(grupo) < 2 or not key:
                continue
            
            suma = grupo['Valor'].sum()
            if abs(suma - reg_c['Valor']) < TOLERANCIA_VALOR_AGRUPACION:
                sim_media = grupo['Texto_Busqueda'].apply(
                    lambda x: calcular_similitud(reg_c['Texto_Busqueda'], x)
                ).mean()
                
                if sim_media >= UMBRAL_SIMILITUD_BAJA:
                    for i, (_, reg_b) in enumerate(grupo.sort_values(['Fecha', 'Valor']).iterrows()):
                        conciliaciones.append({
                            'ID_Conciliacion': id_conc,
                            'Tipo': f'3. Agrupación 1→N ({len(grupo)} reg)',
                            'Similitud': round(sim_media, 3),
                            'Fecha_Banco': reg_b['Fecha'],
                            'Concepto_Banco': reg_b['Concepto'],
                            'Valor_Banco': reg_b['Valor'],
                            'Descripcion_Banco': reg_b['Descripción'],
                            'Fecha_Contable': reg_c['Fecha'] if i == 0 else None,
                            'Concepto_Contable': reg_c['Concepto'] if i == 0 else '',
                            'Valor_Contable': reg_c['Valor'] if i == 0 else None,
                            'Descripcion_Contable': reg_c['Descripción'] if i == 0 else '',
                            'Diferencia': abs(suma - reg_c['Valor']) if i == 0 else None
                        })
                    
                    contable.loc[idx_c, 'Conciliado'] = True
                    banco.loc[grupo.index, 'Conciliado'] = True
                    contador += 1
                    id_conc += 1
                    break
    
    print(f"✓ Agrupaciones 1→N: {contador}")
    return contador

# ============================================================================
# 🎯 ESTRATEGIA 4: N↔M
# ============================================================================

def conciliacion_n_a_m_inteligente(banco, contable, conciliaciones):
    print("\n" + "="*70)
    print("🎯 ESTRATEGIA 4: N↔M (Varios ↔ Varios)")
    print("="*70)
    
    contador = 0
    id_conc = len(conciliaciones) + 1 if conciliaciones else 1
    
    cont_temp = contable[~contable['Conciliado']].copy()
    cont_temp['PalKey'] = cont_temp['Texto_Busqueda'].apply(
        lambda x: ' '.join(sorted(extraer_palabras_clave(x)))
    )
    grupos_c = cont_temp.groupby('PalKey').filter(lambda x: len(x) >= 2)
    
    for key_c, grupo_c in grupos_c.groupby('PalKey'):
        if len(grupo_c) < 2 or not key_c:
            continue
        
        grupo_c = grupo_c[~grupo_c.index.isin(contable[contable['Conciliado']].index)]
        if len(grupo_c) < 2:
            continue
        
        suma_c = grupo_c['Valor'].sum()
        f_media = grupo_c['Fecha'].mean()
        f_min = f_media - timedelta(days=VENTANA_DIAS_AGRUPACION)
        f_max = f_media + timedelta(days=VENTANA_DIAS_AGRUPACION)
        
        cands_b = banco[
            (~banco['Conciliado']) &
            (banco['Fecha'] >= f_min) &
            (banco['Fecha'] <= f_max)
        ].copy()
        
        if len(cands_b) < 2:
            continue
        
        cands_b['PalKey'] = cands_b['Texto_Busqueda'].apply(
            lambda x: ' '.join(sorted(extraer_palabras_clave(x)))
        )
        
        for key_b, grupo_b in cands_b.groupby('PalKey'):
            if len(grupo_b) < 2 or not key_b:
                continue
            
            suma_b = grupo_b['Valor'].sum()
            
            if abs(suma_b - suma_c) < TOLERANCIA_VALOR_AGRUPACION:
                sim_total = sum(
                    calcular_similitud(rb['Texto_Busqueda'], rc['Texto_Busqueda'])
                    for _, rb in grupo_b.iterrows()
                    for _, rc in grupo_c.iterrows()
                )
                sim_media = sim_total / (len(grupo_b) * len(grupo_c))
                
                if sim_media >= UMBRAL_SIMILITUD_BAJA:
                    regs_b = grupo_b.sort_values(['Fecha', 'Valor'])
                    regs_c = grupo_c.sort_values(['Fecha', 'Valor'])
                    max_regs = max(len(regs_b), len(regs_c))
                    
                    for i in range(max_regs):
                        rb = regs_b.iloc[i] if i < len(regs_b) else None
                        rc = regs_c.iloc[i] if i < len(regs_c) else None
                        
                        conciliaciones.append({
                            'ID_Conciliacion': id_conc,
                            'Tipo': f'4. Agrupación N↔M ({len(regs_b)}↔{len(regs_c)})',
                            'Similitud': round(sim_media, 3),
                            'Fecha_Banco': rb['Fecha'] if rb is not None else None,
                            'Concepto_Banco': rb['Concepto'] if rb is not None else '',
                            'Valor_Banco': rb['Valor'] if rb is not None else None,
                            'Descripcion_Banco': rb['Descripción'] if rb is not None else '',
                            'Fecha_Contable': rc['Fecha'] if rc is not None else None,
                            'Concepto_Contable': rc['Concepto'] if rc is not None else '',
                            'Valor_Contable': rc['Valor'] if rc is not None else None,
                            'Descripcion_Contable': rc['Descripción'] if rc is not None else '',
                            'Diferencia': abs(suma_b - suma_c) if i == 0 else None
                        })
                    
                    banco.loc[regs_b.index, 'Conciliado'] = True
                    contable.loc[regs_c.index, 'Conciliado'] = True
                    contador += 1
                    id_conc += 1
                    break
    
    print(f"✓ Agrupaciones N↔M: {contador}")
    return contador

# ============================================================================
# 🎯 ESTRATEGIA 5: IMPUESTOS DGII
# ============================================================================

def conciliacion_impuestos(banco, contable, conciliaciones):
    print("\n" + "="*70)
    print("🎯 ESTRATEGIA 5: IMPUESTOS DGII (0.15%)")
    print("="*70)
    
    contador = 0
    id_conc = len(conciliaciones) + 1 if conciliaciones else 1
    
    banco_imp = banco[(~banco['Conciliado']) & (banco['Es_Impuesto'])].copy()
    
    if len(banco_imp) == 0:
        print("⊘ No hay impuestos pendientes en banco")
        return 0
    
    contable_imp = contable[
        (~contable['Conciliado']) &
        (contable['Es_Impuesto'])
    ].copy()
    
    if len(contable_imp) == 0:
        print("⊘ No hay impuestos pendientes en contable")
        return 0
    
    banco_imp['Mes'] = banco_imp['Fecha'].dt.to_period('M')
    
    for mes, grupo_banco in banco_imp.groupby('Mes'):
        suma_banco = abs(grupo_banco['Valor'].sum())
        
        for idx_c, reg_c in contable_imp.iterrows():
            if contable.loc[idx_c, 'Conciliado']:
                continue
            
            if abs(abs(reg_c['Valor']) - suma_banco) < TOLERANCIA_VALOR_AGRUPACION:
                for i, (idx_b, reg_b) in enumerate(grupo_banco.iterrows()):
                    conciliaciones.append({
                        'ID_Conciliacion': id_conc,
                        'Tipo': f'5. Impuestos DGII ({len(grupo_banco)}→1)',
                        'Similitud': 1.0,
                        'Fecha_Banco': reg_b['Fecha'],
                        'Concepto_Banco': reg_b['Concepto'],
                        'Valor_Banco': reg_b['Valor'],
                        'Descripcion_Banco': reg_b['Descripción'],
                        'Fecha_Contable': reg_c['Fecha'] if i == 0 else None,
                        'Concepto_Contable': reg_c['Concepto'] if i == 0 else '',
                        'Valor_Contable': reg_c['Valor'] if i == 0 else None,
                        'Descripcion_Contable': reg_c['Descripción'] if i == 0 else '',
                        'Diferencia': abs(suma_banco - abs(reg_c['Valor'])) if i == 0 else None
                    })
                
                banco.loc[grupo_banco.index, 'Conciliado'] = True
                contable.loc[idx_c, 'Conciliado'] = True
                contador += 1
                id_conc += 1
                break
    
    print(f"✓ Impuestos conciliados: {contador}")
    return contador

# ============================================================================
# 🎯 ESTRATEGIA 6: SEGUNDA PASADA FLEXIBLE - OPTIMIZADA
# ============================================================================

def segunda_pasada_inteligente(banco, contable, conciliaciones):
    if not EJECUTAR_SEGUNDA_PASADA:
        return 0
    
    print("\n" + "="*70)
    print("🎯 ESTRATEGIA 6: SEGUNDA PASADA (Búsqueda Flexible)")
    print("="*70)
    
    contador = 0
    id_conc = len(conciliaciones) + 1 if conciliaciones else 1
    
    for idx_b, reg_b in banco[~banco['Conciliado']].iterrows():
        if banco.loc[idx_b, 'Conciliado']:
            continue
        
        f_min = reg_b['Fecha'] - timedelta(days=VENTANA_DIAS_FLEXIBLE)
        f_max = reg_b['Fecha'] + timedelta(days=VENTANA_DIAS_FLEXIBLE)
        
        cands = contable[
            (~contable['Conciliado']) &
            (contable['Fecha'] >= f_min) &
            (contable['Fecha'] <= f_max)
        ].copy()
        
        if len(cands) == 0:
            continue
        
        valor_objetivo = reg_b['Valor']
        mejor_combo, mejor_diff = None, float('inf')
        
        # ⚡ OPTIMIZACIÓN: Limitar partidas y combinaciones
        max_items = min(len(cands), MAX_PARTIDAS_AGRUPACION)
        cands_limitados = cands.head(max_items)
        
        total_combinaciones = 0
        for r in range(2, min(len(cands_limitados) + 1, 8)):
            if mejor_diff < TOLERANCIA_VALOR_AGRUPACION:
                break
            if total_combinaciones >= MAX_COMBINACIONES_POR_BUSQUEDA:
                break
            
            for combo in combinations(cands_limitados.index, r):
                total_combinaciones += 1
                if total_combinaciones > MAX_COMBINACIONES_POR_BUSQUEDA:
                    break
                
                suma = contable.loc[list(combo), 'Valor'].sum()
                diff = abs(suma - valor_objetivo)
                
                if diff < mejor_diff:
                    mejor_diff = diff
                    mejor_combo = combo
                
                if diff < TOLERANCIA_VALOR_AGRUPACION:
                    break
        
        if mejor_combo and mejor_diff < TOLERANCIA_VALOR_AGRUPACION:
            grupo = contable.loc[list(mejor_combo)]
            suma = grupo['Valor'].sum()
            
            sim_media = grupo['Texto_Busqueda'].apply(
                lambda x: calcular_similitud(reg_b['Texto_Busqueda'], x)
            ).mean()
            
            for i, (_, reg_c) in enumerate(grupo.sort_values(['Fecha', 'Valor']).iterrows()):
                conciliaciones.append({
                    'ID_Conciliacion': id_conc,
                    'Tipo': f'6. Segunda Pasada ({len(grupo)} reg)',
                    'Similitud': round(sim_media, 3),
                    'Fecha_Banco': reg_b['Fecha'] if i == 0 else None,
                    'Concepto_Banco': reg_b['Concepto'] if i == 0 else '',
                    'Valor_Banco': reg_b['Valor'] if i == 0 else None,
                    'Descripcion_Banco': reg_b['Descripción'] if i == 0 else '',
                    'Fecha_Contable': reg_c['Fecha'],
                    'Concepto_Contable': reg_c['Concepto'],
                    'Valor_Contable': reg_c['Valor'],
                    'Descripcion_Contable': reg_c['Descripción'],
                    'Diferencia': mejor_diff if i == 0 else None
                })
            
            banco.loc[idx_b, 'Conciliado'] = True
            contable.loc[list(mejor_combo), 'Conciliado'] = True
            contador += 1
            id_conc += 1
    
    print(f"✓ Segunda pasada: {contador}")
    return contador

# ============================================================================
# 🎯 ESTRATEGIA 7: BÚSQUEDA EXHAUSTIVA FINAL [v6.1] - OPTIMIZADA
# ============================================================================

def busqueda_exhaustiva_final(banco, contable, conciliaciones):
    if not EJECUTAR_BUSQUEDA_EXHAUSTIVA:
        return 0
    
    print("\n" + "="*70)
    print("🎯 ESTRATEGIA 7: BÚSQUEDA EXHAUSTIVA (Sin restricción fechas)")
    print("="*70)
    
    contador = 0
    id_conc = len(conciliaciones) + 1 if conciliaciones else 1
    
    banco_pend = banco[~banco['Conciliado']].copy()
    contable_pend = contable[~contable['Conciliado']].copy()
    
    n_banco_pend = len(banco_pend)
    n_contable_pend = len(contable_pend)
    
    print(f"  📊 Pendientes: Banco={n_banco_pend}, Contable={n_contable_pend}")
    
    if n_banco_pend > UMBRAL_PARTIDAS_EXHAUSTIVA or n_contable_pend > UMBRAL_PARTIDAS_EXHAUSTIVA:
        print(f"  ⊘ Demasiadas partidas (umbral={UMBRAL_PARTIDAS_EXHAUSTIVA})")
        return 0
    
    if n_banco_pend == 0 or n_contable_pend == 0:
        print("  ⊘ No hay partidas en ambos lados")
        return 0
    
    # CASO 1: Buscar N contable → 1 banco
    if n_contable_pend <= 10:
        print(f"\n  🔍 Buscando: Banco → Contable...")
        for idx_c, reg_c in contable_pend.iterrows():
            if contable.loc[idx_c, 'Conciliado']:
                continue
            
            valor_objetivo = reg_c['Valor']
            banco_disponible = banco[~banco['Conciliado']]
            
            if len(banco_disponible) == 0:
                continue
            
            mejor_combo, mejor_diff = None, float('inf')
            indices = banco_disponible.index.tolist()[:MAX_PARTIDAS_AGRUPACION]
            
            total_comb = 0
            for r in range(1, len(indices) + 1):
                if mejor_diff < TOLERANCIA_VALOR_AGRUPACION or total_comb > MAX_COMBINACIONES_EXHAUSTIVA:
                    break
                
                for combo in combinations(indices, r):
                    total_comb += 1
                    if total_comb > MAX_COMBINACIONES_EXHAUSTIVA:
                        break
                    
                    suma = banco.loc[list(combo), 'Valor'].sum()
                    diff = abs(suma - valor_objetivo)
                    
                    if diff < mejor_diff:
                        mejor_diff = diff
                        mejor_combo = combo
                    
                    if diff < TOLERANCIA_VALOR_AGRUPACION:
                        break
            
            if mejor_combo and mejor_diff < TOLERANCIA_VALOR_AGRUPACION:
                grupo = banco.loc[list(mejor_combo)]
                print(f"    ✓ {len(grupo)} banco → 1 contable (diff=${mejor_diff:.2f})")
                
                for i, (idx_b, reg_b) in enumerate(grupo.sort_values(['Fecha', 'Valor']).iterrows()):
                    conciliaciones.append({
                        'ID_Conciliacion': id_conc,
                        'Tipo': f'7. Exhaustiva ({len(grupo)}→1)',
                        'Similitud': 0.0,
                        'Fecha_Banco': reg_b['Fecha'],
                        'Concepto_Banco': reg_b['Concepto'],
                        'Valor_Banco': reg_b['Valor'],
                        'Descripcion_Banco': reg_b['Descripción'],
                        'Fecha_Contable': reg_c['Fecha'] if i == 0 else None,
                        'Concepto_Contable': reg_c['Concepto'] if i == 0 else '',
                        'Valor_Contable': reg_c['Valor'] if i == 0 else None,
                        'Descripcion_Contable': reg_c['Descripción'] if i == 0 else '',
                        'Diferencia': mejor_diff if i == 0 else None
                    })
                
                banco.loc[list(mejor_combo), 'Conciliado'] = True
                contable.loc[idx_c, 'Conciliado'] = True
                contador += 1
                id_conc += 1
    
    # CASO 2: Buscar N banco → 1 contable
    banco_pend = banco[~banco['Conciliado']].copy()
    if len(banco_pend) <= 10 and len(banco_pend) > 0:
        print(f"\n  🔍 Buscando: Contable → Banco...")
        for idx_b, reg_b in banco_pend.iterrows():
            if banco.loc[idx_b, 'Conciliado']:
                continue
            
            valor_objetivo = reg_b['Valor']
            contable_disponible = contable[~contable['Conciliado']]
            
            if len(contable_disponible) == 0:
                continue
            
            mejor_combo, mejor_diff = None, float('inf')
            indices = contable_disponible.index.tolist()[:MAX_PARTIDAS_AGRUPACION]
            
            total_comb = 0
            for r in range(1, len(indices) + 1):
                if mejor_diff < TOLERANCIA_VALOR_AGRUPACION or total_comb > MAX_COMBINACIONES_EXHAUSTIVA:
                    break
                
                for combo in combinations(indices, r):
                    total_comb += 1
                    if total_comb > MAX_COMBINACIONES_EXHAUSTIVA:
                        break
                    
                    suma = contable.loc[list(combo), 'Valor'].sum()
                    diff = abs(suma - valor_objetivo)
                    
                    if diff < mejor_diff:
                        mejor_diff = diff
                        mejor_combo = combo
                    
                    if diff < TOLERANCIA_VALOR_AGRUPACION:
                        break
            
            if mejor_combo and mejor_diff < TOLERANCIA_VALOR_AGRUPACION:
                grupo = contable.loc[list(mejor_combo)]
                print(f"    ✓ 1 banco → {len(grupo)} contable (diff=${mejor_diff:.2f})")
                
                for i, (_, reg_c) in enumerate(grupo.sort_values(['Fecha', 'Valor']).iterrows()):
                    conciliaciones.append({
                        'ID_Conciliacion': id_conc,
                        'Tipo': f'7. Exhaustiva (1→{len(grupo)})',
                        'Similitud': 0.0,
                        'Fecha_Banco': reg_b['Fecha'] if i == 0 else None,
                        'Concepto_Banco': reg_b['Concepto'] if i == 0 else '',
                        'Valor_Banco': reg_b['Valor'] if i == 0 else None,
                        'Descripcion_Banco': reg_b['Descripción'] if i == 0 else '',
                        'Fecha_Contable': reg_c['Fecha'],
                        'Concepto_Contable': reg_c['Concepto'],
                        'Valor_Contable': reg_c['Valor'],
                        'Descripcion_Contable': reg_c['Descripción'],
                        'Diferencia': mejor_diff if i == 0 else None
                    })
                
                banco.loc[idx_b, 'Conciliado'] = True
                contable.loc[list(mejor_combo), 'Conciliado'] = True
                contador += 1
                id_conc += 1
    
    print(f"\n✓ Búsqueda exhaustiva: {contador}")
    return contador

# ============================================================================
# 🔍 DETECCIÓN DE CASOS ESPECIALES
# ============================================================================

def detectar_casos_especiales(banco, contable):
    print("\n" + "="*70)
    print("🔍 DETECTANDO CASOS ESPECIALES")
    print("="*70)
    
    casos_especiales = []
    
    if not DETECTAR_CASOS_ESPECIALES:
        print("⊘ Detección desactivada")
        return casos_especiales
    
    banco_pend = banco[~banco['Conciliado']].copy()
    contable_pend = contable[~contable['Conciliado']].copy()
    
    for idx_b, reg_b in banco_pend.iterrows():
        for prov_id in contable_pend['Proveedor_ID'].dropna().unique():
            grupo = contable_pend[contable_pend['Proveedor_ID'] == prov_id]
            if len(grupo) < 1:
                continue
            
            suma = grupo['Valor'].sum()
            
            if reg_b['Valor'] != 0:
                porcentaje = abs(suma) / abs(reg_b['Valor']) * 100
            else:
                continue
            
            if 80 <= porcentaje < 100:
                diferencia = abs(reg_b['Valor']) - abs(suma)
                
                sim = calcular_similitud(
                    reg_b['Texto_Busqueda'],
                    ' '.join(grupo['Texto_Busqueda'].tolist())
                )
                
                if sim > 0.1:
                    casos_especiales.append({
                        'Tipo': 'PARCIAL',
                        'Descripcion': f'{len(grupo)} partidas suman {porcentaje:.1f}% del banco',
                        'Banco_Fecha': reg_b['Fecha'],
                        'Banco_Concepto': reg_b['Concepto'],
                        'Banco_Valor': reg_b['Valor'],
                        'Banco_Descripcion': reg_b['Descripción'],
                        'Contable_Registros': len(grupo),
                        'Contable_Suma': suma,
                        'Diferencia_Pendiente': diferencia,
                        'Porcentaje_Conciliado': round(porcentaje, 1),
                        'Similitud': round(sim, 3)
                    })
                    break
        
        mes_banco = reg_b['Fecha'].month if pd.notna(reg_b['Fecha']) else None
        
        if mes_banco:
            otros_meses = contable_pend[
                contable_pend['Fecha'].dt.month != mes_banco
            ]
            
            for _, reg_c in otros_meses.iterrows():
                if abs(reg_c['Valor'] - reg_b['Valor']) < TOLERANCIA_VALOR_EXACTA:
                    sim = calcular_similitud(reg_b['Texto_Busqueda'], reg_c['Texto_Busqueda'])
                    
                    if sim > 0.2:
                        casos_especiales.append({
                            'Tipo': 'PERIODO_DIFERENTE',
                            'Descripcion': f'Monto exacto pero mes diferente',
                            'Banco_Fecha': reg_b['Fecha'],
                            'Banco_Concepto': reg_b['Concepto'],
                            'Banco_Valor': reg_b['Valor'],
                            'Banco_Descripcion': reg_b['Descripción'],
                            'Contable_Registros': 1,
                            'Contable_Suma': reg_c['Valor'],
                            'Diferencia_Pendiente': 0,
                            'Porcentaje_Conciliado': 100.0,
                            'Similitud': round(sim, 3)
                        })
    
    casos_unicos = []
    valores_vistos = set()
    for caso in casos_especiales:
        key = (caso['Banco_Valor'], caso['Contable_Suma'])
        if key not in valores_vistos:
            valores_vistos.add(key)
            casos_unicos.append(caso)
    
    print(f"⚠️ Casos especiales detectados: {len(casos_unicos)}")
    return casos_unicos

# ============================================================================
# 🎨 FORMATO EXCEL PROFESIONAL
# ============================================================================

def aplicar_formato_profesional(ruta):
    """Aplica formato moderno y profesional al Excel"""
    wb = openpyxl.load_workbook(ruta)
    
    C_PRI = "1A237E"
    C_SEC = "3F51B5"
    C_EXI = "2E7D32"
    C_ERR = "C62828"
    C_ADV = "EF6C00"
    C_GRI = "FAFAFA"
    C_GRI2 = "EEEEEE"
    
    f_tit = Font(name='Segoe UI', size=22, bold=True, color="FFFFFF")
    f_sub = Font(name='Segoe UI', size=12, italic=True, color="757575")
    f_enc = Font(name='Segoe UI', size=11, bold=True, color="FFFFFF")
    f_sec = Font(name='Segoe UI', size=13, bold=True, color=C_SEC)
    f_nor = Font(name='Segoe UI', size=10)
    f_num = Font(name='Segoe UI', size=10, bold=True)
    
    r_pri = PatternFill(start_color=C_PRI, end_color=C_PRI, fill_type="solid")
    r_sec = PatternFill(start_color=C_SEC, end_color=C_SEC, fill_type="solid")
    r_exi = PatternFill(start_color=C_EXI, end_color=C_EXI, fill_type="solid")
    r_err = PatternFill(start_color=C_ERR, end_color=C_ERR, fill_type="solid")
    r_adv = PatternFill(start_color=C_ADV, end_color=C_ADV, fill_type="solid")
    r_gri = PatternFill(start_color=C_GRI, end_color=C_GRI, fill_type="solid")
    r_gri2 = PatternFill(start_color=C_GRI2, end_color=C_GRI2, fill_type="solid")
    
    tipos_r = {
        '1': PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid"),
        '2': PatternFill(start_color="BBDEFB", end_color="BBDEFB", fill_type="solid"),
        '3': PatternFill(start_color="FFE082", end_color="FFE082", fill_type="solid"),
        '4': PatternFill(start_color="FFCCBC", end_color="FFCCBC", fill_type="solid"),
        '5': PatternFill(start_color="E1BEE7", end_color="E1BEE7", fill_type="solid"),
        '6': PatternFill(start_color="B2DFDB", end_color="B2DFDB", fill_type="solid"),
        '7': PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid"),
    }
    
    borde = Border(
        left=Side(style='thin', color="BDBDBD"),
        right=Side(style='thin', color="BDBDBD"),
        top=Side(style='thin', color="BDBDBD"),
        bottom=Side(style='thin', color="BDBDBD")
    )
    
    if 'RESUMEN' in wb.sheetnames:
        ws = wb['RESUMEN']
        ws.insert_rows(1, 3)
        
        ws.merge_cells('A1:B1')
        ws['A1'].value = "📊 CONCILIACIÓN BANCARIA v6.5 DEFINITIVO"
        ws['A1'].font = f_tit
        ws['A1'].fill = r_pri
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 50
        
        ws.merge_cells('A2:B2')
        ws['A2'].value = "GREEN PARK - BANRESERBAS"
        ws['A2'].font = Font(name='Segoe UI', size=14, bold=True, color="424242")
        ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[2].height = 28
        
        ws.merge_cells('A3:B3')
        ws['A3'].value = f"📅 Generado: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
        ws['A3'].font = f_sub
        ws['A3'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[3].height = 24
        
        for r in range(4, ws.max_row + 1):
            for c in [1, 2]:
                cell = ws.cell(r, c)
                cell.border = borde
                cell.font = f_nor
                
                if c == 1 and cell.value and isinstance(cell.value, str):
                    if '═══' in str(cell.value):
                        cell.font = f_sec
                        cell.fill = r_gri2
                    elif '───' in str(cell.value):
                        cell.font = Font(name='Segoe UI', size=11, bold=True, color=C_SEC)
                        cell.fill = r_gri
                
                if c == 2 and isinstance(cell.value, (int, float)):
                    cell.font = f_num
                    cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                    if cell.value < 0:
                        cell.font = Font(name='Segoe UI', size=10, bold=True, color=C_ERR)
                    elif cell.value > 0:
                        cell.font = Font(name='Segoe UI', size=10, bold=True, color=C_EXI)
        
        ws.column_dimensions['A'].width = 62
        ws.column_dimensions['B'].width = 35
    
    if 'CONCILIADOS' in wb.sheetnames:
        ws = wb['CONCILIADOS']
        if ws.max_row > 1:
            ws.insert_rows(1)
            max_c = ws.max_column
            
            ws.merge_cells(f'A1:{get_column_letter(max_c)}1')
            ws['A1'].value = "✅ PARTIDAS CONCILIADAS"
            ws['A1'].font = f_tit
            ws['A1'].fill = r_exi
            ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
            ws.row_dimensions[1].height = 45
            
            for c in range(1, max_c + 1):
                cell = ws.cell(2, c)
                cell.fill = r_sec
                cell.font = f_enc
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = borde
            ws.row_dimensions[2].height = 38
            
            for r in range(3, ws.max_row + 1):
                tipo_val = str(ws.cell(r, 2).value) if ws.cell(r, 2).value else ""
                relleno = None
                
                for t_num, t_fill in tipos_r.items():
                    if tipo_val.startswith(f'{t_num}.'):
                        relleno = t_fill
                        break
                
                if not relleno and r % 2 == 0:
                    relleno = r_gri
                
                for c in range(1, max_c + 1):
                    cell = ws.cell(r, c)
                    if relleno:
                        cell.fill = relleno
                    cell.border = borde
                    cell.font = f_nor
                    cell.alignment = Alignment(vertical='center', wrap_text=True)
                    
                    hdr = ws.cell(2, c)
                    if hdr.value:
                        if 'Fecha' in str(hdr.value) and isinstance(cell.value, datetime):
                            cell.number_format = 'DD/MM/YYYY'
                        elif ('Valor' in str(hdr.value) or 'Diferencia' in str(hdr.value)):
                            if isinstance(cell.value, (int, float)):
                                cell.number_format = '#,##0.00'
                                color = C_ERR if cell.value < 0 else C_EXI
                                cell.font = Font(name='Segoe UI', size=10, bold=True, color=color)
            
            for c in range(1, max_c + 1):
                col_letter = get_column_letter(c)
                max_len = 0
                for r in range(1, min(ws.max_row + 1, 100)):
                    val = ws.cell(r, c).value
                    if val:
                        max_len = max(max_len, len(str(val)))
                ws.column_dimensions[col_letter].width = min(max_len + 4, 55)
    
    if 'NO_CONCILIADOS' in wb.sheetnames:
        ws = wb['NO_CONCILIADOS']
        if ws.max_row > 1:
            ws.insert_rows(1)
            max_c = ws.max_column
            
            ws.merge_cells(f'A1:{get_column_letter(max_c)}1')
            ws['A1'].value = "⚠️ PARTIDAS PENDIENTES DE CONCILIACIÓN"
            ws['A1'].font = f_tit
            ws['A1'].fill = r_err
            ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
            ws.row_dimensions[1].height = 45
            
            for c in range(1, max_c + 1):
                cell = ws.cell(2, c)
                cell.font = f_enc
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = borde
                if cell.value and 'Banco' in str(cell.value):
                    cell.fill = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
                elif cell.value and 'Contable' in str(cell.value):
                    cell.fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
                else:
                    cell.fill = r_sec
            ws.row_dimensions[2].height = 38
            
            for r in range(3, ws.max_row + 1):
                relleno_fila = r_gri if r % 2 == 0 else None
                for c in range(1, max_c + 1):
                    cell = ws.cell(r, c)
                    if relleno_fila:
                        cell.fill = relleno_fila
                    cell.border = borde
                    cell.font = f_nor
                    cell.alignment = Alignment(vertical='center', wrap_text=True)
                    
                    hdr = ws.cell(2, c)
                    if hdr.value:
                        if 'Fecha' in str(hdr.value) and isinstance(cell.value, datetime):
                            cell.number_format = 'DD/MM/YYYY'
                        elif 'Valor' in str(hdr.value):
                            if isinstance(cell.value, (int, float)):
                                cell.number_format = '#,##0.00'
                                color = C_ERR if cell.value < 0 else C_EXI
                                cell.font = Font(name='Segoe UI', size=10, bold=True, color=color)
            
            for c in range(1, max_c + 1):
                col_letter = get_column_letter(c)
                max_len = 0
                for r in range(1, min(ws.max_row + 1, 100)):
                    val = ws.cell(r, c).value
                    if val:
                        max_len = max(max_len, len(str(val)))
                ws.column_dimensions[col_letter].width = min(max_len + 4, 55)
    
    if 'CASOS_ESPECIALES' in wb.sheetnames:
        ws = wb['CASOS_ESPECIALES']
        if ws.max_row > 1:
            ws.insert_rows(1)
            max_c = ws.max_column
            
            ws.merge_cells(f'A1:{get_column_letter(max_c)}1')
            ws['A1'].value = "🔍 CASOS ESPECIALES - Requieren Revisión Manual"
            ws['A1'].font = f_tit
            ws['A1'].fill = r_adv
            ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
            ws.row_dimensions[1].height = 45
            
            for c in range(1, max_c + 1):
                cell = ws.cell(2, c)
                cell.fill = r_sec
                cell.font = f_enc
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = borde
            ws.row_dimensions[2].height = 38
            
            for r in range(3, ws.max_row + 1):
                for c in range(1, max_c + 1):
                    cell = ws.cell(r, c)
                    cell.border = borde
                    cell.font = f_nor
                    cell.fill = PatternFill(start_color="FFF8E1", end_color="FFF8E1", fill_type="solid")
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = '#,##0.00'
                        cell.font = f_num
            
            for c in range(1, max_c + 1):
                col_letter = get_column_letter(c)
                max_len = 0
                for r in range(1, min(ws.max_row + 1, 100)):
                    val = ws.cell(r, c).value
                    if val:
                        max_len = max(max_len, len(str(val)))
                ws.column_dimensions[col_letter].width = min(max_len + 4, 55)
    
    wb.save(ruta)
    print("✨ Formato profesional aplicado exitosamente")

# ============================================================================
# 🚀 FUNCIÓN PRINCIPAL
# ============================================================================

def main():

    """Función principal de conciliación"""
    import time
    tiempo_inicio = time.time()
    
    print("\n" + "="*70)
    print("  🏦 CONCILIACIÓN BANCARIA v6.5 DEFINITIVO")
    print("  🔥 Motor: v6.0 + v6.1 + Optimizaciones")
    print("="*70)
    print(f"\n📅 {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")
    print(f"📁 Proyecto: GREEN PARK - BANRESERBAS")
    
    print("\n⚙️ PARÁMETROS:")
    print("─"*70)
    print(f"  💰 Tolerancia exacta: ${TOLERANCIA_VALOR_EXACTA}")
    print(f"  💰 Tolerancia agrupación: ${TOLERANCIA_VALOR_AGRUPACION}")
    print(f"  📅 Ventanas: ±{VENTANA_DIAS_EXACTA}d / ±{VENTANA_DIAS_AGRUPACION}d / ±{VENTANA_DIAS_FLEXIBLE}d")
    print(f"  ⚡ Límites: {MAX_PARTIDAS_AGRUPACION} partidas / {MAX_COMBINACIONES_POR_BUSQUEDA:,} comb")
    print(f"  🔧 Segunda pasada: {'Sí' if EJECUTAR_SEGUNDA_PASADA else 'No'}")
    print(f"  🔧 Búsqueda exhaustiva: {'Sí' if EJECUTAR_BUSQUEDA_EXHAUSTIVA else 'No'}")
    
    print("\n🏦 BANCOS:")
    print("─"*70)
    print(f"  1) BanReservas")
    print(f"  2) Popular")
    print(f"  3) BHD")
    print(f"  4) Santa Cruz")

    print(f" 🔧 Ingrese el numero para su selección")

    valid = False
    while valid == False:
        tipoBanco = input("\nIngrese aqui el banco que esta utilizando: ")
        if tipoBanco.isdigit() and int(tipoBanco) in [1, 2, 3, 4]:
            tipoBanco = int(tipoBanco)
            valid = True
        else:
            print(f"❌ Entrada inválida.")

    print("\n" + "="*70)
    print("📂 CARGANDO DATOS")
    print("="*70)
    
    # Carga de Banco
    if tipoBanco == 1:
        try:
            ARCHIVO_BANCO = r"BANRESERVAS.xlsx"
            banco = cargar_banreservas(ARCHIVO_BANCO, "BANCO")
        except Exception as e:
            print(f"❌ Error al cargar banco: {e}")
            return
        
    if tipoBanco == 2:
        try:
            ARCHIVO_BANCO = r"POPULAR.xlsx"
            banco = cargar_popular(ARCHIVO_BANCO, "BANCO")
        except Exception as e:
            print(f"❌ Error al cargar banco: {e}")
            return
        
    if tipoBanco == 3:
        try:
            ARCHIVO_BANCO = r"BHD.xlsx"
            banco = cargar_bhd(ARCHIVO_BANCO, "BANCO")
        except Exception as e:
            print(f"❌ Error al cargar banco: {e}")
            return

    if tipoBanco == 4:
        try:
            ARCHIVO_BANCO = r"SANTACRUZ.xlsx"
            banco = cargar_santaCruz(ARCHIVO_BANCO, "BANCO")
        except Exception as e:
            print(f"❌ Error al cargar banco: {e}")
            return
        
    # Carga de Contable
    try:
        ARCHIVO_CONTABLE = r"LIBRO_CONTABLE.xlsx"
        contable = cargar_contable(ARCHIVO_CONTABLE, "CONTABLE")
    except Exception as e:
        print(f"❌ Error al cargar contable: {e}")
        return
    
    print("\n" + "="*70)
    print("🔄 EJECUTANDO MOTOR v6.5")
    print("="*70)
    
    conciliaciones = []
    
    t1 = conciliacion_por_monto_exacto(banco, contable, conciliaciones)
    t1_5 = conciliacion_con_comisiones(banco, contable, conciliaciones)
    t1_6 = conciliacion_comisiones_agrupadas(banco, contable, conciliaciones)
    t2 = conciliacion_n_a_1_inteligente(banco, contable, conciliaciones)
    t3 = conciliacion_1_a_n_inteligente(banco, contable, conciliaciones)
    t4 = conciliacion_n_a_m_inteligente(banco, contable, conciliaciones)
    t5 = conciliacion_impuestos(banco, contable, conciliaciones)
    t6 = segunda_pasada_inteligente(banco, contable, conciliaciones)
    t7 = busqueda_exhaustiva_final(banco, contable, conciliaciones)
    
    casos_especiales = detectar_casos_especiales(banco, contable)
    
    if conciliaciones:
        df_conc = pd.DataFrame(conciliaciones).sort_values(['Tipo', 'ID_Conciliacion'])
    else:
        df_conc = pd.DataFrame()
    
    if casos_especiales:
        df_especiales = pd.DataFrame(casos_especiales)
    else:
        df_especiales = pd.DataFrame()
    
    no_b = banco[~banco['Conciliado']][['Fecha', 'Concepto', 'Valor', 'Descripción']].reset_index(drop=True)
    no_c = contable[~contable['Conciliado']][['Fecha', 'Concepto', 'Valor', 'Descripción']].reset_index(drop=True)
    
    max_r = max(len(no_b), len(no_c)) if len(no_b) > 0 or len(no_c) > 0 else 1
    
    df_no_conc = pd.DataFrame({
        'Fecha_Banco': list(no_b['Fecha']) + [None] * (max_r - len(no_b)),
        'Concepto_Banco': list(no_b['Concepto']) + [''] * (max_r - len(no_b)),
        'Valor_Banco': list(no_b['Valor']) + [None] * (max_r - len(no_b)),
        'Descripcion_Banco': list(no_b['Descripción']) + [''] * (max_r - len(no_b)),
        'Fecha_Contable': list(no_c['Fecha']) + [None] * (max_r - len(no_c)),
        'Concepto_Contable': list(no_c['Concepto']) + [''] * (max_r - len(no_c)),
        'Valor_Contable': list(no_c['Valor']) + [None] * (max_r - len(no_c)),
        'Descripcion_Contable': list(no_c['Descripción']) + [''] * (max_r - len(no_c))
    })
    
    total_grupos = len(df_conc['ID_Conciliacion'].unique()) if len(df_conc) > 0 else 0
    total_estrategias = t1 + t1_5 + t1_6 + t2 + t3 + t4 + t5 + t6 + t7
    
    tiempo_total = time.time() - tiempo_inicio
    
    resumen = pd.DataFrame({
        'Concepto': [
            '══════════════════════════════════════════════════════════════',
            'CONCILIACIÓN BANCARIA v6.5 DEFINITIVO',
            '══════════════════════════════════════════════════════════════',
            '',
            '─── INFORMACIÓN GENERAL ───',
            'Total registros Banco',
            'Total registros Contable',
            'Suma total Banco',
            'Suma total Contable',
            '',
            '─── CONCILIACIONES POR ESTRATEGIA ───',
            '1️⃣  Monto Exacto (1:1)',
            '1.5️⃣ Transferencias + Comisión',
            '1.6️⃣ Comisiones Agrupadas [v6.1]',
            '2️⃣  Agrupaciones N→1',
            '3️⃣  Agrupaciones 1→N',
            '4️⃣  Agrupaciones N↔M',
            '5️⃣  Impuestos DGII',
            '6️⃣  Segunda Pasada',
            '7️⃣  Búsqueda Exhaustiva [v6.1]',
            '📦  TOTAL GRUPOS CONCILIADOS',
            '',
            '─── REGISTROS PROCESADOS ───',
            'Registros Banco conciliados',
            'Registros Contable conciliados',
            'Registros Banco pendientes',
            'Registros Contable pendientes',
            '',
            '─── MONTOS ───',
            'Suma conciliados Banco',
            'Suma conciliados Contable',
            'Suma pendientes Banco',
            'Suma pendientes Contable',
            '',
            '─── INDICADORES ───',
            '📊 % Conciliación Banco',
            '📊 % Conciliación Contable',
            '💰 DIFERENCIA MONETARIA TOTAL',
            '⚠️  Casos especiales detectados',
            '⏱️  Tiempo de ejecución (seg)',
        ],
        'Valor': [
            '', '', '', '',
            '',
            len(banco),
            len(contable),
            round(banco['Valor'].sum(), 2),
            round(contable['Valor'].sum(), 2),
            '',
            '',
            t1, t1_5, t1_6, t2, t3, t4, t5, t6, t7,
            total_grupos,
            '',
            '',
            banco['Conciliado'].sum(),
            contable['Conciliado'].sum(),
            len(no_b),
            len(no_c),
            '',
            '',
            round(banco[banco['Conciliado']]['Valor'].sum(), 2),
            round(contable[contable['Conciliado']]['Valor'].sum(), 2),
            round(no_b['Valor'].sum(), 2) if len(no_b) > 0 else 0,
            round(no_c['Valor'].sum(), 2) if len(no_c) > 0 else 0,
            '',
            '',
            round(banco['Conciliado'].sum() / len(banco) * 100, 2),
            round(contable['Conciliado'].sum() / len(contable) * 100, 2),
            round(banco['Valor'].sum() - contable['Valor'].sum(), 2),
            len(casos_especiales),
            round(tiempo_total, 1),
        ]
    })
    
    print("\n" + "="*70)
    print("💾 EXPORTANDO RESULTADOS")
    print("="*70)
    
    with pd.ExcelWriter(RUTA_SALIDA, engine='openpyxl') as writer:
        resumen.to_excel(writer, sheet_name='RESUMEN', index=False)
        
        if len(df_conc) > 0:
            df_conc.to_excel(writer, sheet_name='CONCILIADOS', index=False)
        else:
            pd.DataFrame({'Mensaje': ['No hay conciliaciones']}).to_excel(
                writer, sheet_name='CONCILIADOS', index=False
            )
        
        df_no_conc.to_excel(writer, sheet_name='NO_CONCILIADOS', index=False)
        
        if len(df_especiales) > 0:
            df_especiales.to_excel(writer, sheet_name='CASOS_ESPECIALES', index=False)
    
    print("✅ Datos exportados correctamente")
    
    if APLICAR_FORMATO_PROFESIONAL:
        print("🎨 Aplicando formato profesional...")
        try:
            aplicar_formato_profesional(RUTA_SALIDA)
        except Exception as e:
            print(f"⚠️ Error al aplicar formato: {e}")
    
    perc_b = banco['Conciliado'].sum() / len(banco) * 100
    perc_c = contable['Conciliado'].sum() / len(contable) * 100
    
    print("\n" + "="*70)
    print("       ✅ CONCILIACIÓN COMPLETADA EXITOSAMENTE")
    print("="*70)
    print(f"\n📁 Archivo generado: {RUTA_SALIDA}\n")
    
    print("📊 RESUMEN EJECUTIVO:")
    print("─"*70)
    print(f"📥 Datos Originales:")
    print(f"   • Banco:     {len(banco):,} registros (${banco['Valor'].sum():,.2f})")
    print(f"   • Contable:  {len(contable):,} registros (${contable['Valor'].sum():,.2f})")
    
    print(f"\n✅ Conciliaciones por Estrategia:")
    print(f"   1️⃣  Monto Exacto:             {t1:,}")
    print(f"   1.5️⃣ Transf + Comisión:        {t1_5:,}")
    print(f"   1.6️⃣ Comisiones Agrupadas:     {t1_6:,}")
    print(f"   2️⃣  N→1:                      {t2:,}")
    print(f"   3️⃣  1→N:                      {t3:,}")
    print(f"   4️⃣  N↔M:                      {t4:,}")
    print(f"   5️⃣  Impuestos DGII:           {t5:,}")
    print(f"   6️⃣  Segunda Pasada:           {t6:,}")
    print(f"   7️⃣  Búsqueda Exhaustiva:      {t7:,}")
    print(f"   {'─'*32}")
    print(f"   📦 TOTAL GRUPOS:              {total_grupos:,}")
    
    print(f"\n📈 Resultados:")
    print(f"   • Banco:     {banco['Conciliado'].sum():,} conciliados ({perc_b:.1f}%)")
    print(f"   • Contable:  {contable['Conciliado'].sum():,} conciliados ({perc_c:.1f}%)")
    print(f"   • Pendientes: Banco {len(no_b):,} | Contable {len(no_c):,}")
    
    diferencia = banco['Valor'].sum() - contable['Valor'].sum()
    print(f"\n💰 DIFERENCIA TOTAL: ${diferencia:,.2f}")
    
    if len(casos_especiales) > 0:
        print(f"\n⚠️ Casos Especiales: {len(casos_especiales)}")
        print(f"   → Revisa la hoja 'CASOS_ESPECIALES' para más detalles")
    
    print(f"\n⏱️ Tiempo de Ejecución: {tiempo_total:.1f} segundos")
    
    print(f"\n🎯 Evaluación:")
    if perc_b >= 99 and perc_c >= 99:
        print("   ⭐⭐⭐⭐⭐ PERFECTA (≥99%)")
    elif perc_b >= 95 and perc_c >= 95:
        print("   ⭐⭐⭐⭐ EXCELENTE (≥95%)")
    elif perc_b >= 85 and perc_c >= 85:
        print("   ⭐⭐⭐ MUY BUENA (≥85%)")
    elif perc_b >= 75 and perc_c >= 75:
        print("   ⭐⭐ BUENA (≥75%)")
    elif perc_b >= 60 and perc_c >= 60:
        print("   ⭐ REGULAR (≥60%)")
    else:
        print("   ⚠️ NECESITA REVISIÓN (<60%)")
    
    print("\n" + "="*70)
    print("💡 TIPS:")
    print("─"*70)
    print("  1. Revisa RESUMEN para estadísticas generales")
    print("  2. Valida CONCILIADOS para verificar agrupaciones")
    print("  3. Analiza NO_CONCILIADOS para identificar diferencias")
    if len(casos_especiales) > 0:
        print("  4. Revisa CASOS_ESPECIALES para coincidencias parciales")
    print("="*70 + "\n")

if __name__ == "__main__":
    main()