"""
============================================================================
CONCILIACIÓN BANCARIA INTELIGENTE v0.9.6
============================================================================
Proyecto: GREEN PARK - BANRESERVAS

[OK] TODAS LAS ESTRATEGIAS (7 estrategias completas):
   1.   Monto Exacto (1:1)
   1.5  Transferencias + Comisión ($7)
   1.6  Comisiones Agrupadas Multi-fecha
   2.   Agrupaciones N→1 (4 métodos)
   3.   Agrupaciones 1→N
   4.   Agrupaciones N↔M
   5.   Impuestos DGII (0.15%)
   6.   Segunda Pasada Flexible
   7.   Búsqueda Exhaustiva Final

[OK] OPTIMIZACIONES:
   - Lectura rápida: solo filas con Fecha Y Valor válidos
   - Límites por estrategia para evitar timeouts
   - Alias TC/LEGAL mejorados: TC LEGAL ↔ TC Corporativa ↔ Legalizaciones ↔ IPI
   - Descripción vacía funcional (crea columna si no existe)
   - Score combinado para desempate (Concepto_Norm del v5 restaurado)

[TARGET] OBJETIVO: < 1 minuto de ejecución con máxima conciliación
============================================================================
"""

import pandas as pd
import csv
import numpy as np
from datetime import datetime, timedelta
import os
import glob
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import re
from difflib import SequenceMatcher
import unicodedata
from itertools import combinations
import warnings
warnings.filterwarnings('ignore')
import sys
import argparse
from pathlib import Path

def base_path():
    if getattr(sys, 'frozen', False):
        return Path(sys.executable).parent
    return Path(__file__).parent

# ============================================================================
# [NEW] NUEVO: GESTIÓN DE FIDEICOMISOS
# ============================================================================

def natural_sort_key(text):
    """
    Genera una clave para ordenamiento natural (números ordenados numéricamente)
    Ejemplo: '1', '2', '10', '11' en lugar de '1', '10', '11', '2'
    """
    parts = []
    for part in re.split(r'(\d+)', str(text)):
        if part.isdigit():
            parts.append(int(part))
        else:
            parts.append(part.lower())
    return parts

def detectar_fideicomisos():
    """
    Detecta todos los fideicomisos (carpetas) en el directorio actual
    Retorna: lista de nombres de fideicomisos encontrados
    """
    carpeta_script = base_path()
    fideicomisos = []
    
    # Buscar todas las carpetas que contengan la estructura esperada
    for item in carpeta_script.iterdir():
        if item.is_dir() and item.name not in ['__pycache__', '.git', '.vscode']:
            # Verificar si tiene las subcarpetas requeridas
            archivos_banco = item / 'Archivos Banco'
            archivos_contable = item / 'Archivos Libro Contable'
            
            # Si tiene al menos una de las carpetas esperadas, es un fideicomiso
            if archivos_banco.exists() or archivos_contable.exists():
                fideicomisos.append(item.name)
    
    # Ordenar usando ordenamiento natural (números se ordenan numéricamente)
    return sorted(fideicomisos, key=natural_sort_key)

def crear_estructura_fideicomiso(nombre_fideicomiso):
    """
    Crea la estructura de carpetas para un nuevo fideicomiso
    
    Estructura creada:
    nombre_fideicomiso/
    ├── Archivos Banco/
    ├── Archivos Libro Contable/
    └── Resultados/
    """
    carpeta_script = base_path()
    carpeta_fideicomiso = carpeta_script / nombre_fideicomiso
    
    # Verificar si ya existe
    if carpeta_fideicomiso.exists():
        print(f"[WARN]  El fideicomiso '{nombre_fideicomiso}' ya existe")
        return None
    
    # Crear carpeta principal del fideicomiso
    carpeta_fideicomiso.mkdir(exist_ok=True)
    
    # Crear subcarpetas
    (carpeta_fideicomiso / 'Archivos Banco').mkdir(exist_ok=True)
    (carpeta_fideicomiso / 'Archivos Libro Contable').mkdir(exist_ok=True)
    (carpeta_fideicomiso / 'Resultados').mkdir(exist_ok=True)

    print(f"\n[OK] Fideicomiso '{nombre_fideicomiso}' creado exitosamente")
    print(f"📁 Ubicación: {carpeta_fideicomiso}")
    print(f"\n[LIST] Estructura creada:")
    print(f"   ├── Archivos Banco/")
    print(f"   ├── Archivos Libro Contable/")
    print(f"   └── Resultados/")
    print(f"\n[INFO] Ahora coloca los archivos en las carpetas correspondientes")
    
    return str(carpeta_fideicomiso)

def seleccionar_fideicomiso():
    """
    Permite al usuario seleccionar un fideicomiso existente o crear uno nuevo
    Retorna: ruta del fideicomiso seleccionado/creado
    """
    print("\n" + "="*70)
    print("[DIR] GESTIÓN DE FIDEICOMISOS")
    print("="*70)
    
    fideicomisos = detectar_fideicomisos()
    
    if fideicomisos:
        print(f"\n[LIST] Fideicomisos encontrados: {len(fideicomisos)}\n")
        for i, fideicomiso in enumerate(fideicomisos, 1):
            # Mostrar info adicional del fideicomiso
            carpeta_fideicomiso = base_path() / fideicomiso
            archivos_banco = list((carpeta_fideicomiso / 'Archivos Banco').glob('*.*')) if (carpeta_fideicomiso / 'Archivos Banco').exists() else []
            archivos_contable = list((carpeta_fideicomiso / 'Archivos Libro Contable').glob('*.*')) if (carpeta_fideicomiso / 'Archivos Libro Contable').exists() else []
            
            n_banco = len([f for f in archivos_banco if f.suffix in ['.xlsx', '.xls', '.csv']])
            n_contable = len([f for f in archivos_contable if f.suffix in ['.xlsx', '.xls', '.csv']])
            
            print(f"   {i}) {fideicomiso}")
            print(f"      [STATS] Banco: {n_banco} archivo(s) | Contable: {n_contable} archivo(s) \n")
    else:
        print("\n[WARN]  No se encontraron fideicomisos existentes")
    
    print(f"\n   {'0' if fideicomisos else '1'}) [PLUS] Crear nuevo fideicomiso")
    
    while True:
        if fideicomisos:
            entrada = input(f"\nSelecciona una opción (0-{len(fideicomisos)}): ").strip()
        else:
            entrada = input(f"\nSelecciona una opción (1 para crear nuevo): ").strip()
        
        # Opción: Crear nuevo fideicomiso
        if entrada == '0' or (not fideicomisos and entrada == '1'):
            print("\n" + "─"*70)
            print("[PLUS] CREAR NUEVO FIDEICOMISO")
            print("─"*70)
            
            while True:
                nombre_caso = input("\n[NOTE] Nombre del nuevo fideicomiso: ").strip()
                
                if not nombre_caso:
                    print("[ERROR] El nombre no puede estar vacío")
                    continue
                
                # Validar nombre (sin caracteres especiales peligrosos)
                if not re.match(r'^[\w\s\-\.]+$', nombre_caso):
                    print("[ERROR] El nombre contiene caracteres inválidos")
                    print("   Usa solo letras, números, espacios, guiones y puntos")
                    continue
                
                # Confirmar creación
                confirmar = input(f"\n¿Crear fideicomiso '{nombre_caso}'? (s/n): ").strip().lower()
                if confirmar == 's':
                    ruta_caso = crear_estructura_fideicomiso(nombre_caso)
                    if ruta_caso:
                        return ruta_caso
                    else:
                        # Si ya existe, volver al menú
                        break
                else:
                    print("[ERROR] Creación cancelada")
                    break

        # Opción: Seleccionar fideicomiso existente
        elif fideicomisos and entrada.isdigit() and 1 <= int(entrada) <= len(fideicomisos):
            caso_seleccionado = fideicomisos[int(entrada) - 1]
            carpeta_caso = base_path() / caso_seleccionado
            
            print(f"\n[OK] Fideicomiso seleccionado: {caso_seleccionado}")
            print(f"📁 Ubicación: {carpeta_caso}")
            
            return str(carpeta_caso)
        else:
            print("[ERROR] Selección inválida")

def configurar_rutas_caso(carpeta_caso):
    """
    Configura las rutas globales para el caso seleccionado
    """
    global CARPETA_TRABAJO, CARPETA_BANCOS, CARPETA_CONTABLE, CARPETA_RESULTADOS
    
    CARPETA_TRABAJO = Path(carpeta_caso)
    CARPETA_BANCOS = CARPETA_TRABAJO / 'Archivos Banco'
    CARPETA_CONTABLE = CARPETA_TRABAJO / 'Archivos Libro Contable'
    CARPETA_RESULTADOS = CARPETA_TRABAJO / 'Resultados'
    
    # Crear carpetas si no existen
    CARPETA_BANCOS.mkdir(exist_ok=True)
    CARPETA_CONTABLE.mkdir(exist_ok=True)
    CARPETA_RESULTADOS.mkdir(exist_ok=True)
    
    print(f"\n[DIR] Rutas configuradas:")
    print(f"   [BANK] Bancos:     {CARPETA_BANCOS}")
    print(f"   [BOOK] Contable:   {CARPETA_CONTABLE}")
    print(f"   [SAVE] Resultados: {CARPETA_RESULTADOS}")

# ============================================================================
# [CONFIG] PARÁMETROS AJUSTABLES
# ============================================================================

# [MONEY] TOLERANCIAS DE VALOR
TOLERANCIA_VALOR_EXACTA = 0.01
TOLERANCIA_VALOR_AGRUPACION = 1.00
TOLERANCIA_PORCENTAJE_PARCIAL = 0.02

# [DATE] VENTANAS DE TIEMPO
VENTANA_DIAS_EXACTA = 10
VENTANA_DIAS_AGRUPACION = 20
VENTANA_DIAS_FLEXIBLE = 30
VENTANA_DIAS_COMISIONES = 45

# [TARGET] UMBRALES DE SIMILITUD
UMBRAL_SIMILITUD_BAJA = 0.05
UMBRAL_SIMILITUD_MEDIA = 0.20
UMBRAL_SIMILITUD_ALTA = 0.40

# [TOOL] CONFIGURACIÓN AVANZADA
PERMITIR_SOLO_MONTO = True
USAR_FECHAS_PARA_DESAMBIGUAR = True
DETECTAR_CASOS_ESPECIALES = True
APLICAR_FORMATO_PROFESIONAL = True
EJECUTAR_SEGUNDA_PASADA = True
EJECUTAR_BUSQUEDA_EXHAUSTIVA = True

# [USD] COMISIONES BANCARIAS
COMISION_TRANSFERENCIA_USD = 7.00
DETECTAR_COMISIONES = True

# [PERF] LÍMITES DE RENDIMIENTO (OPTIMIZACIÓN)
MAX_PARTIDAS_AGRUPACION = 30        # Reducido de 100 a 15 para velocidad
MAX_COMBINACIONES_POR_BUSQUEDA = 10000  # Límite por cada búsqueda individual
UMBRAL_PARTIDAS_EXHAUSTIVA = 25     # Aumentado de 20 a 25
MAX_COMBINACIONES_EXHAUSTIVA = 100000

# ============================================================================
# [CONFIG] CONFIGURACIÓN DE RUTAS Y ARCHIVOS RESULTADO
# ============================================================================

import calendar
def generar_nombre_conciliacion(codigo_banco, df_banco):
    """
    Genera nombre tipo:
    Conciliacion_POPULAR_Marzo_2024.xlsx
    """
    # Tomar la primera fecha válida del banco
    fecha_ref = df_banco['Fecha'].dropna().iloc[0]

    mes_nombre = calendar.month_name[fecha_ref.month]
    anio = fecha_ref.year

    # Capitalizar mes (March -> Marzo si luego quieres traducir)
    mes_nombre = mes_nombre.capitalize()

    return f"Conciliacion_{codigo_banco}_{mes_nombre}_{anio}.xlsx"

def obtener_ruta_unica(ruta_base):
    """
    Genera un nombre de archivo único.
    Si el archivo existe, agrega (1), (2), etc. como en descargas estándar.
    Ejemplo: Conciliacion_Resultados(1).xlsx, Conciliacion_Resultados(2).xlsx
    """
    if not Path(ruta_base).exists():
        return ruta_base
    
    # Dividir ruta base en directorio, nombre y extensión
    p = Path(ruta_base)
    directorio = p.parent
    nombre_sin_ext = p.stem
    extension = p.suffix
    
    contador = 1
    while True:
        nueva_ruta = directorio / f"{nombre_sin_ext}({contador}){extension}"
        if not nueva_ruta.exists():
            return str(nueva_ruta)
        contador += 1

# ============================================================================
# [BANK] PASO 1: SISTEMA DE RECONOCIMIENTO AUTOMÁTICO DE BANCOS
# ============================================================================

# Lista de bancos dominicanos soportados
BANCOS_SOPORTADOS = {
    'POPULAR': ['POPULAR', 'BANCO POPULAR', 'BPD', 'BP'],
    'BANRESERVAS': ['BANRESERVAS', 'BANCO DE RESERVAS', 'RESERVAS', 'BDR'],
    'BHD': ['BHD', 'BANCO BHD', 'BHD LEON', 'BHDLEON'],
    'BANESCO': ['BANESCO', 'BANCO BANESCO'],
    'SCOTIABANK': ['SCOTIABANK', 'SCOTIA', 'BANCO SCOTIA'],
    'PROMERICA': ['PROMERICA', 'BANCO PROMERICA'],
    'SANTA_CRUZ': ['SANTA CRUZ', 'SANTACRUZ', 'BANCO SANTA CRUZ', 'BSC'],
    'ADEMI': ['ADEMI', 'BANCO ADEMI'],
    'LOPEZ_DE_HARO': ['LOPEZ DE HARO', 'LOPEZDEHARO', 'BANCO LOPEZ'],
    'BELLBANK': ['BELLBANK', 'BELL BANK'],
    'APAP': ['APAP', 'BANCO APAP', 'ASOCIACION POPULAR DE AHORROS Y PRESTAMOS', 'ASOCIACION POPULAR']
}

def normalizar_nombre_banco(texto):
    """Normaliza el nombre para búsqueda (sin acentos, mayúsculas, sin espacios ni símbolos)"""
    if not texto:
        return ""
    texto = quitar_acentos(str(texto)).upper()
    texto = re.sub(r'[^A-Z0-9]', '', texto)  # Solo letras y números
    return texto

def detectar_banco_en_nombre_archivo(nombre_archivo):
    """
    Detecta qué banco es basándose en el nombre del archivo
    Retorna: (codigo_banco, nombre_legible) o (None, None)
    """
    nombre_norm = normalizar_nombre_banco(nombre_archivo)
    
    for codigo_banco, variantes in BANCOS_SOPORTADOS.items():
        for variante in variantes:
            variante_norm = normalizar_nombre_banco(variante)
            if variante_norm in nombre_norm:
                return codigo_banco, variante
    
    return None, None

def buscar_archivos_en_carpeta():
    """
    Busca archivos de banco en 'Archivos Banco'
    y libro contable en 'Archivos Libro Contable'
    """

    print("\n" + "="*70)
    print("[SEARCH] PASO 1: BUSCANDO ARCHIVOS")
    print("="*70)

    print(f"[BANK] Carpeta Bancos:   {CARPETA_BANCOS}")
    print(f"[BOOK] Carpeta Contable: {CARPETA_CONTABLE}\n")

    # ─────────────────────────────────────────────
    # Buscar archivos de BANCO
    # ─────────────────────────────────────────────
    archivos_banco = []
    for extension in ['*.xlsx', '*.xls', '*.xlsm', '*.csv']:
        for ruta_archivo in glob.glob(os.path.join(CARPETA_BANCOS, extension)):
            nombre_archivo = os.path.basename(ruta_archivo)
            nombre_sin_ext = os.path.splitext(nombre_archivo)[0]

            banco_det, nombre_det = detectar_banco_en_nombre_archivo(nombre_sin_ext)
            if banco_det:
                archivos_banco.append((ruta_archivo, banco_det, nombre_det, nombre_archivo))
                print(f"[OK] BANCO identificado: {nombre_archivo}")
                print(f"   [BANK] Banco: {banco_det} ({nombre_det})")

    # ─────────────────────────────────────────────
    # Buscar archivo CONTABLE
    # ─────────────────────────────────────────────
    archivos_contable = []
    for extension in ['*.xlsx', '*.xls', '*.xlsm', '*.csv']:
        for ruta_archivo in glob.glob(os.path.join(CARPETA_CONTABLE, extension)):
            nombre_archivo = os.path.basename(ruta_archivo)
            archivos_contable.append((ruta_archivo, 'CONTABLE', 'Libro Contable', nombre_archivo))
            print(f"[OK] CONTABLE identificado: {nombre_archivo}")

    print("\n" + "─"*70)

    # ─────────────────────────────────────────────
    # Validaciones
    # ─────────────────────────────────────────────
    if not archivos_banco:
        print("[ERROR] ERROR: No se encontraron archivos de banco")
        return None, None, None, None

    if not archivos_contable:
        print("[ERROR] ERROR: No se encontró archivo del libro contable")
        return None, None, None, None

    # ─────────────────────────────────────────────
    # Selección BANCO
    # ─────────────────────────────────────────────
    if len(archivos_banco) > 1:
        print("\n[!] MÚLTIPLES ARCHIVOS DE BANCO:\n")
        for i, (_, codigo, nombre, archivo) in enumerate(archivos_banco, 1):
            print(f"   {i}) {archivo} — {codigo} ({nombre})")

        while True:
            entrada = input(f"\nSelecciona banco (1-{len(archivos_banco)}): ").strip()
            if entrada.isdigit() and 1 <= int(entrada) <= len(archivos_banco):
                archivo_banco, codigo_banco, nombre_banco, _ = archivos_banco[int(entrada) - 1]
                break
            print("[ERROR] Selección inválida")
    else:
        archivo_banco, codigo_banco, nombre_banco, _ = archivos_banco[0]

    # ─────────────────────────────────────────────
    # Selección CONTABLE
    # ─────────────────────────────────────────────
    if len(archivos_contable) > 1:
        print("\n[!] MÚLTIPLES ARCHIVOS CONTABLES:\n")
        for i, (_, _, _, archivo) in enumerate(archivos_contable, 1):
            print(f"   {i}) {archivo}")

        while True:
            entrada = input(f"\nSelecciona libro contable (1-{len(archivos_contable)}): ").strip()
            if entrada.isdigit() and 1 <= int(entrada) <= len(archivos_contable):
                archivo_contable, _, _, _ = archivos_contable[int(entrada) - 1]
                break
            print("[ERROR] Selección inválida")
    else:
        archivo_contable, _, _, _ = archivos_contable[0]

    # ─────────────────────────────────────────────
    # Resultado final
    # ─────────────────────────────────────────────
    print("\n[OK] Archivos validados correctamente")
    print(f"   [BANK] Banco:    {os.path.basename(archivo_banco)}")
    print(f"   [BOOK] Contable: {os.path.basename(archivo_contable)}")

    return archivo_banco, archivo_contable, codigo_banco, nombre_banco

# ============================================================================
# [NOTE] ALIAS DE EMPRESAS Y CONCEPTOS (AMPLIADO PARA TC/LEGAL)
# ============================================================================

ALIAS_EMPRESAS = {
    # RG-RB / CONCRESOL
    'RGRB': ['RG RB', 'RGRB', 'RG-RB', 'CONCRESOL', 'CONCRESOL SRL'],
    'CONCRESOL': ['RG RB', 'RGRB', 'RG-RB', 'CONCRESOL', 'CONCRESOL SRL'],
    
    # TC / TARJETA CORPORATIVA / LEGALIZACIONES [NUEVO v6.5]
    'TC_CORPORATIVA': [
        'TC LEGAL', 'TC CORPORATIVA', 'TARJETA CORPORATIVA', 'TC',
        'LEGAL', 'LEGALIZACIONES', 'LEGALIZACION', 'IPI',
        'CORTE TC', 'CORTE TC LEGAL', 'PAGO TC', 'REEMBOLSO TC',
        'REEMBOLSO CORTE TC', 'REEMBOLSO TC LEGAL', 'DEBITO PAGO',
        'DEBITO PAGO RD', 'PAGO RD TC', 'TC CORPORATIVA'
    ],
}

# Patrones específicos
PATRONES_TC = [
    'TC LEGAL', 'TC CORPORATIVA', 'TARJETA CORPORATIVA', 'CORTE TC',
    'REEMBOLSO TC', 'DEBITO PAGO', 'LEGALIZACIONES', 'IPI',
    'PAGO RD TC', 'REEMBOLSO CORTE'
]

PATRONES_IMPUESTO = ['IMP', 'IMPUESTO', 'DGII', '0.15', '0.15%', 'ITF', 'RETENCION']

PATRONES_COMISION = [
    'COM', 'COMISION', 'COM.', 'CARGO', 'COSTO', 'GASTO BANCARIO',
    'EXI', 'COMISION BANCARIA', 'COMISIONES', 'LBTR',
    'TRANS APROBA', 'DEBITO CUENTA', 'GASTO', 'SERVICIO'
]

PATRONES_TRANSFERENCIA = ['TRANSFER', 'TRANSF', 'ACH', 'CR TRANSFERENCIA', 'REALIT LLC','LLC','REALIT']

PALABRAS_COMUNES = {
    'DE', 'DEL', 'LA', 'EL', 'LOS', 'LAS', 'CON', 'SIN', 'PARA', 'POR', 'QUE',
    'COMO', 'DESDE', 'HASTA', 'ENTRE', 'ACH', 'TRANSFERENCIA', 'PAGO', 'TRANSF',
    'SAS', 'LTDA', 'SA', 'SRL', 'CIA', 'RECAUDO', 'INMOB', 'FIDEICOMISO',
    'DEPOSITO', 'BANCO', 'CTA', 'CTAS', 'FAC', 'FACT', 'FACTURA', 'NUM',
    'NUMERO', 'NO', 'Y', 'A', 'EN', 'AL', 'O', 'U', 'E', 'USD', 'DOP', 'RD'
}

# ============================================================================
# [NOTE] PATRONES PARA CAJA CHICA (AÑADIR A LA SECCIÓN DE ALIAS)
# ============================================================================

PATRONES_CAJA_CHICA = [
    'CAJA CHICA', 'CAJACHICA', 'REPOSICION', 'REPOSICION DE CAJA',
    'REP CAJA', 'REPO CAJA', 'REEMBOLSO CAJA', 'CAJA MENOR'
]

# ============================================================================
# [TOOL] FUNCIONES DE NORMALIZACIÓN
# ============================================================================

def quitar_acentos(texto):
    if pd.isna(texto) or texto is None:
        return ""
    texto = str(texto).strip()
    if not texto:
        return ""
    texto_nfd = unicodedata.normalize('NFD', texto)
    return ''.join(c for c in texto_nfd if unicodedata.category(c) != 'Mn')

def normalizar_texto(texto):
    if pd.isna(texto) or texto is None:
        return ""
    texto = quitar_acentos(str(texto))
    texto = texto.upper()
    texto = re.sub(r'[^A-Z0-9\s]', ' ', texto)
    texto = re.sub(r'\s+', ' ', texto)
    return texto.strip()

def es_patron_tc(texto):
    """Detecta si el texto corresponde a TC/Legalizaciones"""
    texto_norm = normalizar_texto(texto)
    return any(patron in texto_norm for patron in PATRONES_TC)

def normalizar_nombre_empresa(texto):
    """Normaliza nombre de empresa + detecta TC/LEGAL"""
    texto_norm = normalizar_texto(texto)
    if not texto_norm:
        return ""
    
    # PRIMERO: Verificar si es TC/LEGAL (prioridad alta)
    if es_patron_tc(texto):
        return 'TC_CORPORATIVA'
    
    # Eliminar sufijos legales
    sufijos = ['SRL', 'SA', 'SAS', 'LTDA', 'CIA', 'INC', 'LLC', 'CORP']
    for sufijo in sufijos:
        texto_norm = re.sub(rf'\b{sufijo}\b', '', texto_norm)
    
    texto_norm = re.sub(r'\s+', ' ', texto_norm).strip()
    texto_compacto = texto_norm.replace(' ', '')
    
    # Buscar en alias
    for clave, aliases in ALIAS_EMPRESAS.items():
        for alias in aliases:
            alias_compacto = normalizar_texto(alias).replace(' ', '')
            if alias_compacto and texto_compacto:
                if alias_compacto in texto_compacto or texto_compacto in alias_compacto:
                    return clave
    
    return texto_norm

def extraer_palabras_clave(texto):
    texto_norm = normalizar_texto(texto)
    if not texto_norm:
        return []
    palabras = [p for p in texto_norm.split() if len(p) >= 3 and p not in PALABRAS_COMUNES]
    numeros = re.findall(r'\d{5,}', texto_norm)
    return list(set(palabras + numeros))

def extraer_identificador_proveedor(texto):
    texto_norm = normalizar_texto(texto)
    if not texto_norm:
        return None
    match = re.search(r'^\s*(\d{8,})', texto_norm)
    return match.group(1) if match else None

def calcular_similitud(texto1, texto2):
    # Comparar nombres de empresa primero
    emp1 = normalizar_nombre_empresa(texto1)
    emp2 = normalizar_nombre_empresa(texto2)
    
    if emp1 and emp2 and emp1 == emp2:
        return 1.0
    
    palabras1 = set(extraer_palabras_clave(texto1))
    palabras2 = set(extraer_palabras_clave(texto2))
    
    if not palabras1 or not palabras2:
        return 0.0
    
    coincidencias = len(palabras1 & palabras2)
    
    # Coincidencias difusas
    for p1 in palabras1:
        for p2 in palabras2:
            if len(p1) >= 4 and len(p2) >= 4:
                if SequenceMatcher(None, p1, p2).ratio() > 0.85:
                    coincidencias += 0.5
                    break
    
    max_palabras = max(len(palabras1), len(palabras2))
    return coincidencias / max_palabras if max_palabras > 0 else 0.0

def es_patron_impuesto(texto):
    texto_norm = normalizar_texto(texto)
    return any(patron in texto_norm for patron in PATRONES_IMPUESTO)

def es_patron_comision(texto):
    texto_norm = normalizar_texto(texto)
    return any(patron in texto_norm for patron in PATRONES_COMISION)

def es_patron_transferencia(texto):
    texto_norm = normalizar_texto(texto)
    return any(patron in texto_norm for patron in PATRONES_TRANSFERENCIA)

# ============================================================================
# [BANK] PASO 2: SISTEMA DE LIMPIEZA ESPECÍFICO POR BANCO
# ============================================================================

def limpiar_banco_popular(df_original):
    """
    Limpieza específica para POPULAR en formato EXCEL
    
    Estructura esperada (sin headers):
    Fecha de Posteo; Descripcion Corta (CONCEPTO); MONTO; USELESS; USELESS; USELESS; Descripcion
    
    Las primeras 10 filas deben ignorarse.
    """
    print("\n   Aplicando limpieza: POPULAR")
    df = df_original.copy()
    df = df.reset_index(drop=True)
    # Eliminar filas completamente vacías
    df = df.dropna(how='all')

    # Buscar fila de cabecera en las primeras N filas
    header_row = None
    max_search = min(25, len(df))
    for i in range(max_search):
        vals = ' '.join([str(x) for x in df.iloc[i].values if pd.notna(x)])
        up = quitar_acentos(vals).upper()
        if 'FECHA POSTEO' in up and ("DESCRIPCION CORTA" in up or 'DESCRIPCION' in up or 'MONTO' in up or 'MONTO TRANSACCION' in up or 'TRANSACCION' in up):
            header_row = i
            break

    # Fallback: buscar solo 'FECHA'
    if header_row is None:
        for i in range(max_search):
            vals = ' '.join([str(x) for x in df.iloc[i].values if pd.notna(x)])
            up = quitar_acentos(vals).upper()
            if 'FECHA' in up:
                header_row = i
                break

    if header_row is None:
        # No se detectó cabecera; devolver df con filas vacías eliminadas
        return df

    # Aplicar la fila encontrada como cabecera y devolver solo las filas de datos
    raw_header = df.iloc[header_row].astype(str).tolist()
    df = df.iloc[header_row + 1 :].reset_index(drop=True)
    df.columns = raw_header

    # ===== Detectar segunda cabecera (inicio de Débitos) =====
    header_norm = [quitar_acentos(str(c)).upper() for c in raw_header]

    idx_separador = None

    for i in range(len(df)):
        fila_norm = [
            quitar_acentos(str(x)).upper()
            for x in df.iloc[i].values
            if pd.notna(x)
        ]

        # Si la fila contiene TODAS las columnas de cabecera → es cabecera repetida
        if all(any(h in v for v in fila_norm) for h in header_norm):
            idx_separador = i
            break

    if idx_separador is not None:
        print(f"  [MERGE] Segunda cabecera detectada en fila: {idx_separador}")
    else:
        print("  [WARN] No se detectó segunda cabecera (Débitos)")

    # Normalizar nombres de columnas mínimamente (sin cambiar valores)
    rename_map = {}
    for col in list(df.columns):
        key = quitar_acentos(str(col)).upper()
        if 'FECHA POSTEO' in key:
            rename_map[col] = 'Fecha'
        elif 'DESCRIPCION CORTA' in key:
            rename_map[col] = 'Concepto'
        elif any(k in key for k in ['DETALLE', 'OBSERV', 'DESCRIPCIÓN']):
            rename_map[col] = 'Descripción'
        elif any(k in key for k in ['VALOR', 'MONTO', 'IMPORTE', 'AMOUNT']):
            rename_map[col] = 'Valor'

    if rename_map:
        df = df.rename(columns=rename_map)

    # Convertir Valor
    df['Valor'] = pd.to_numeric(df['Valor'], errors='coerce')

    # Aplicar signo: Débitos negativos
    if idx_separador is not None:
        df.loc[idx_separador + 1 :, 'Valor'] *= -1

    # Quitar filas de totales u otros rótulos (líneas que contienen 'TOTAL' u 'NETO')
    mask_tot = df.apply(lambda r: any('TOTAL' in str(x).upper() or 'NETO' in str(x).upper() for x in r.values), axis=1)
    if mask_tot.any():
        df = df[~mask_tot].reset_index(drop=True)

    # Eliminar filas vacías residuales
    df = df.dropna(how='all').reset_index(drop=True)

    return df

def detectar_delimitador_popular(ruta_archivo):
    with open(ruta_archivo, 'r', encoding='latin1') as f:
        for line in f:
            up = quitar_acentos(line).upper()
            if 'FECHA POSTEO' in up and 'MONTO' in up:
                return ';' if line.count(';') > line.count(',') else ','
    raise ValueError("No se pudo detectar el delimitador")

def detectar_headers_en_df(df):
    """
    Devuelve una lista de índices donde aparece la cabecera
    (Créditos, Débitos, etc.)
    """
    headers = []
    for i in range(len(df)):
        txt = ' '.join(str(x) for x in df.iloc[i].values if pd.notna(x))
        up = quitar_acentos(txt).upper()
        if 'FECHA' in up and 'MONTO' in up:
            headers.append(i)
    return headers

from io import StringIO
def read_dirty_csv_popular(ruta_archivo, sep):
    """
    Reads a bank CSV embedded in garbage text.
    Keeps ONLY lines that contain the delimiter.
    """
    good_lines = []

    with open(ruta_archivo, 'r', encoding='latin1') as f:
        for line in f:
            if sep in line:
                good_lines.append(line)

    if not good_lines:
        raise ValueError("No se encontraron líneas con delimitador")

    return pd.read_csv(
        StringIO(''.join(good_lines)),
        sep=sep,
        header=None,
        engine='python',
        dtype=str
    )

def limpiar_banco_popular_csv(ruta_archivo):

    print("\n   Aplicando limpieza: POPULAR (CSV)")

    try:
        # ─────────────────────────────────────────────
        # 1. Detectar delimitador
        # ─────────────────────────────────────────────
        sep = detectar_delimitador_popular(ruta_archivo)
        print(f"  [SEARCH] Delimitador detectado: '{sep}'")

        # ─────────────────────────────────────────────
        # 2. Leer archivo CSV ignorando basura
        # ─────────────────────────────────────────────
        df = read_dirty_csv_popular(ruta_archivo, sep)


        print(f"  [STATS] CSV cargado: {df.shape[0]} filas × {df.shape[1]} columnas")

        # ─────────────────────────────────────────────
        # 3. Detectar cabeceras dentro del DF
        # ─────────────────────────────────────────────
        headers = detectar_headers_en_df(df)

        if len(headers) == 0:
            raise ValueError("No se detectó ninguna tabla")

        idx_creditos = headers[0]
        idx_debitos = headers[1] if len(headers) > 1 else None

        print(f"  [SEARCH] Inicio Créditos: fila {idx_creditos}")
        if idx_debitos is not None:
            print(f"  [SEARCH] Inicio Débitos: fila {idx_debitos}")

        # ─────────────────────────────────────────────
        # 4. Recortar DF a partir de Créditos
        # ─────────────────────────────────────────────
        df = df.iloc[idx_creditos + 1:].reset_index(drop=True)

        df.columns = [f'col_{i}' for i in range(df.shape[1])]

        # ─────────────────────────────────────────────
        # 5. Construir DF limpio
        # ─────────────────────────────────────────────
        df_limpio = pd.DataFrame({
            'Fecha': df['col_0'],
            'Concepto': df['col_1'],
            'Valor': df['col_2'],
            'Descripción': df['col_6'],
        })

        # ─────────────────────────────────────────────
        # 6. Detectar cabecera Débitos (ya alineada)
        # ─────────────────────────────────────────────
        idx_debitos_df = None
        for i in range(len(df)):
            txt = ' '.join(str(x) for x in df.iloc[i].values if pd.notna(x))
            up = quitar_acentos(txt).upper()
            if 'FECHA' in up and 'MONTO' in up:
                idx_debitos_df = i
                break

        # ─────────────────────────────────────────────
        # 7. Tipos
        # ─────────────────────────────────────────────
        df_limpio['Fecha'] = pd.to_datetime(
            df_limpio['Fecha'].str.strip(),
            errors='coerce',
            dayfirst=True
        )

        df_limpio['Valor'] = pd.to_numeric(
            df_limpio['Valor'].str.strip(),
            errors='coerce'
        )

        # ─────────────────────────────────────────────
        # 8. Aplicar signo Débitos
        # ─────────────────────────────────────────────
        if idx_debitos_df is not None:
            print(f"  [MERGE] Aplicando Débitos desde fila {idx_debitos_df}")
            df_limpio.loc[idx_debitos_df + 1:, 'Valor'] *= -1
            df_limpio = df_limpio.drop(index=idx_debitos_df).reset_index(drop=True)

        # ─────────────────────────────────────────────
        # 9. Limpieza final
        # ─────────────────────────────────────────────
        df_limpio = df_limpio.dropna(subset=['Valor']).reset_index(drop=True)

        print(f"  [OK] Limpieza POPULAR CSV completada: {df_limpio.shape[0]} filas")
        return df_limpio

    except Exception as e:
        print(f"  [ERROR] Error al limpiar CSV de POPULAR: {e}")
        raise

def limpiar_banreservas(df_original):
    """Limpieza ligera para BANRESERVAS.

    Objetivo: eliminar filas arriba de la tabla de datos (texto libre, títulos,
    metadatos) para que pandas pueda leer correctamente la tabla. NO debe
    transformar o combinar columnas (p.ej. Debito/Credito) — eso lo hace el
    loader universal posteriormente.
    """
    print("\n   Aplicando limpieza: BANRESERVAS (solo remover encabezados no-datos)")
    df = df_original.copy()
    df = df.reset_index(drop=True)
    # Eliminar filas completamente vacías
    df = df.dropna(how='all')

    # Buscar fila de cabecera en las primeras N filas
    header_row = None
    max_search = min(25, len(df))
    for i in range(max_search):
        vals = ' '.join([str(x) for x in df.iloc[i].values if pd.notna(x)])
        up = quitar_acentos(vals).upper()
        if 'FECHA' in up and ("TRANSACC" in up or 'CONCEP' in up or 'NO DE TRANSACCION' in up or 'NO.' in up or 'NO DE TRANSAC' in up):
            header_row = i
            break

    # Fallback: buscar solo 'FECHA'
    if header_row is None:
        for i in range(max_search):
            vals = ' '.join([str(x) for x in df.iloc[i].values if pd.notna(x)])
            up = quitar_acentos(vals).upper()
            if 'FECHA' in up:
                header_row = i
                break

    if header_row is None:
        # No se detectó cabecera; devolver df con filas vacías eliminadas
        return df

    # Aplicar la fila encontrada como cabecera y devolver solo las filas de datos
    raw_header = df.iloc[header_row].astype(str).tolist()
    df = df.iloc[header_row + 1 :].reset_index(drop=True)
    df.columns = raw_header

    # Normalizar nombres de columnas mínimamente (sin cambiar valores)
    rename_map = {}
    for col in list(df.columns):
        key = quitar_acentos(str(col)).upper()
        if 'FECHA' in key:
            rename_map[col] = 'Fecha'
        elif 'CONCEP' in key:
            rename_map[col] = 'Concepto'
        elif 'DEB' in key and 'DEBITO' in key or 'DEBITO' in key:
            rename_map[col] = 'Debito'
        elif 'CRED' in key or 'CREDITO' in key:
            rename_map[col] = 'Credito'
        elif 'REFER' in key:
            rename_map[col] = 'Referencia'
        elif any(k in key for k in ['DESCRIP', 'DETALLE', 'OBSERV', 'DESCRIPCIÓN']):
            rename_map[col] = 'Descripción'
        elif any(k in key for k in ['VALOR', 'MONTO', 'IMPORTE', 'AMOUNT']):
            rename_map[col] = 'Valor'

    if rename_map:
        df = df.rename(columns=rename_map)

    # Quitar filas de totales u otros rótulos (líneas que contienen 'TOTAL' u 'NETO')
    mask_tot = df.apply(lambda r: any('TOTAL' in str(x).upper() or 'NETO' in str(x).upper() for x in r.values), axis=1)
    if mask_tot.any():
        df = df[~mask_tot].reset_index(drop=True)

    # Eliminar filas vacías residuales
    df = df.dropna(how='all').reset_index(drop=True)

    return df

def limpiar_bhd(df_original):
    """Limpieza específica para BHD (Excel)"""
    print("\n   Aplicando limpieza: BHD")
    df = df_original.copy()
    df = df.reset_index(drop=True)
    df = df.dropna(how='all')
    return df

def limpiar_bhd_csv(ruta_archivo):
    """
    Limpieza específica para BHD en formato CSV
    
    Estructura esperada (sin headers):
    Fecha, USELESS, USELESS, USELESS, Concepto, Debito, Credito, USELESS, Balance, Hour
    
    La primera fila es un resumen y debe ignorarse.
    """
    print("\n   Aplicando limpieza: BHD (CSV)")
    
    try:
        # Leer CSV sin headers (no_column_names)
        df = pd.read_csv(ruta_archivo, header=None)
        print(f"  [STATS] CSV cargado: {df.shape[0]} filas × {df.shape[1]} columnas")
        
        # Asignar nombres temporales a las columnas (0-indexed)
        df.columns = [f'col_{i}' for i in range(df.shape[1])]
        
        # Saltar la primera fila (resumen/encabezado)
        df = df.iloc[1:].reset_index(drop=True)
        print(f"  [OK] Salteada primera fila (resumen), quedaron: {df.shape[0]} filas")
        
        # Mapear columnas: Fecha=0, Concepto=4, Debito=5, Credito=6
        # Las demás columnas (1,2,3,7,8,9) son inútiles
        df_limpio = pd.DataFrame({
            'Fecha': df['col_0'],
            'Concepto': df['col_4'],
            'Descripción': '',  # Campo vacío por defecto
            'Debito': df['col_5'],
            'Credito': df['col_6']
        })
        
        # Limpiar datos
        df_limpio = df_limpio.reset_index(drop=True)
        df_limpio = df_limpio.dropna(how='all')
        
        # Convertir a tipos de datos correctos
        df_limpio['Fecha'] = pd.to_datetime(df_limpio['Fecha'], errors='coerce')
        df_limpio['Debito'] = pd.to_numeric(df_limpio['Debito'], errors='coerce')
        df_limpio['Credito'] = pd.to_numeric(df_limpio['Credito'], errors='coerce')
        
        print(f"  [OK] Limpieza BHD CSV completada: {df_limpio.shape[0]} filas")
        return df_limpio
        
    except Exception as e:
        print(f"  [ERROR] Error al limpiar CSV de BHD: {e}")
        raise

def limpiar_santa_cruz(df_original):
    """Limpieza ligera para SANTA CRUZ.

    Objetivo: eliminar filas arriba de la tabla de datos (texto libre, títulos,
    metadatos) para que pandas pueda leer correctamente la tabla. NO debe
    transformar o combinar columnas (p.ej. Debito/Credito) — eso lo hace el
    loader universal posteriormente.
    
    Mapeo de columnas Santa Cruz:
    - Fecha de Posteo → IGNORADA
    - Fecha Efectiva → Fecha (IMPORTANTE)
    - No. Cheque → IGNORADA
    - No. Referencia → Concepto (ID de transacción)
    - Descripcion → Descripción
    - Retiros → Debito
    - Despositos → Credito
    - Balance → IGNORADA
    """
    print("\n   Aplicando limpieza: SANTA CRUZ (solo remover encabezados no-datos)")
    df = df_original.copy()
    df = df.reset_index(drop=True)
    # Eliminar filas completamente vacías
    df = df.dropna(how='all')

    # Buscar fila de cabecera en las primeras N filas
    header_row = None
    max_search = min(25, len(df))
    for i in range(max_search):
        vals = ' '.join([str(x) for x in df.iloc[i].values if pd.notna(x)])
        up = quitar_acentos(vals).upper()
        # Buscar palabras clave de Santa Cruz
        if any(k in up for k in ['FECHA EFECTIVA', 'FECHA DE POSTEO', 'NO. REFERENCIA', 'REFERENCIA', 'RETIROS', 'DESPOSITOS', 'DEPOSITOS']):
            header_row = i
            break

    # Fallback: buscar solo 'FECHA'
    if header_row is None:
        for i in range(max_search):
            vals = ' '.join([str(x) for x in df.iloc[i].values if pd.notna(x)])
            up = quitar_acentos(vals).upper()
            if 'FECHA' in up:
                header_row = i
                break

    if header_row is None:
        # No se detectó cabecera; devolver df con filas vacías eliminadas
        return df

    # Aplicar la fila encontrada como cabecera y devolver solo las filas de datos
    raw_header = df.iloc[header_row].astype(str).tolist()
    df = df.iloc[header_row + 1 :].reset_index(drop=True)
    df.columns = raw_header

    # Normalizar nombres de columnas para Santa Cruz
    rename_map = {}
    for col in list(df.columns):
        key = quitar_acentos(str(col)).upper()
        
        # Mapeo específico para Santa Cruz
        if 'FECHA EFECTIVA' in key:
            rename_map[col] = 'Fecha'
        elif 'FECHA DE POSTEO' in key or 'FECHA POSTEO' in key:
            # Ignorar Fecha de Posteo (no usar para mapeo)
            pass
        elif 'NO. REFERENCIA' in key or 'NO REFERENCIA' in key or 'REFERENCIA' in key:
            rename_map[col] = 'Concepto'
        elif 'DESCRIPCION' in key or 'DESCRIP' in key:
            rename_map[col] = 'Descripción'
        elif 'RETIROS' in key or 'RETIRO' in key:
            rename_map[col] = 'Debito'
        elif 'DESPOSITOS' in key or 'DEPOSITOS' in key or 'DEPOSITO' in key:
            rename_map[col] = 'Credito'
        elif 'NO. CHEQUE' in key or 'NO CHEQUE' in key or 'CHEQUE' in key:
            # Ignorar No. Cheque
            pass
        elif 'BALANCE' in key or 'SALDO' in key:
            # Ignorar Balance/Saldo
            pass

    if rename_map:
        df = df.rename(columns=rename_map)

    # Quitar filas de totales u otros rótulos (líneas que contienen 'TOTAL' u 'NETO')
    mask_tot = df.apply(lambda r: any('TOTAL' in str(x).upper() or 'NETO' in str(x).upper() for x in r.values), axis=1)
    if mask_tot.any():
        df = df[~mask_tot].reset_index(drop=True)

    # Eliminar filas vacías residuales
    df = df.dropna(how='all').reset_index(drop=True)

    return df

def limpiar_apap(df_original):
    """Limpieza ligera para APAP.

    Objetivo: eliminar filas arriba de la tabla de datos (texto libre, títulos,
    metadatos) para que pandas pueda leer correctamente la tabla. NO debe
    transformar o combinar columnas (p.ej. Debito/Credito) — eso lo hace el
    loader universal posteriormente.
    """
    print("\n   Aplicando limpieza: APAP (solo remover encabezados no-datos)")
    df = df_original.copy()
    df = df.reset_index(drop=True)
    # Eliminar filas completamente vacías
    df = df.dropna(how='all')

    # Buscar fila de cabecera en las primeras N filas
    header_row = None
    max_search = min(25, len(df))
    for i in range(max_search):
        vals = ' '.join([str(x) for x in df.iloc[i].values if pd.notna(x)])
        up = quitar_acentos(vals).upper()
        if 'FECHA' in up and ("REFER" in up or 'CONCEP' in up or 'NO DE REFERENCIA' in up or 'NO.' in up or 'DESC' in up):
            header_row = i
            break

    # Fallback: buscar solo 'FECHA'
    if header_row is None:
        for i in range(max_search):
            vals = ' '.join([str(x) for x in df.iloc[i].values if pd.notna(x)])
            up = quitar_acentos(vals).upper()
            if 'FECHA' in up:
                header_row = i
                break

    if header_row is None:
        # No se detectó cabecera; devolver df con filas vacías eliminadas
        return df

    # Aplicar la fila encontrada como cabecera y devolver solo las filas de datos
    raw_header = df.iloc[header_row].astype(str).tolist()
    df = df.iloc[header_row + 1 :].reset_index(drop=True)
    df.columns = raw_header

    # Normalizar nombres de columnas mínimamente (sin cambiar valores)
    rename_map = {}
    for col in list(df.columns):
        key = quitar_acentos(str(col)).upper()
        if 'FECHA DE ENTRADA' in key:
            rename_map[col] = 'Fecha'
        elif 'NO. DE REFERENCIA' in key:
            rename_map[col] = 'Concepto'
        elif any(k in key for k in ['DESCRIP', 'DETALLE', 'OBSERV', 'DESCRIPCIÓN']):
            rename_map[col] = 'Descripción'
        elif any(k in key for k in ['VALOR', 'MONTO', 'IMPORTE', 'AMOUNT']):
            rename_map[col] = 'Valor'

    if rename_map:
        df = df.rename(columns=rename_map)

    # Quitar filas de totales u otros rótulos (líneas que contienen 'TOTAL' u 'NETO')
    mask_tot = df.apply(lambda r: any('TOTAL' in str(x).upper() or 'NETO' in str(x).upper() for x in r.values), axis=1)
    if mask_tot.any():
        df = df[~mask_tot].reset_index(drop=True)

    # Eliminar filas vacías residuales
    df = df.dropna(how='all').reset_index(drop=True)

    return df

def limpiar_ademi(df_original):
    """Limpieza específica para ADEMI"""
    print("\n   Aplicando limpieza: ADEMI")
    df = df_original.copy()
    df = df.reset_index(drop=True)
    df = df.dropna(how='all')
    return df

def limpiar_lopez_de_haro(df_original):
    """Limpieza específica para LOPEZ_DE_HARO"""
    print("\n   Aplicando limpieza: LOPEZ_DE_HARO")
    df = df_original.copy()
    df = df.reset_index(drop=True)
    df = df.dropna(how='all')
    return df

def limpiar_bellbank(df_original):
    """Limpieza específica para BELLBANK"""
    print("\n   Aplicando limpieza: BELLBANK")
    df = df_original.copy()
    df = df.reset_index(drop=True)
    df = df.dropna(how='all')
    return df

def limpiar_banesco(df_original):
    """Limpieza específica para BANESCO"""
    print("\n   Aplicando limpieza: BANESCO")
    df = df_original.copy()
    df = df.reset_index(drop=True)
    df = df.dropna(how='all')
    return df

def limpiar_scotiabank(df_original):
    """Limpieza específica para SCOTIABANK"""
    print("\n   Aplicando limpieza: SCOTIABANK")
    df = df_original.copy()
    df = df.reset_index(drop=True)
    df = df.dropna(how='all')
    return df

def limpiar_promerica(df_original):
    """Limpieza específica para PROMERICA"""
    print("\n   Aplicando limpieza: PROMERICA")
    df = df_original.copy()
    df = df.reset_index(drop=True)
    df = df.dropna(how='all')
    return df

# Mapeo de bancos a funciones de limpieza
FUNCIONES_LIMPIEZA_BANCO = {
    'POPULAR': limpiar_banco_popular,
    'BANRESERVAS': limpiar_banreservas,
    'BHD': limpiar_bhd,
    'BANESCO': limpiar_banesco,
    'SCOTIABANK': limpiar_scotiabank,
    'PROMERICA': limpiar_promerica,
    'SANTA_CRUZ': limpiar_santa_cruz,
    'ADEMI': limpiar_ademi,
    'LOPEZ_DE_HARO': limpiar_lopez_de_haro,
    'BELLBANK': limpiar_bellbank,
    'APAP': limpiar_apap
}

FUNCIONES_LIMPIEZA_BANCO_CSV = {
    'BHD': limpiar_bhd_csv,
    'POPULAR': limpiar_banco_popular_csv
}

# ============================================================================
# [BANK] PASO 2B: LIMPIEZA DEL ARCHIVO CONTABLE (UNIVERSAL)
# ============================================================================

def limpiar_archivo_contable(ruta_archivo, moneda='RD$'):
    """
    Limpia el archivo contable (formato estándar para todos los casos)
    
    PARÁMETROS:
    - ruta_archivo: ruta del archivo Excel
    - moneda: 'RD$' (pesos) o 'USD' (dólares) - para futura implementación

    ---CODIGO PARA LIMPIEZA---
    
    """
    print("\n  Cargando archivo CONTABLE ({} )... (limpieza ligera)".format(moneda))
    # Leer Excel y eliminar filas vacías por encima de la tabla principal, sin
    # interpretar la naturaleza (Natu) ni transformar signos. Solo establecer
    # la fila de cabecera correcta para que el loader universal mapée columnas.
    try:
        df = pd.read_excel(ruta_archivo)
    except Exception as e:
        raise ValueError(f"[ERROR] Error al leer contable '{ruta_archivo}': {e}")

    df = df.reset_index(drop=True)
    df = df.dropna(how='all')

    header_row = None
    max_search = min(30, len(df))
    for i in range(max_search):
        vals = ' '.join([str(x) for x in df.iloc[i].values if pd.notna(x)])
        up = quitar_acentos(vals).upper()
        
        # Skip metadata rows that contain date range info (D.Fecha, H.Fecha)
        if 'D.FECHA' in up or 'H.FECHA' in up or 'DESDE' in up or 'HASTA' in up:
            continue
        
        # Buscar elementos típicos del libro contable: F.Comp, Detalle, Nro.Doc, Periodo
        # Require at least 2 of these keywords to be more specific and avoid false positives
        keyword_matches = sum(1 for k in ['F.COMP', 'F COMP', 'DETALLE', 'NRO', 'NRO DOC', 'PERIODO'] if k in up)
        if keyword_matches >= 2:
            header_row = i
            break

    if header_row is None:
        return df

    raw_header = df.iloc[header_row].astype(str).tolist()
    df = df.iloc[header_row + 1 :].reset_index(drop=True)
    df.columns = raw_header

    # No hacemos más transformaciones aquí (no tocar NATU ni signos)
    df = df.dropna(how='all').reset_index(drop=True)
    return df

# ============================================================================
# [DIR] CARGA DE DATOS OPTIMIZADA (SOLO FILAS VÁLIDAS)
# ============================================================================

# Carga de Archivos
def cargar_banco(ruta, nombre="", codigo_banco=None):
    """Carga datos - OPTIMIZACIÓN: solo lee filas con Fecha Y Valor válidos
    Si se proporciona `codigo_banco`, se aplicará la función de limpieza
    específica definida en `FUNCIONES_LIMPIEZA_BANCO` antes del mapeo general.
    """
    print(f"\n  [DIR] Cargando: {nombre.upper() if nombre else ruta}")

    # Resolver ruta relativa: si el usuario pasa solo el nombre de archivo,
    # buscarlo en el mismo directorio del script y en el cwd.
    try:
        ruta_path = Path(ruta)
    except Exception:
        ruta_path = Path(str(ruta))

    if not ruta_path.is_absolute():
        script_dir = base_path()
        candidates = [script_dir / ruta_path, Path.cwd() / ruta_path]

        # If no suffix provided, try common extensions too
        if not ruta_path.suffix:
            exts = ['.xlsx', '.xls', '.csv']
        else:
            exts = [ruta_path.suffix]

        found = None
        for base in candidates:
            for ext in exts:
                cand = base.with_suffix(ext) if not ruta_path.suffix else base
                if cand.exists():
                    found = cand
                    break
            if found:
                break

        if found:
            ruta = str(found)
        else:
            # if the literal provided (e.g., 'BANRESERVAS.xlsx') exists in script_dir
            alt = script_dir / ruta_path.name
            if alt.exists():
                ruta = str(alt)

    try:
        # Detectar si es CSV o Excel basado en la extensión del archivo
        if ruta.lower().endswith('.csv'):
            print(f"  [LIST] Detectado archivo CSV")
            # Aplicar limpieza específica si existe para CSV
            if codigo_banco and codigo_banco in FUNCIONES_LIMPIEZA_BANCO_CSV:
                print(f"  [CLEAN] Aplicando limpieza específica para {codigo_banco} (CSV)...")
                df = FUNCIONES_LIMPIEZA_BANCO_CSV[codigo_banco](ruta)
            else:
                # Para otros CSVs, cargar directamente
                df = pd.read_csv(ruta)
        else:
            # Leer Excel
            df = pd.read_excel(ruta)

        # [PERF] OPTIMIZACIÓN 1: Eliminar filas completamente vacías PRIMERO
        df = df.dropna(how='all')

        # Si se indicó un código de banco y existe una función de limpieza,
        # aplicar la limpieza específica antes del mapeo automático.
        # (Para CSV de BHD, la limpieza ya se aplicó arriba)
        if codigo_banco and codigo_banco in FUNCIONES_LIMPIEZA_BANCO and not (ruta.lower().endswith('.csv')):
            try:
                print(f"  [CLEAN] Aplicando limpieza específica para {codigo_banco}...")
                df = FUNCIONES_LIMPIEZA_BANCO[codigo_banco](df)
            except Exception as e:
                raise ValueError(f"[ERROR] Error en limpieza específica de {codigo_banco}: {e}")

        # Evitar errores por columnas duplicadas en el Excel (por ejemplo cabeceras repetidas).
        # Hacemos los nombres de columnas únicos antes del mapeo (col -> col, col_1, col_2...)
        cols = [str(c) if c is not None else '' for c in df.columns]
        seen = {}
        unique_cols = []
        for c in cols:
            if c in seen:
                seen[c] += 1
                new_c = f"{c}_{seen[c]}"
            else:
                seen[c] = 0
                new_c = c
            unique_cols.append(new_c)
        df.columns = unique_cols

    except Exception as e:
        raise ValueError(f"[ERROR] Error al cargar '{ruta}': {e}")
    
    print(f"  [LIST] Columnas encontradas: {list(df.columns)}")
    
    # Mapeo automático de columnas
    columnas_map = {}
    tiene_debito_credito = False
    for col in df.columns:
        col_upper = str(col).upper().strip()
        if any(x in col_upper for x in ['FECHA', 'DATE', 'EFECTIVA']) and 'Fecha' not in columnas_map:
            columnas_map['Fecha'] = col
        elif any(x in col_upper for x in ['CONCEPTO', 'REFERENCIA', 'CO', 'CEPTO']) and 'Concepto' not in columnas_map:
            columnas_map['Concepto'] = col
        elif any(x in col_upper for x in ['DÉBITO', 'DEBITO', 'RETIRO', 'RETIROS']) and 'Debito' not in columnas_map:
            columnas_map['Debito'] = col
        elif any(x in col_upper for x in ['CRÉDITO', 'CREDITO', 'DEPOSITO', 'DEPOSITOS']) and 'Credito' not in columnas_map:
            columnas_map['Credito'] = col
        elif any(x in col_upper for x in ['VALOR', 'MONTO', 'IMPORTE']) and 'Valor' not in columnas_map:
            columnas_map['Valor'] = col
        elif any(x in col_upper for x in ['DESCRIP', 'DETALLE', 'OBSERV', 'DESCRIPCIÓN']) and 'Descripción' not in columnas_map:
            columnas_map['Descripción'] = col

    # Fallback: detect Valor-like column with additional keywords (English variants)
    if 'Valor' not in columnas_map:
        for col in df.columns:
            cu = str(col).upper()
            if any(k in cu for k in ['VALOR', 'MONTO', 'IMPORTE', 'AMOUNT', 'AMT']):
                columnas_map['Valor'] = col
                break
    
    # Detectar si hay Debito y/o Credito
    tiene_debito_credito = ('Debito' in columnas_map) or ('Credito' in columnas_map)
    print(f"  [SEARCH] Formato detectado: {'Debito/Credito' if tiene_debito_credito else 'Valor Único'}")
    # Si no hay Descripción, crearla vacía
    if 'Descripción' not in columnas_map:
        df['Descripción'] = ''
        columnas_map['Descripción'] = 'Descripción'
        print("  [WARN] Columna 'Descripción' no encontrada - creada vacía")
    
    # Verificar columnas requeridas
    if tiene_debito_credito:
        # Requerir Fecha, Concepto y al menos una de Debito/Credito
        requeridas = ['Fecha', 'Concepto']
        if 'Debito' in columnas_map:
            requeridas.append('Debito')
        if 'Credito' in columnas_map:
            requeridas.append('Credito')
    else:
        requeridas = ['Fecha', 'Concepto', 'Valor']
    
    faltantes = [r for r in requeridas if r not in columnas_map]
    if faltantes:
        raise ValueError(f"[ERROR] Faltan columnas requeridas: {faltantes}")
    
    # Renombrar columnas - hacerlo de forma segura para evitar reindex errors
    try:
        mapping = {v: k for k, v in columnas_map.items()}
        # Construir nueva lista de columnas aplicando mapping y asegurando unicidad
        new_cols = []
        seen = {}
        for orig in list(df.columns):
            target = mapping.get(orig, orig)
            if target in seen:
                seen[target] += 1
                target = f"{target}_{seen[target]}"
            else:
                seen[target] = 0
            new_cols.append(target)
        df.columns = new_cols
    except Exception as e:
        # Proveer información diagnóstica
        msg = (
            f"Error renombrando columnas: {e}\n"
            f"Columnas actuales: {list(df.columns)}\n"
            f"Columnas mapeadas: {columnas_map}"
        )
        raise ValueError(msg)
    
    # Convertir tipos y combinar Debito/Credito si es necesario
    df['Fecha'] = pd.to_datetime(df['Fecha'], errors='coerce')
    df['Concepto'] = df['Concepto'].fillna('').astype(str)
    df['Descripción'] = df['Descripción'].fillna('').astype(str)
    
    if tiene_debito_credito:
        # Asegurar que ambas columnas existen en el DataFrame (llenar con 0 si faltan)
        if 'Debito' not in df.columns:
            df['Debito'] = 0
        if 'Credito' not in df.columns:
            df['Credito'] = 0
        # Convertir Debito y Credito a numéricos
        df['Debito'] = pd.to_numeric(df['Debito'], errors='coerce').fillna(0)
        df['Credito'] = pd.to_numeric(df['Credito'], errors='coerce').fillna(0)
        # Crear Valor: Credito positivo, Debito negativo
        df['Valor'] = df['Credito'] - df['Debito']
        # Eliminar columnas originales si existían
        drop_cols = [c for c in ['Debito', 'Credito'] if c in df.columns]
        if drop_cols:
            df = df.drop(columns=drop_cols)
    else:
        df['Valor'] = pd.to_numeric(df['Valor'], errors='coerce')
    
    # [PERF] OPTIMIZACIÓN 2: Filtrar SOLO filas con Fecha Y Valor válidos
    registros_antes = len(df)
    df = df.dropna(subset=['Fecha', 'Valor'])
    df = df[df['Valor'] != 0]  # Eliminar valores $0
    registros_despues = len(df)
    
    if registros_antes != registros_despues:
        print(f"  [WARN] Filtradas {registros_antes - registros_despues} filas sin Fecha/Valor válidos")
    
    # Campos de búsqueda
    df['Texto_Busqueda'] = (df['Concepto'].astype(str) + ' ' + df['Descripción'].astype(str)).apply(normalizar_texto)
    df['Concepto_Norm'] = df['Concepto'].apply(normalizar_texto)  # ← RESTAURADO del v5
    df['Proveedor_ID'] = df['Descripción'].apply(extraer_identificador_proveedor)
    df['Empresa_Norm'] = df['Texto_Busqueda'].apply(normalizar_nombre_empresa)
    df['Es_Impuesto'] = df['Texto_Busqueda'].apply(es_patron_impuesto)
    df['Es_Comision'] = df['Texto_Busqueda'].apply(es_patron_comision)
    df['Es_TC'] = df['Texto_Busqueda'].apply(es_patron_tc)  # ← NUEVO v6.5
    df['Es_Caja_Chica'] = df['Texto_Busqueda'].apply(es_patron_caja_chica)  # ← NUEVO v6.6
    
    # Control
    df['ID_Original'] = range(len(df))
    df['Conciliado'] = False
    
    print(f"  [OK] Cargados: {len(df)} registros válidos")
    return df

# Cargas de Archivos CONTABLE
def cargar_contable(ruta, usa_dolares, nombre=""):
    """Carga datos - OPTIMIZACIÓN: solo lee filas con Fecha Y Valor válidos"""
    print(f"\n  [DIR] Cargando: {nombre.upper() if nombre else ruta}")

    # Resolver ruta relativa: si el usuario pasa solo el nombre de archivo,
    # buscarlo en el mismo directorio del script y en el cwd.
    try:
        ruta_path = Path(ruta)
    except Exception:
        ruta_path = Path(str(ruta))

    if not ruta_path.is_absolute():
        script_dir = base_path()
        candidates = [script_dir / ruta_path, Path.cwd() / ruta_path]

        # If no suffix provided, try common extensions too
        if not ruta_path.suffix:
            exts = ['.xlsx', '.xls', '.csv']
        else:
            exts = [ruta_path.suffix]

        found = None
        for base in candidates:
            for ext in exts:
                cand = base.with_suffix(ext) if not ruta_path.suffix else base
                if cand.exists():
                    found = cand
                    break
            if found:
                break

        if found:
            ruta = str(found)
        else:
            # if the literal provided (e.g., 'BANRESERVAS.xlsx') exists in script_dir
            alt = script_dir / ruta_path.name
            if alt.exists():
                ruta = str(alt)

    try:
        # Usar la limpieza ligera para el libro contable (detecta y corta
        # filas no-datos encima de la tabla principal). Esta función lee el
        # archivo y devuelve un DataFrame ya recortado.
        moneda_str = 'USD' if usa_dolares else 'RD$'
        df = limpiar_archivo_contable(ruta, moneda_str)
    except Exception as e:
        raise ValueError(f"[ERROR] Error al cargar contable '{ruta}': {e}")
    
    print(f"  [LIST] Columnas encontradas: {list(df.columns)}")
    
    # Mapeo automático de columnas (normaliza espacios/acentos al buscar)
    columnas_map = {}
    for col in df.columns:
        col_upper = str(col).upper().strip()
        if any(x in col_upper for x in ['F.COMP', 'FECHA']) and 'Fecha' not in columnas_map:
            columnas_map['Fecha'] = col
        elif any(x in col_upper for x in ['DETALLE', 'CONCEPTO', 'CONCEP', 'OBSERV', 'REFERENCIA', 'BENEFICIARIO', 'BENEF', 'CO', 'CEPTO']) and 'Concepto' not in columnas_map:
            columnas_map['Concepto'] = col
        elif 'NATU' in col_upper and 'Natu' not in columnas_map:
            columnas_map['Natu'] = col

        # Detectar "Valor Moneda Extranjera" para USD (tiene signos ya aplicados)
        elif 'EXTRANJERA' in col_upper and 'Valor_USD' not in columnas_map:
                columnas_map['Valor_USD'] = col

        # Valor (único) - detecta nombres comunes (incluye variantes en inglés)
        elif any(x in col_upper for x in ['VALOR', 'MONTO', 'IMPORTE', 'AMOUNT', 'AMT']) and 'Valor' not in columnas_map:
            columnas_map['Valor'] = col
        
        # Descripción
        elif any(x in col_upper for x in ['DESCRIP', 'DETALLE', 'OBSERV', 'REFERENCIA', 'BENEFICIARIO', 'BENEF']) and 'Descripción' not in columnas_map:
            columnas_map['Descripción'] = col
    
    # Detectar libro contable YA LIMPIO
    es_limpio = (
        'Valor' in df.columns and
        'Natu' not in df.columns and
        'Valor_USD' not in df.columns
    )

    # Detectar si hay columna de naturaleza (Natu)
    tiene_natu = 'Natu' in columnas_map

    # Si no hay Descripción, crearla vacía
    if 'Descripción' not in columnas_map:
        df['Descripción'] = ''
        columnas_map['Descripción'] = 'Descripción'
        print("  [WARN] Columna 'Descripción' no encontrada - creada vacía")
    
    # Verificar columnas requeridas según moneda
    if usa_dolares and 'Valor_USD' in columnas_map:
        # USD: solo necesita Valor_USD
        requeridas = ['Fecha', 'Concepto', 'Valor_USD']
    elif tiene_natu:
        # RD: necesita Valor y Natu
        requeridas = ['Fecha', 'Concepto', 'Valor', 'Natu']
    else: # Chequeo de documento limpiado
        # Sin Natu: necesita solo Valor
        requeridas = ['Fecha', 'Concepto', 'Valor']
    
    faltantes = [r for r in requeridas if r not in columnas_map]
    if faltantes:
        raise ValueError(f"[ERROR] Faltan columnas requeridas: {faltantes}")
    
    # Renombrar columnas
    df = df.rename(columns={v: k for k, v in columnas_map.items()})
    
    # [PERF] OPTIMIZACIÓN 3: Eliminar filas completamente vacías (espacios en blanco bajo encabezados)
    df = df.dropna(how='all')
    df = df.reset_index(drop=True)
    
    # [BLOCK] FILTRO ESTADO: Ignorar filas marcadas como "A" o "X" en la columna Estado
    if 'Estado' in df.columns:
        registros_antes_estado = len(df)
        # Convertir Estado a string y eliminar filas con A o X
        df['Estado'] = df['Estado'].fillna('').astype(str).str.strip().str.upper()
        df = df[~df['Estado'].isin(['A', 'X'])]
        registros_eliminados_estado = registros_antes_estado - len(df)
        if registros_eliminados_estado > 0:
            print(f"  [BLOCK] Eliminadas {registros_eliminados_estado} filas con Estado 'A' o 'X'")
        df = df.drop(columns=['Estado'], errors='ignore')  # Eliminar la columna Estado después de filtrar
    
    # Convertir tipos y combinar/ajustar columnas según formato
    df['Fecha'] = pd.to_datetime(df['Fecha'], errors='coerce')
    df['Concepto'] = df['Concepto'].fillna('').astype(str)
    df['Descripción'] = df['Descripción'].fillna('').astype(str)

    # Procesar Valor según moneda detectada
    if es_limpio:
        print("  [CLEAN] Libro contable ya limpio — usando Valor directamente")
        df['Valor'] = pd.to_numeric(df['Valor'], errors='coerce')
    
    elif usa_dolares == True and 'Valor_USD' in df.columns:
        # USD: usar Valor Moneda Extranjera directamente (ya tiene signos correctos)
        print("  [USD] Detectada moneda USD - usando 'Valor Moneda Extranjera'")
        # Descartar columnas RD primero
        if 'Valor' in df.columns:
            df = df.drop(columns=['Valor'], errors='ignore')
        if 'Natu' in df.columns:
            df = df.drop(columns=['Natu'], errors='ignore')
        # Luego usar USD como Valor
        df['Valor'] = pd.to_numeric(df['Valor_USD'], errors='coerce')
        df = df.drop(columns=['Valor_USD'], errors='ignore')
    
    elif tiene_natu == True and 'Valor' in df.columns:
        # RD: usar Valor aplicando interpretación de NATU
        print("  Detectada moneda RD - usando 'Valor' interpretando 'Natu'")
        df['Valor'] = pd.to_numeric(df['Valor'], errors='coerce')

        if 'Natu' in df.columns:
            df['Natu'] = df['Natu'].astype(str).str.strip().str.upper()

            # [-] APLICAR SIGNO SEGÚN NATU
            df.loc[df['Natu'] == 'E', 'Valor'] *= -1
            df.loc[df['Natu'] == 'I', 'Valor'] = df.loc[df['Natu'] == 'I', 'Valor'].abs()

            df = df.drop(columns=['Natu'])

        # Descartar columna USD si existe
        if 'Valor_USD' in df.columns:
            df = df.drop(columns=['Valor_USD'], errors='ignore')
    
    # [PERF] OPTIMIZACIÓN 2: Filtrar SOLO filas con Fecha Y Valor válidos
    registros_antes = len(df)
    df = df.dropna(subset=['Fecha', 'Valor'])
    df = df[df['Valor'] != 0]  # Eliminar valores $0
    registros_despues = len(df)
    
    if registros_antes != registros_despues:
        print(f"  [WARN] Filtradas {registros_antes - registros_despues} filas sin Fecha/Valor válidos")
    
    # Campos de búsqueda
    df['Texto_Busqueda'] = (df['Concepto'].astype(str) + ' ' + df['Descripción'].astype(str)).apply(normalizar_texto)
    df['Concepto_Norm'] = df['Concepto'].apply(normalizar_texto)  # ← RESTAURADO del v5
    df['Proveedor_ID'] = df['Descripción'].apply(extraer_identificador_proveedor)
    df['Empresa_Norm'] = df['Texto_Busqueda'].apply(normalizar_nombre_empresa)
    df['Es_Impuesto'] = df['Texto_Busqueda'].apply(es_patron_impuesto)
    df['Es_Comision'] = df['Texto_Busqueda'].apply(es_patron_comision)
    df['Es_TC'] = df['Texto_Busqueda'].apply(es_patron_tc)  # ← NUEVO v6.5
    
    # Control
    df['ID_Original'] = range(len(df))
    df['Conciliado'] = False
    
    print(f"  [OK] Cargados: {len(df)} registros válidos")
    return df

# ============================================================================
# [TARGET] ESTRATEGIA 1: MONTO EXACTO (1:1) - CON SCORE COMBINADO
# ============================================================================

def conciliacion_por_monto_exacto(banco, contable, conciliaciones):
    print("\n" + "="*70)
    print("[TARGET] ESTRATEGIA 1: MONTO EXACTO (1:1)")
    print("="*70)
    
    contador = 0
    id_conc = len(conciliaciones) + 1
    
    for idx_b, reg_b in banco[~banco['Conciliado']].iterrows():
        f_min = reg_b['Fecha'] - timedelta(days=VENTANA_DIAS_EXACTA)
        f_max = reg_b['Fecha'] + timedelta(days=VENTANA_DIAS_EXACTA)
        
        cands = contable[
            (~contable['Conciliado']) &
            (contable['Fecha'] >= f_min) &
            (contable['Fecha'] <= f_max) &
            (abs(contable['Valor'] - reg_b['Valor']) < TOLERANCIA_VALOR_EXACTA)
        ]
        
        if len(cands) == 0:
            continue
        
        if len(cands) == 1:
            reg_c = cands.iloc[0]
            sim = calcular_similitud(reg_b['Texto_Busqueda'], reg_c['Texto_Busqueda'])
            
            if PERMITIR_SOLO_MONTO or sim >= UMBRAL_SIMILITUD_BAJA:
                conciliaciones.append({
                    'ID_Conciliacion': id_conc,
                    'Tipo': '1. Monto Exacto (1:1)',
                    'Similitud': round(sim, 3),
                    'Fecha_Banco': reg_b['Fecha'],
                    'Concepto_Banco': reg_b['Concepto'],
                    'Valor_Banco': reg_b['Valor'],
                    'Descripcion_Banco': reg_b['Descripción'],
                    'Fecha_Contable': reg_c['Fecha'],
                    'Concepto_Contable': reg_c['Concepto'],
                    'Valor_Contable': reg_c['Valor'],
                    'Descripcion_Contable': reg_c['Descripción'],
                    'Diferencia': abs(reg_b['Valor'] - reg_c['Valor'])
                })
                banco.loc[idx_b, 'Conciliado'] = True
                contable.loc[reg_c.name, 'Conciliado'] = True
                contador += 1
                id_conc += 1
        else:
            # [*] SCORE COMBINADO para desempatar (restaurado del v5)
            cands = cands.copy()
            cands['Sim'] = cands['Texto_Busqueda'].apply(
                lambda x: calcular_similitud(reg_b['Texto_Busqueda'], x)
            )
            cands['Dias_Diff'] = abs((cands['Fecha'] - reg_b['Fecha']).dt.days)
            cands['Score'] = (cands['Sim'] * 0.7) + ((VENTANA_DIAS_EXACTA - cands['Dias_Diff']) / VENTANA_DIAS_EXACTA * 0.3)
            
            mejor_idx = cands['Score'].idxmax()
            mejor_sim = cands.loc[mejor_idx, 'Sim']
            
            if PERMITIR_SOLO_MONTO or mejor_sim >= UMBRAL_SIMILITUD_BAJA:
                reg_c = contable.loc[mejor_idx]
                conciliaciones.append({
                    'ID_Conciliacion': id_conc,
                    'Tipo': '1. Monto Exacto (1:1)',
                    'Similitud': round(mejor_sim, 3),
                    'Fecha_Banco': reg_b['Fecha'],
                    'Concepto_Banco': reg_b['Concepto'],
                    'Valor_Banco': reg_b['Valor'],
                    'Descripcion_Banco': reg_b['Descripción'],
                    'Fecha_Contable': reg_c['Fecha'],
                    'Concepto_Contable': reg_c['Concepto'],
                    'Valor_Contable': reg_c['Valor'],
                    'Descripcion_Contable': reg_c['Descripción'],
                    'Diferencia': abs(reg_b['Valor'] - reg_c['Valor'])
                })
                banco.loc[idx_b, 'Conciliado'] = True
                contable.loc[mejor_idx, 'Conciliado'] = True
                contador += 1
                id_conc += 1
    
    print(f"[v] Conciliaciones 1:1: {contador}")
    return contador

# ============================================================================
# [TARGET] ESTRATEGIA 1.5: TRANSFERENCIAS CON COMISIÓN ($7)
# ============================================================================

def conciliacion_con_comisiones(banco, contable, conciliaciones):
    if not DETECTAR_COMISIONES:
        return 0
    
    print("\n" + "="*70)
    print("[TARGET] ESTRATEGIA 1.5: TRANSFERENCIAS CON COMISIÓN ($7)")
    print("="*70)
    
    contador = 0
    id_conc = len(conciliaciones) + 1 if conciliaciones else 1
    
    banco_no_conc = banco[~banco['Conciliado']].copy()
    
    for idx_com, reg_com in banco_no_conc[banco_no_conc['Es_Comision']].iterrows():
        if abs(abs(reg_com['Valor']) - COMISION_TRANSFERENCIA_USD) > 0.50:
            continue
        
        fecha_com = reg_com['Fecha']
        candidatos_transf = banco_no_conc[
            (~banco_no_conc['Conciliado']) &
            (banco_no_conc['Fecha'] == fecha_com) &
            (banco_no_conc['Valor'] > 0) &
            (banco_no_conc.index != idx_com)
        ]
        
        for idx_transf, reg_transf in candidatos_transf.iterrows():
            valor_neto = reg_transf['Valor'] - COMISION_TRANSFERENCIA_USD
            
            f_min = fecha_com - timedelta(days=VENTANA_DIAS_EXACTA)
            f_max = fecha_com + timedelta(days=VENTANA_DIAS_EXACTA)
            
            cands_cont = contable[
                (~contable['Conciliado']) &
                (contable['Fecha'] >= f_min) &
                (contable['Fecha'] <= f_max) &
                (abs(contable['Valor'] - valor_neto) < TOLERANCIA_VALOR_EXACTA)
            ]
            
            if len(cands_cont) >= 1:
                reg_c = cands_cont.iloc[0]
                sim = calcular_similitud(reg_transf['Texto_Busqueda'], reg_c['Texto_Busqueda'])
                
                conciliaciones.append({
                    'ID_Conciliacion': id_conc,
                    'Tipo': '1.5. Transf+Comisión (2→1)',
                    'Similitud': round(sim, 3),
                    'Fecha_Banco': reg_com['Fecha'],
                    'Concepto_Banco': reg_com['Concepto'],
                    'Valor_Banco': reg_com['Valor'],
                    'Descripcion_Banco': reg_com['Descripción'],
                    'Fecha_Contable': None,
                    'Concepto_Contable': '(Comisión bancaria)',
                    'Valor_Contable': None,
                    'Descripcion_Contable': '',
                    'Diferencia': None
                })
                
                conciliaciones.append({
                    'ID_Conciliacion': id_conc,
                    'Tipo': '1.5. Transf+Comisión (2→1)',
                    'Similitud': round(sim, 3),
                    'Fecha_Banco': reg_transf['Fecha'],
                    'Concepto_Banco': reg_transf['Concepto'],
                    'Valor_Banco': reg_transf['Valor'],
                    'Descripcion_Banco': reg_transf['Descripción'],
                    'Fecha_Contable': reg_c['Fecha'],
                    'Concepto_Contable': reg_c['Concepto'],
                    'Valor_Contable': reg_c['Valor'],
                    'Descripcion_Contable': reg_c['Descripción'],
                    'Diferencia': COMISION_TRANSFERENCIA_USD
                })
                
                banco.loc[idx_com, 'Conciliado'] = True
                banco.loc[idx_transf, 'Conciliado'] = True
                contable.loc[reg_c.name, 'Conciliado'] = True
                
                contador += 1
                id_conc += 1
                break
    
    print(f"[v] Transferencias con comisión: {contador}")
    return contador

# ============================================================================
# [TARGET] ESTRATEGIA 1.6: COMISIONES AGRUPADAS MULTI-FECHA [v6.1]
# ============================================================================

def conciliacion_comisiones_agrupadas(banco, contable, conciliaciones):
    print("\n" + "="*70)
    print("[TARGET] ESTRATEGIA 1.6: COMISIONES AGRUPADAS (Multi-fecha)")
    print("="*70)
    
    contador = 0
    id_conc = len(conciliaciones) + 1 if conciliaciones else 1
    
    contable_com = contable[(~contable['Conciliado']) & (contable['Es_Comision'])].copy()
    
    if len(contable_com) == 0:
        print("(0) No hay comisiones pendientes en contable")
        return 0
    
    banco_com = banco[(~banco['Conciliado']) & (banco['Es_Comision'])].copy()
    
    if len(banco_com) == 0:
        print("(0) No hay comisiones pendientes en banco")
        return 0
    
    for idx_c, reg_c in contable_com.iterrows():
        if contable.loc[idx_c, 'Conciliado']:
            continue
        
        valor_objetivo = reg_c['Valor']
        
        banco_com_pend = banco_com[~banco_com.index.isin(banco[banco['Conciliado']].index)]
        if len(banco_com_pend) == 0:
            continue
        
        # Ventana amplia para comisiones dispersas
        mes_contable = reg_c['Fecha'].month if pd.notna(reg_c['Fecha']) else None
        if mes_contable:
            banco_com_mes = banco_com_pend[
                (banco_com_pend['Fecha'].dt.month >= mes_contable - 1) &
                (banco_com_pend['Fecha'].dt.month <= mes_contable + 1)
            ]
        else:
            banco_com_mes = banco_com_pend
        
        if len(banco_com_mes) == 0:
            continue
        
        # Búsqueda combinatoria LIMITADA
        mejor_combo, mejor_diff = None, float('inf')
        indices = banco_com_mes.index.tolist()[:MAX_PARTIDAS_AGRUPACION]
        
        total_combinaciones = 0
        for r in range(1, len(indices) + 1):
            if mejor_diff < TOLERANCIA_VALOR_AGRUPACION:
                break
            if total_combinaciones >= MAX_COMBINACIONES_POR_BUSQUEDA:
                break
            
            for combo in combinations(indices, r):
                total_combinaciones += 1
                if total_combinaciones > MAX_COMBINACIONES_POR_BUSQUEDA:
                    break
                
                suma = banco.loc[list(combo), 'Valor'].sum()
                diff = abs(suma - valor_objetivo)
                
                if diff < mejor_diff:
                    mejor_diff = diff
                    mejor_combo = combo
                
                if diff < TOLERANCIA_VALOR_AGRUPACION:
                    break
        
        if mejor_combo and mejor_diff < TOLERANCIA_VALOR_AGRUPACION:
            grupo = banco.loc[list(mejor_combo)]
            
            for i, (idx_b, reg_b) in enumerate(grupo.sort_values(['Fecha', 'Valor']).iterrows()):
                conciliaciones.append({
                    'ID_Conciliacion': id_conc,
                    'Tipo': f'1.6. Comisiones Agrupadas ({len(grupo)}→1)',
                    'Similitud': 1.0,
                    'Fecha_Banco': reg_b['Fecha'],
                    'Concepto_Banco': reg_b['Concepto'],
                    'Valor_Banco': reg_b['Valor'],
                    'Descripcion_Banco': reg_b['Descripción'],
                    'Fecha_Contable': reg_c['Fecha'] if i == 0 else None,
                    'Concepto_Contable': reg_c['Concepto'] if i == 0 else '',
                    'Valor_Contable': reg_c['Valor'] if i == 0 else None,
                    'Descripcion_Contable': reg_c['Descripción'] if i == 0 else '',
                    'Diferencia': mejor_diff if i == 0 else None
                })
            
            banco.loc[list(mejor_combo), 'Conciliado'] = True
            contable.loc[idx_c, 'Conciliado'] = True
            contador += 1
            id_conc += 1
    
    print(f"[v] Comisiones agrupadas: {contador}")
    return contador

# ============================================================================
# [TARGET] ESTRATEGIA 2: N→1 INTELIGENTE (4 MÉTODOS) - RESTAURADO DEL v5
# ============================================================================

def conciliacion_n_a_1_inteligente(banco, contable, conciliaciones):
    print("\n" + "="*70)
    print("[TARGET] ESTRATEGIA 2: N→1 (Varios Contable → Uno Banco)")
    print("="*70)
    
    contador = 0
    id_conc = len(conciliaciones) + 1 if conciliaciones else 1
    
    for idx_b, reg_b in banco[~banco['Conciliado']].iterrows():
        if banco.loc[idx_b, 'Conciliado']:
            continue
            
        f_min = reg_b['Fecha'] - timedelta(days=VENTANA_DIAS_AGRUPACION)
        f_max = reg_b['Fecha'] + timedelta(days=VENTANA_DIAS_AGRUPACION)
        
        cands = contable[
            (~contable['Conciliado']) &
            (contable['Fecha'] >= f_min) &
            (contable['Fecha'] <= f_max)
        ].copy()
        
        if len(cands) < 2:
            continue
        
        encontrado = False
        
        # MÉTODO 1: Por Proveedor_ID con desambiguación por fechas
        if USAR_FECHAS_PARA_DESAMBIGUAR and 'Proveedor_ID' in cands.columns:
            for prov_id in cands['Proveedor_ID'].dropna().unique():
                if encontrado:
                    break
                    
                grupo_prov = cands[cands['Proveedor_ID'] == prov_id]
                if len(grupo_prov) < 2:
                    continue
                
                for fecha_grupo in grupo_prov['Fecha'].unique():
                    grupo = grupo_prov[
                        (grupo_prov['Fecha'] >= fecha_grupo - timedelta(days=3)) &
                        (grupo_prov['Fecha'] <= fecha_grupo + timedelta(days=3))
                    ]
                    
                    if len(grupo) < 2:
                        continue
                    
                    suma = grupo['Valor'].sum()
                    if abs(suma - reg_b['Valor']) < TOLERANCIA_VALOR_AGRUPACION:
                        sim_media = grupo['Texto_Busqueda'].apply(
                            lambda x: calcular_similitud(reg_b['Texto_Busqueda'], x)
                        ).mean()
                        
                        for i, (_, reg_c) in enumerate(grupo.sort_values(['Fecha', 'Valor']).iterrows()):
                            conciliaciones.append({
                                'ID_Conciliacion': id_conc,
                                'Tipo': f'2. Agrupación N→1 ({len(grupo)} reg)',
                                'Similitud': round(sim_media, 3),
                                'Fecha_Banco': reg_b['Fecha'] if i == 0 else None,
                                'Concepto_Banco': reg_b['Concepto'] if i == 0 else '',
                                'Valor_Banco': reg_b['Valor'] if i == 0 else None,
                                'Descripcion_Banco': reg_b['Descripción'] if i == 0 else '',
                                'Fecha_Contable': reg_c['Fecha'],
                                'Concepto_Contable': reg_c['Concepto'],
                                'Valor_Contable': reg_c['Valor'],
                                'Descripcion_Contable': reg_c['Descripción'],
                                'Diferencia': abs(reg_b['Valor'] - suma) if i == 0 else None
                            })
                        
                        banco.loc[idx_b, 'Conciliado'] = True
                        contable.loc[grupo.index, 'Conciliado'] = True
                        contador += 1
                        id_conc += 1
                        encontrado = True
                        break
        
        if encontrado:
            continue
        
        # [*] MÉTODO 2: Por Concepto_Norm (RESTAURADO del v5)
        for concepto in cands['Concepto_Norm'].unique():
            if encontrado or not concepto:
                break
            
            grupo = cands[cands['Concepto_Norm'] == concepto]
            if len(grupo) < 2:
                continue
            
            suma = grupo['Valor'].sum()
            if abs(suma - reg_b['Valor']) < TOLERANCIA_VALOR_AGRUPACION:
                sim_media = grupo['Texto_Busqueda'].apply(
                    lambda x: calcular_similitud(reg_b['Texto_Busqueda'], x)
                ).mean()
                
                for i, (_, reg_c) in enumerate(grupo.sort_values(['Fecha', 'Valor']).iterrows()):
                    conciliaciones.append({
                        'ID_Conciliacion': id_conc,
                        'Tipo': f'2. Agrupación N→1 ({len(grupo)} reg)',
                        'Similitud': round(sim_media, 3),
                        'Fecha_Banco': reg_b['Fecha'] if i == 0 else None,
                        'Concepto_Banco': reg_b['Concepto'] if i == 0 else '',
                        'Valor_Banco': reg_b['Valor'] if i == 0 else None,
                        'Descripcion_Banco': reg_b['Descripción'] if i == 0 else '',
                        'Fecha_Contable': reg_c['Fecha'],
                        'Concepto_Contable': reg_c['Concepto'],
                        'Valor_Contable': reg_c['Valor'],
                        'Descripcion_Contable': reg_c['Descripción'],
                        'Diferencia': abs(reg_b['Valor'] - suma) if i == 0 else None
                    })
                
                banco.loc[idx_b, 'Conciliado'] = True
                contable.loc[grupo.index, 'Conciliado'] = True
                contador += 1
                id_conc += 1
                encontrado = True
        
        if encontrado:
            continue
        
        # MÉTODO 3: Por empresa normalizada
        for emp_norm in cands['Empresa_Norm'].unique():
            if encontrado or not emp_norm:
                break
                
            grupo = cands[cands['Empresa_Norm'] == emp_norm]
            if len(grupo) < 2:
                continue
            
            suma = grupo['Valor'].sum()
            if abs(suma - reg_b['Valor']) < TOLERANCIA_VALOR_AGRUPACION:
                sim_media = grupo['Texto_Busqueda'].apply(
                    lambda x: calcular_similitud(reg_b['Texto_Busqueda'], x)
                ).mean()
                
                for i, (_, reg_c) in enumerate(grupo.sort_values(['Fecha', 'Valor']).iterrows()):
                    conciliaciones.append({
                        'ID_Conciliacion': id_conc,
                        'Tipo': f'2. Agrupación N→1 ({len(grupo)} reg)',
                        'Similitud': round(sim_media, 3),
                        'Fecha_Banco': reg_b['Fecha'] if i == 0 else None,
                        'Concepto_Banco': reg_b['Concepto'] if i == 0 else '',
                        'Valor_Banco': reg_b['Valor'] if i == 0 else None,
                        'Descripcion_Banco': reg_b['Descripción'] if i == 0 else '',
                        'Fecha_Contable': reg_c['Fecha'],
                        'Concepto_Contable': reg_c['Concepto'],
                        'Valor_Contable': reg_c['Valor'],
                        'Descripcion_Contable': reg_c['Descripción'],
                        'Diferencia': abs(reg_b['Valor'] - suma) if i == 0 else None
                    })
                
                banco.loc[idx_b, 'Conciliado'] = True
                contable.loc[grupo.index, 'Conciliado'] = True
                contador += 1
                id_conc += 1
                encontrado = True
        
        if encontrado:
            continue
        
        # MÉTODO 4: Por palabras clave
        cands['PalKey'] = cands['Texto_Busqueda'].apply(
            lambda x: ' '.join(sorted(extraer_palabras_clave(x)))
        )
        
        for key, grupo in cands.groupby('PalKey'):
            if len(grupo) < 2 or not key:
                continue
            
            suma = grupo['Valor'].sum()
            if abs(suma - reg_b['Valor']) < TOLERANCIA_VALOR_AGRUPACION:
                sim_media = grupo['Texto_Busqueda'].apply(
                    lambda x: calcular_similitud(reg_b['Texto_Busqueda'], x)
                ).mean()
                
                if sim_media >= UMBRAL_SIMILITUD_BAJA:
                    for i, (_, reg_c) in enumerate(grupo.sort_values(['Fecha', 'Valor']).iterrows()):
                        conciliaciones.append({
                            'ID_Conciliacion': id_conc,
                            'Tipo': f'2. Agrupación N→1 ({len(grupo)} reg)',
                            'Similitud': round(sim_media, 3),
                            'Fecha_Banco': reg_b['Fecha'] if i == 0 else None,
                            'Concepto_Banco': reg_b['Concepto'] if i == 0 else '',
                            'Valor_Banco': reg_b['Valor'] if i == 0 else None,
                            'Descripcion_Banco': reg_b['Descripción'] if i == 0 else '',
                            'Fecha_Contable': reg_c['Fecha'],
                            'Concepto_Contable': reg_c['Concepto'],
                            'Valor_Contable': reg_c['Valor'],
                            'Descripcion_Contable': reg_c['Descripción'],
                            'Diferencia': abs(reg_b['Valor'] - suma) if i == 0 else None
                        })
                    
                    banco.loc[idx_b, 'Conciliado'] = True
                    contable.loc[grupo.index, 'Conciliado'] = True
                    contador += 1
                    id_conc += 1
                    break
    
    print(f"[v] Agrupaciones N→1: {contador}")
    return contador

# ============================================================================
# [TARGET] ESTRATEGIA 3: 1→N
# ============================================================================

def conciliacion_1_a_n_inteligente(banco, contable, conciliaciones):
    print("\n" + "="*70)
    print("[TARGET] ESTRATEGIA 3: 1→N (Uno Contable → Varios Banco)")
    print("="*70)
    
    contador = 0
    id_conc = len(conciliaciones) + 1 if conciliaciones else 1
    
    for idx_c, reg_c in contable[~contable['Conciliado']].iterrows():
        if contable.loc[idx_c, 'Conciliado']:
            continue
            
        f_min = reg_c['Fecha'] - timedelta(days=VENTANA_DIAS_AGRUPACION)
        f_max = reg_c['Fecha'] + timedelta(days=VENTANA_DIAS_AGRUPACION)
        
        cands = banco[
            (~banco['Conciliado']) &
            (banco['Fecha'] >= f_min) &
            (banco['Fecha'] <= f_max)
        ].copy()
        
        if len(cands) < 2:
            continue
        
        cands['PalKey'] = cands['Texto_Busqueda'].apply(
            lambda x: ' '.join(sorted(extraer_palabras_clave(x)))
        )
        
        for key, grupo in cands.groupby('PalKey'):
            if len(grupo) < 2 or not key:
                continue
            
            suma = grupo['Valor'].sum()
            if abs(suma - reg_c['Valor']) < TOLERANCIA_VALOR_AGRUPACION:
                sim_media = grupo['Texto_Busqueda'].apply(
                    lambda x: calcular_similitud(reg_c['Texto_Busqueda'], x)
                ).mean()
                
                if sim_media >= UMBRAL_SIMILITUD_BAJA:
                    for i, (_, reg_b) in enumerate(grupo.sort_values(['Fecha', 'Valor']).iterrows()):
                        conciliaciones.append({
                            'ID_Conciliacion': id_conc,
                            'Tipo': f'3. Agrupación 1→N ({len(grupo)} reg)',
                            'Similitud': round(sim_media, 3),
                            'Fecha_Banco': reg_b['Fecha'],
                            'Concepto_Banco': reg_b['Concepto'],
                            'Valor_Banco': reg_b['Valor'],
                            'Descripcion_Banco': reg_b['Descripción'],
                            'Fecha_Contable': reg_c['Fecha'] if i == 0 else None,
                            'Concepto_Contable': reg_c['Concepto'] if i == 0 else '',
                            'Valor_Contable': reg_c['Valor'] if i == 0 else None,
                            'Descripcion_Contable': reg_c['Descripción'] if i == 0 else '',
                            'Diferencia': abs(suma - reg_c['Valor']) if i == 0 else None
                        })
                    
                    contable.loc[idx_c, 'Conciliado'] = True
                    banco.loc[grupo.index, 'Conciliado'] = True
                    contador += 1
                    id_conc += 1
                    break
    
    print(f"[v] Agrupaciones 1→N: {contador}")
    return contador

# ============================================================================
# [TARGET] ESTRATEGIA 4: N↔M
# ============================================================================

def conciliacion_n_a_m_inteligente(banco, contable, conciliaciones):
    print("\n" + "="*70)
    print("[TARGET] ESTRATEGIA 4: N↔M (Varios ↔ Varios)")
    print("="*70)
    
    contador = 0
    id_conc = len(conciliaciones) + 1 if conciliaciones else 1
    
    cont_temp = contable[~contable['Conciliado']].copy()
    cont_temp['PalKey'] = cont_temp['Texto_Busqueda'].apply(
        lambda x: ' '.join(sorted(extraer_palabras_clave(x)))
    )
    grupos_c = cont_temp.groupby('PalKey').filter(lambda x: len(x) >= 2)
    
    for key_c, grupo_c in grupos_c.groupby('PalKey'):
        if len(grupo_c) < 2 or not key_c:
            continue
        
        grupo_c = grupo_c[~grupo_c.index.isin(contable[contable['Conciliado']].index)]
        if len(grupo_c) < 2:
            continue
        
        suma_c = grupo_c['Valor'].sum()
        f_media = grupo_c['Fecha'].mean()
        f_min = f_media - timedelta(days=VENTANA_DIAS_AGRUPACION)
        f_max = f_media + timedelta(days=VENTANA_DIAS_AGRUPACION)
        
        cands_b = banco[
            (~banco['Conciliado']) &
            (banco['Fecha'] >= f_min) &
            (banco['Fecha'] <= f_max)
        ].copy()
        
        if len(cands_b) < 2:
            continue
        
        cands_b['PalKey'] = cands_b['Texto_Busqueda'].apply(
            lambda x: ' '.join(sorted(extraer_palabras_clave(x)))
        )
        
        for key_b, grupo_b in cands_b.groupby('PalKey'):
            if len(grupo_b) < 2 or not key_b:
                continue
            
            suma_b = grupo_b['Valor'].sum()
            
            if abs(suma_b - suma_c) < TOLERANCIA_VALOR_AGRUPACION:
                sim_total = sum(
                    calcular_similitud(rb['Texto_Busqueda'], rc['Texto_Busqueda'])
                    for _, rb in grupo_b.iterrows()
                    for _, rc in grupo_c.iterrows()
                )
                sim_media = sim_total / (len(grupo_b) * len(grupo_c))
                
                if sim_media >= UMBRAL_SIMILITUD_BAJA:
                    regs_b = grupo_b.sort_values(['Fecha', 'Valor'])
                    regs_c = grupo_c.sort_values(['Fecha', 'Valor'])
                    max_regs = max(len(regs_b), len(regs_c))
                    
                    for i in range(max_regs):
                        rb = regs_b.iloc[i] if i < len(regs_b) else None
                        rc = regs_c.iloc[i] if i < len(regs_c) else None
                        
                        conciliaciones.append({
                            'ID_Conciliacion': id_conc,
                            'Tipo': f'4. Agrupación N↔M ({len(regs_b)}↔{len(regs_c)})',
                            'Similitud': round(sim_media, 3),
                            'Fecha_Banco': rb['Fecha'] if rb is not None else None,
                            'Concepto_Banco': rb['Concepto'] if rb is not None else '',
                            'Valor_Banco': rb['Valor'] if rb is not None else None,
                            'Descripcion_Banco': rb['Descripción'] if rb is not None else '',
                            'Fecha_Contable': rc['Fecha'] if rc is not None else None,
                            'Concepto_Contable': rc['Concepto'] if rc is not None else '',
                            'Valor_Contable': rc['Valor'] if rc is not None else None,
                            'Descripcion_Contable': rc['Descripción'] if rc is not None else '',
                            'Diferencia': abs(suma_b - suma_c) if i == 0 else None
                        })
                    
                    banco.loc[regs_b.index, 'Conciliado'] = True
                    contable.loc[regs_c.index, 'Conciliado'] = True
                    contador += 1
                    id_conc += 1
                    break
    
    print(f"[v] Agrupaciones N↔M: {contador}")
    return contador

# ============================================================================
# [TARGET] ESTRATEGIA 5: IMPUESTOS DGII
# ============================================================================

def conciliacion_impuestos(banco, contable, conciliaciones):
    print("\n" + "="*70)
    print("[TARGET] ESTRATEGIA 5: IMPUESTOS DGII (0.15%)")
    print("="*70)
    
    contador = 0
    id_conc = len(conciliaciones) + 1 if conciliaciones else 1
    
    banco_imp = banco[(~banco['Conciliado']) & (banco['Es_Impuesto'])].copy()
    
    if len(banco_imp) == 0:
        print("(0) No hay impuestos pendientes en banco")
        return 0
    
    contable_imp = contable[
        (~contable['Conciliado']) &
        (contable['Es_Impuesto'])
    ].copy()
    
    if len(contable_imp) == 0:
        print("(0) No hay impuestos pendientes en contable")
        return 0
    
    banco_imp['Mes'] = banco_imp['Fecha'].dt.to_period('M')
    
    for mes, grupo_banco in banco_imp.groupby('Mes'):
        suma_banco = abs(grupo_banco['Valor'].sum())
        
        for idx_c, reg_c in contable_imp.iterrows():
            if contable.loc[idx_c, 'Conciliado']:
                continue
            
            if abs(abs(reg_c['Valor']) - suma_banco) < TOLERANCIA_VALOR_AGRUPACION:
                for i, (idx_b, reg_b) in enumerate(grupo_banco.iterrows()):
                    conciliaciones.append({
                        'ID_Conciliacion': id_conc,
                        'Tipo': f'5. Impuestos DGII ({len(grupo_banco)}→1)',
                        'Similitud': 1.0,
                        'Fecha_Banco': reg_b['Fecha'],
                        'Concepto_Banco': reg_b['Concepto'],
                        'Valor_Banco': reg_b['Valor'],
                        'Descripcion_Banco': reg_b['Descripción'],
                        'Fecha_Contable': reg_c['Fecha'] if i == 0 else None,
                        'Concepto_Contable': reg_c['Concepto'] if i == 0 else '',
                        'Valor_Contable': reg_c['Valor'] if i == 0 else None,
                        'Descripcion_Contable': reg_c['Descripción'] if i == 0 else '',
                        'Diferencia': abs(suma_banco - abs(reg_c['Valor'])) if i == 0 else None
                    })
                
                banco.loc[grupo_banco.index, 'Conciliado'] = True
                contable.loc[idx_c, 'Conciliado'] = True
                contador += 1
                id_conc += 1
                break
    
    print(f"[v] Impuestos conciliados: {contador}")
    return contador

# ============================================================================
# [TARGET] ESTRATEGIA 6: SEGUNDA PASADA FLEXIBLE - OPTIMIZADA
# ============================================================================

def segunda_pasada_inteligente(banco, contable, conciliaciones):
    if not EJECUTAR_SEGUNDA_PASADA:
        return 0
    
    print("\n" + "="*70)
    print("[TARGET] ESTRATEGIA 6: SEGUNDA PASADA (Búsqueda Flexible)")
    print("="*70)
    
    contador = 0
    id_conc = len(conciliaciones) + 1 if conciliaciones else 1
    
    for idx_b, reg_b in banco[~banco['Conciliado']].iterrows():
        if banco.loc[idx_b, 'Conciliado']:
            continue
        
        f_min = reg_b['Fecha'] - timedelta(days=VENTANA_DIAS_FLEXIBLE)
        f_max = reg_b['Fecha'] + timedelta(days=VENTANA_DIAS_FLEXIBLE)
        
        cands = contable[
            (~contable['Conciliado']) &
            (contable['Fecha'] >= f_min) &
            (contable['Fecha'] <= f_max)
        ].copy()
        
        if len(cands) == 0:
            continue
        
        valor_objetivo = reg_b['Valor']
        mejor_combo, mejor_diff = None, float('inf')
        
        # [PERF] OPTIMIZACIÓN: Limitar partidas y combinaciones
        max_items = min(len(cands), MAX_PARTIDAS_AGRUPACION)
        cands_limitados = cands.head(max_items)
        
        total_combinaciones = 0
        for r in range(2, min(len(cands_limitados) + 1, 8)):
            if mejor_diff < TOLERANCIA_VALOR_AGRUPACION:
                break
            if total_combinaciones >= MAX_COMBINACIONES_POR_BUSQUEDA:
                break
            
            for combo in combinations(cands_limitados.index, r):
                total_combinaciones += 1
                if total_combinaciones > MAX_COMBINACIONES_POR_BUSQUEDA:
                    break
                
                suma = contable.loc[list(combo), 'Valor'].sum()
                diff = abs(suma - valor_objetivo)
                
                if diff < mejor_diff:
                    mejor_diff = diff
                    mejor_combo = combo
                
                if diff < TOLERANCIA_VALOR_AGRUPACION:
                    break
        
        if mejor_combo and mejor_diff < TOLERANCIA_VALOR_AGRUPACION:
            grupo = contable.loc[list(mejor_combo)]
            suma = grupo['Valor'].sum()
            
            sim_media = grupo['Texto_Busqueda'].apply(
                lambda x: calcular_similitud(reg_b['Texto_Busqueda'], x)
            ).mean()
            
            for i, (_, reg_c) in enumerate(grupo.sort_values(['Fecha', 'Valor']).iterrows()):
                conciliaciones.append({
                    'ID_Conciliacion': id_conc,
                    'Tipo': f'6. Segunda Pasada ({len(grupo)} reg)',
                    'Similitud': round(sim_media, 3),
                    'Fecha_Banco': reg_b['Fecha'] if i == 0 else None,
                    'Concepto_Banco': reg_b['Concepto'] if i == 0 else '',
                    'Valor_Banco': reg_b['Valor'] if i == 0 else None,
                    'Descripcion_Banco': reg_b['Descripción'] if i == 0 else '',
                    'Fecha_Contable': reg_c['Fecha'],
                    'Concepto_Contable': reg_c['Concepto'],
                    'Valor_Contable': reg_c['Valor'],
                    'Descripcion_Contable': reg_c['Descripción'],
                    'Diferencia': mejor_diff if i == 0 else None
                })
            
            banco.loc[idx_b, 'Conciliado'] = True
            contable.loc[list(mejor_combo), 'Conciliado'] = True
            contador += 1
            id_conc += 1
    
    print(f"[v] Segunda pasada: {contador}")
    return contador

# ============================================================================
# [TARGET] ESTRATEGIA 7: BÚSQUEDA EXHAUSTIVA FINAL [v6.1] - OPTIMIZADA
# ============================================================================

def busqueda_exhaustiva_final(banco, contable, conciliaciones):
    if not EJECUTAR_BUSQUEDA_EXHAUSTIVA:
        return 0
    
    print("\n" + "="*70)
    print("[TARGET] ESTRATEGIA 7: BÚSQUEDA EXHAUSTIVA (Sin restricción fechas)")
    print("="*70)
    
    contador = 0
    id_conc = len(conciliaciones) + 1 if conciliaciones else 1
    
    banco_pend = banco[~banco['Conciliado']].copy()
    contable_pend = contable[~contable['Conciliado']].copy()
    
    n_banco_pend = len(banco_pend)
    n_contable_pend = len(contable_pend)
    
    print(f"  [STATS] Pendientes: Banco={n_banco_pend}, Contable={n_contable_pend}")
    
    if n_banco_pend > UMBRAL_PARTIDAS_EXHAUSTIVA or n_contable_pend > UMBRAL_PARTIDAS_EXHAUSTIVA:
        print(f"  (0) Demasiadas partidas (umbral={UMBRAL_PARTIDAS_EXHAUSTIVA})")
        return 0
    
    if n_banco_pend == 0 or n_contable_pend == 0:
        print("  (0) No hay partidas en ambos lados")
        return 0
    
    # CASO 1: Buscar N contable → 1 banco
    if n_contable_pend <= 10:
        print(f"\n  [SEARCH] Buscando: Banco → Contable...")
        for idx_c, reg_c in contable_pend.iterrows():
            if contable.loc[idx_c, 'Conciliado']:
                continue
            
            valor_objetivo = reg_c['Valor']
            banco_disponible = banco[~banco['Conciliado']]
            
            if len(banco_disponible) == 0:
                continue
            
            mejor_combo, mejor_diff = None, float('inf')
            indices = banco_disponible.index.tolist()[:MAX_PARTIDAS_AGRUPACION]
            
            total_comb = 0
            for r in range(1, len(indices) + 1):
                if mejor_diff < TOLERANCIA_VALOR_AGRUPACION or total_comb > MAX_COMBINACIONES_EXHAUSTIVA:
                    break
                
                for combo in combinations(indices, r):
                    total_comb += 1
                    if total_comb > MAX_COMBINACIONES_EXHAUSTIVA:
                        break
                    
                    suma = banco.loc[list(combo), 'Valor'].sum()
                    diff = abs(suma - valor_objetivo)
                    
                    if diff < mejor_diff:
                        mejor_diff = diff
                        mejor_combo = combo
                    
                    if diff < TOLERANCIA_VALOR_AGRUPACION:
                        break
            
            if mejor_combo and mejor_diff < TOLERANCIA_VALOR_AGRUPACION:
                grupo = banco.loc[list(mejor_combo)]
                print(f"    [v] {len(grupo)} banco → 1 contable (diff=${mejor_diff:.2f})")
                
                for i, (idx_b, reg_b) in enumerate(grupo.sort_values(['Fecha', 'Valor']).iterrows()):
                    conciliaciones.append({
                        'ID_Conciliacion': id_conc,
                        'Tipo': f'7. Exhaustiva ({len(grupo)}→1)',
                        'Similitud': 0.0,
                        'Fecha_Banco': reg_b['Fecha'],
                        'Concepto_Banco': reg_b['Concepto'],
                        'Valor_Banco': reg_b['Valor'],
                        'Descripcion_Banco': reg_b['Descripción'],
                        'Fecha_Contable': reg_c['Fecha'] if i == 0 else None,
                        'Concepto_Contable': reg_c['Concepto'] if i == 0 else '',
                        'Valor_Contable': reg_c['Valor'] if i == 0 else None,
                        'Descripcion_Contable': reg_c['Descripción'] if i == 0 else '',
                        'Diferencia': mejor_diff if i == 0 else None
                    })
                
                banco.loc[list(mejor_combo), 'Conciliado'] = True
                contable.loc[idx_c, 'Conciliado'] = True
                contador += 1
                id_conc += 1
    
    # CASO 2: Buscar N banco → 1 contable
    banco_pend = banco[~banco['Conciliado']].copy()
    if len(banco_pend) <= 10 and len(banco_pend) > 0:
        print(f"\n  [SEARCH] Buscando: Contable → Banco...")
        for idx_b, reg_b in banco_pend.iterrows():
            if banco.loc[idx_b, 'Conciliado']:
                continue
            
            valor_objetivo = reg_b['Valor']
            contable_disponible = contable[~contable['Conciliado']]
            
            if len(contable_disponible) == 0:
                continue
            
            mejor_combo, mejor_diff = None, float('inf')
            indices = contable_disponible.index.tolist()[:MAX_PARTIDAS_AGRUPACION]
            
            total_comb = 0
            for r in range(1, len(indices) + 1):
                if mejor_diff < TOLERANCIA_VALOR_AGRUPACION or total_comb > MAX_COMBINACIONES_EXHAUSTIVA:
                    break
                
                for combo in combinations(indices, r):
                    total_comb += 1
                    if total_comb > MAX_COMBINACIONES_EXHAUSTIVA:
                        break
                    
                    suma = contable.loc[list(combo), 'Valor'].sum()
                    diff = abs(suma - valor_objetivo)
                    
                    if diff < mejor_diff:
                        mejor_diff = diff
                        mejor_combo = combo
                    
                    if diff < TOLERANCIA_VALOR_AGRUPACION:
                        break
            
            if mejor_combo and mejor_diff < TOLERANCIA_VALOR_AGRUPACION:
                grupo = contable.loc[list(mejor_combo)]
                print(f"    [v] 1 banco → {len(grupo)} contable (diff=${mejor_diff:.2f})")
                
                for i, (_, reg_c) in enumerate(grupo.sort_values(['Fecha', 'Valor']).iterrows()):
                    conciliaciones.append({
                        'ID_Conciliacion': id_conc,
                        'Tipo': f'7. Exhaustiva (1→{len(grupo)})',
                        'Similitud': 0.0,
                        'Fecha_Banco': reg_b['Fecha'] if i == 0 else None,
                        'Concepto_Banco': reg_b['Concepto'] if i == 0 else '',
                        'Valor_Banco': reg_b['Valor'] if i == 0 else None,
                        'Descripcion_Banco': reg_b['Descripción'] if i == 0 else '',
                        'Fecha_Contable': reg_c['Fecha'],
                        'Concepto_Contable': reg_c['Concepto'],
                        'Valor_Contable': reg_c['Valor'],
                        'Descripcion_Contable': reg_c['Descripción'],
                        'Diferencia': mejor_diff if i == 0 else None
                    })
                
                banco.loc[idx_b, 'Conciliado'] = True
                contable.loc[list(mejor_combo), 'Conciliado'] = True
                contador += 1
                id_conc += 1
    
    print(f"\n[v] Búsqueda exhaustiva: {contador}")
    return contador

# ============================================================================
# [TOOL] FUNCIONES AUXILIARES PARA CAJA CHICA (AÑADIR DESPUÉS DE LAS FUNCIONES DE NORMALIZACIÓN)
# ============================================================================

def es_patron_caja_chica(texto):
    """Detecta si el texto corresponde a reposición de caja chica"""
    texto_norm = normalizar_texto(texto)
    return any(patron in texto_norm for patron in PATRONES_CAJA_CHICA)

def extraer_nombre_persona(texto):
    """
    Extrae el nombre completo de una persona del texto
    Busca patrones como: "TRANSFERENCIA A JUAN PEREZ" o "MIGUEL ANGEL SANTIAGO"
    """
    texto_norm = normalizar_texto(texto)
    if not texto_norm:
        return None
    
    # Patrón 1: "TRANSFERENCIA A [NOMBRE]"
    match = re.search(r'(?:TRANSFERENCIA|TRANSF|PAGO|DEBITO)\s+(?:A|PARA|DE)\s+([A-Z\s]{6,})', texto_norm)
    if match:
        nombre = match.group(1).strip()
        # Eliminar tokens que no son parte del nombre (ej. REPOSICION, CAJA, CHICA)
        ruido_tokens = set()
        for pat in PATRONES_CAJA_CHICA:
            for w in pat.split():
                if len(w) >= 3:
                    ruido_tokens.add(w)

        palabras = [p for p in nombre.split() if len(p) >= 3 and p not in PALABRAS_COMUNES and p not in ruido_tokens]
        if len(palabras) >= 2:
            return ' '.join(palabras)
    
    # Patrón 2: Buscar nombres completos (2-4 palabras capitalizadas consecutivas)
    palabras = texto_norm.split()
    nombres_encontrados = []
    
    for i in range(len(palabras) - 1):
        # Buscar secuencias de 2-4 palabras que parezcan nombres
        for longitud in [4, 3, 2]:  # Empezar por las más largas
            if i + longitud <= len(palabras):
                candidato = palabras[i:i+longitud]
                # Filtrar palabras comunes
                candidato_limpio = [p for p in candidato if len(p) >= 3 and p not in PALABRAS_COMUNES]
                
                if len(candidato_limpio) >= 2:
                    # También eliminar tokens de ruido de caja chica
                    ruido_tokens = set()
                    for pat in PATRONES_CAJA_CHICA:
                        for w in pat.split():
                            if len(w) >= 3:
                                ruido_tokens.add(w)

                    candidato_final = [p for p in candidato_limpio if p not in ruido_tokens]
                    if len(candidato_final) >= 2:
                        nombre_completo = ' '.join(candidato_final)
                        # Validar que parezca un nombre (longitud razonable)
                        if 10 <= len(nombre_completo) <= 50:
                            nombres_encontrados.append(nombre_completo)
    
    # Retornar el nombre más largo encontrado (generalmente es el más completo)
    if nombres_encontrados:
        return max(nombres_encontrados, key=len)
    
    return None

def similitud_nombres(nombre1, nombre2):
    """
    Calcula similitud entre dos nombres de persona
    Retorna score de 0.0 a 1.0
    """
    if not nombre1 or not nombre2:
        return 0.0
    
    n1 = normalizar_texto(nombre1)
    n2 = normalizar_texto(nombre2)
    
    palabras1 = set(n1.split())
    palabras2 = set(n2.split())
    
    # Si las palabras principales coinciden, es muy probable que sea la misma persona
    coincidencias = len(palabras1 & palabras2)
    max_palabras = max(len(palabras1), len(palabras2))
    
    if max_palabras == 0:
        return 0.0
    
    return coincidencias / max_palabras

# ============================================================================
# [TARGET] ESTRATEGIA 8: REPOSICIÓN DE CAJA CHICA [NUEVA v6.6]
# ============================================================================

def conciliacion_caja_chica(banco, contable, conciliaciones):
    """
    Estrategia especializada para reposiciones de caja chica
    
    LÓGICA:
    1. Detecta registros de banco con patrón "REPOSICIÓN DE CAJA CHICA"
    2. Extrae el nombre de la persona del banco
    3. Busca TODAS las partidas contables con ese mismo nombre
    4. Agrupa por suma de montos (los conceptos pueden ser diversos)
    5. Concilia si la suma coincide con la reposición del banco
    """
    print("\n" + "="*70)
    print("[TARGET] ESTRATEGIA 8: REPOSICIÓN DE CAJA CHICA")
    print("="*70)
    
    contador = 0
    id_conc = len(conciliaciones) + 1 if conciliaciones else 1
    
    # Buscar registros de caja chica en banco
    banco_caja_chica = banco[
        (~banco['Conciliado']) & 
        (banco['Texto_Busqueda'].apply(es_patron_caja_chica))
    ].copy()
    
    if len(banco_caja_chica) == 0:
        print("(0) No hay reposiciones de caja chica pendientes en banco")
        return 0
    
    print(f"  [LIST] Encontradas {len(banco_caja_chica)} reposiciones de caja chica en banco")
    
    for idx_b, reg_b in banco_caja_chica.iterrows():
        if banco.loc[idx_b, 'Conciliado']:
            continue
        
        # PASO 1: Extraer nombre de persona del banco
        nombre_banco = extraer_nombre_persona(reg_b['Texto_Busqueda'])
        
        if not nombre_banco:
            print(f"  [WARN]  No se pudo extraer nombre de persona de: '{reg_b['Concepto']}'")
            continue
        
        print(f"\n  [SEARCH] Buscando gastos de: {nombre_banco}")
        print(f"      Banco: {reg_b['Concepto']} - ${reg_b['Valor']:,.2f}")
        
        # PASO 2: Buscar TODAS las partidas contables con ese nombre
        # Ventana de tiempo amplia (generalmente las reposiciones cubren varios días)
        f_min = reg_b['Fecha'] - timedelta(days=VENTANA_DIAS_AGRUPACION)
        f_max = reg_b['Fecha'] + timedelta(days=VENTANA_DIAS_AGRUPACION)
        
        # Excluir partidas que claramente son comisiones o impuestos;
        # además filtrar por presencia del nombre extraído en el Texto_Busqueda
        candidatos = contable[
            (~contable['Conciliado']) &
            (~contable.get('Es_Comision', False)) &
            (~contable.get('Es_Impuesto', False)) &
            (contable['Fecha'] >= f_min) &
            (contable['Fecha'] <= f_max) &
            (contable['Texto_Busqueda'].str.contains(nombre_banco))
        ].copy()
        
        if len(candidatos) == 0:
            continue
        
        # Buscar partidas que contengan el nombre de la persona
        partidas_persona = []
        # Diagnóstico: comparar con todas las partidas del periodo que contienen el nombre (sin excluir flags)
        todas_partidas_nombre = []
        for idx_c_all, reg_c_all in contable[(contable['Fecha'] >= f_min) & (contable['Fecha'] <= f_max)].iterrows():
            # Incluir cualquier partida del periodo que contenga el nombre en Texto_Busqueda (no filtrar por flags aquí)
            if nombre_banco and nombre_banco in reg_c_all['Texto_Busqueda']:
                todas_partidas_nombre.append({'index': idx_c_all, 'registro': reg_c_all, 'Es_Comision': reg_c_all.get('Es_Comision', False), 'Es_Impuesto': reg_c_all.get('Es_Impuesto', False)})

        if len(todas_partidas_nombre) != 0:
            suma_todas = sum(r['registro']['Valor'] for r in todas_partidas_nombre)
            excluidas = [r for r in todas_partidas_nombre if r['Es_Comision'] or r['Es_Impuesto']]
            suma_excluidas = sum(r['registro']['Valor'] for r in excluidas) if excluidas else 0
            print(f"      (DEBUG) Partidas totales con nombre en ventana: {len(todas_partidas_nombre)}; suma={suma_todas:,.2f}; excluidas por flag: {len(excluidas)} suma_excluidas={suma_excluidas:,.2f}")
            # Mostrar detalles de las partidas totales (índice, valor, concepto, flags)
            for r in todas_partidas_nombre:
                reg = r['registro']
                print(f"        (DBG-TOT) idx={r['index']} valor={reg['Valor']:,.2f} flag_com={r['Es_Comision']} flag_imp={r['Es_Impuesto']} concept='{reg['Concepto'][:50]}' desc='{reg['Descripción'][:40]}'")
            if excluidas:
                print("        (DBG-TOT) Excluidas:")
                for e in excluidas:
                    reg = e['registro']
                    print(f"          - idx={e['index']} valor={reg['Valor']:,.2f} concept='{reg['Concepto'][:50]}' desc='{reg['Descripción'][:40]}'")
        for idx_c, reg_c in candidatos.iterrows():
            # Asegurar que no se consideren partidas marcadas como comisión/impuesto
            if reg_c.get('Es_Comision', False) or reg_c.get('Es_Impuesto', False):
                continue

            # Si el nombre extraído del banco aparece literalmente en Texto_Busqueda, asumir match directo
            if nombre_banco and nombre_banco in reg_c['Texto_Busqueda']:
                partidas_persona.append({
                    'index': idx_c,
                    'registro': reg_c,
                    'similitud_nombre': 1.0
                })
                continue

            nombre_contable = extraer_nombre_persona(reg_c['Texto_Busqueda'])

            if nombre_contable:
                sim = similitud_nombres(nombre_banco, nombre_contable)

                # Si hay alta similitud en nombres, es la misma persona
                if sim >= 0.6:  # Al menos 60% de coincidencia
                    partidas_persona.append({
                        'index': idx_c,
                        'registro': reg_c,
                        'similitud_nombre': sim
                    })

        # Diagnóstico: imprimir suma de partidas_persona seleccionadas
        if partidas_persona:
            suma_seleccionadas = sum(p['registro']['Valor'] for p in partidas_persona)
            print(f"      (DEBUG) Partidas seleccionadas (sin flags): {len(partidas_persona)}; suma={suma_seleccionadas:,.2f}")
        
        if len(partidas_persona) == 0:
            print(f"      [ERROR] No se encontraron gastos de '{nombre_banco}' en contable")
            continue
        
        # PASO 3: Agrupar todas las partidas de esa persona
        indices_grupo = [p['index'] for p in partidas_persona]
        grupo = contable.loc[indices_grupo]
        
        suma_contable = grupo['Valor'].sum()
        diferencia = abs(reg_b['Valor'] - suma_contable)
        
        print(f"      [v] Encontradas {len(grupo)} partidas:")
        for _, reg in grupo.head(5).iterrows():  # Mostrar solo las primeras 5
            print(f"        • {reg['Concepto'][:40]:40} ${reg['Valor']:>10,.2f}")
        if len(grupo) > 5:
            print(f"        ... y {len(grupo) - 5} partidas más")
        print(f"      [STATS] Suma contable: ${suma_contable:,.2f}")
        print(f"      [STATS] Diferencia:    ${diferencia:,.2f}")
        
        # PASO 4: Conciliar si la suma coincide
        if diferencia < TOLERANCIA_VALOR_AGRUPACION:
            print(f"      [OK] CONCILIADO - Diferencia aceptable: ${diferencia:.2f}")
            
            # Registrar conciliación
            for i, (idx_c, reg_c) in enumerate(grupo.sort_values(['Fecha', 'Valor']).iterrows()):
                conciliaciones.append({
                    'ID_Conciliacion': id_conc,
                    'Tipo': f'8. Caja Chica ({len(grupo)}→1)',
                    'Similitud': round(partidas_persona[0]['similitud_nombre'], 3),
                    'Fecha_Banco': reg_b['Fecha'] if i == 0 else None,
                    'Concepto_Banco': reg_b['Concepto'] if i == 0 else '',
                    'Valor_Banco': reg_b['Valor'] if i == 0 else None,
                    'Descripcion_Banco': reg_b['Descripción'] if i == 0 else '',
                    'Fecha_Contable': reg_c['Fecha'],
                    'Concepto_Contable': reg_c['Concepto'],
                    'Valor_Contable': reg_c['Valor'],
                    'Descripcion_Contable': reg_c['Descripción'],
                    'Diferencia': diferencia if i == 0 else None
                })
            
            # Marcar como conciliadas
            banco.loc[idx_b, 'Conciliado'] = True
            contable.loc[indices_grupo, 'Conciliado'] = True
            
            contador += 1
            id_conc += 1
        else:
            # Intentar conciliar sumando TODAS las reposiciones de banco para la misma persona
            # (caso: varias transferencias que en conjunto cubren los gastos)
            similares_banco = []
            for idx_bb, reg_bb in banco_caja_chica.iterrows():
                if banco.loc[idx_bb, 'Conciliado']:
                    continue
                nombre_bb = extraer_nombre_persona(reg_bb['Texto_Busqueda'])
                if nombre_bb and similitud_nombres(nombre_banco, nombre_bb) >= 0.6:
                    similares_banco.append({'index': idx_bb, 'registro': reg_bb})

            if similares_banco:
                suma_bancos = sum(b['registro']['Valor'] for b in similares_banco)
                diff_grupal = abs(suma_bancos - suma_contable)
                print(f"      (DEBUG) Suma reposiciones banco para {nombre_banco}: {suma_bancos:,.2f}; diff_grupal={diff_grupal:,.2f}")
                if diff_grupal < TOLERANCIA_VALOR_AGRUPACION:
                    print(f"      [OK] CONCILIADO COMO GRUPO - {len(similares_banco)} reposiciones suman {suma_bancos:,.2f}")
                    # Registrar conciliaciones: map each banco y cada contable registro
                    for i_b, binfo in enumerate(sorted(similares_banco, key=lambda x: x['registro']['Fecha'])):
                        for i_c, (idx_c, reg_c) in enumerate(grupo.sort_values(['Fecha', 'Valor']).iterrows()):
                            conciliaciones.append({
                                'ID_Conciliacion': id_conc,
                                'Tipo': f'8. Caja Chica (Grupo {len(similares_banco)}→{len(grupo)})',
                                'Similitud': round(partidas_persona[0]['similitud_nombre'], 3) if partidas_persona else 0.0,
                                'Fecha_Banco': binfo['registro']['Fecha'] if i_c == 0 and i_b == 0 else None,
                                'Concepto_Banco': binfo['registro']['Concepto'] if i_c == 0 and i_b == 0 else '',
                                'Valor_Banco': binfo['registro']['Valor'] if i_c == 0 and i_b == 0 else None,
                                'Descripcion_Banco': binfo['registro']['Descripción'] if i_c == 0 and i_b == 0 else '',
                                'Fecha_Contable': reg_c['Fecha'] if i_b == 0 else None,
                                'Concepto_Contable': reg_c['Concepto'] if i_b == 0 else '',
                                'Valor_Contable': reg_c['Valor'] if i_b == 0 else None,
                                'Descripcion_Contable': reg_c['Descripción'] if i_b == 0 else '',
                                'Diferencia': diff_grupal if i_b == 0 and i_c == 0 else None
                            })

                    # Marcar como conciliadas
                    banco.loc[[b['index'] for b in similares_banco], 'Conciliado'] = True
                    contable.loc[indices_grupo, 'Conciliado'] = True
                    contador += 1
                    id_conc += 1
                else:
                    print(f"      [WARN]  Diferencia muy alta: ${diferencia:.2f} > ${TOLERANCIA_VALOR_AGRUPACION:.2f}")
            else:
                print(f"      [WARN]  Diferencia muy alta: ${diferencia:.2f} > ${TOLERANCIA_VALOR_AGRUPACION:.2f}")
    
    print(f"\n[v] Reposiciones de caja chica conciliadas: {contador}")
    return contador

# ============================================================================
# [SEARCH] DETECCIÓN DE CASOS ESPECIALES
# ============================================================================

def detectar_casos_especiales(banco, contable):
    print("\n" + "="*70)
    print("[SEARCH] DETECTANDO CASOS ESPECIALES")
    print("="*70)
    
    casos_especiales = []
    
    if not DETECTAR_CASOS_ESPECIALES:
        print("(0) Detección desactivada")
        return casos_especiales
    
    banco_pend = banco[~banco['Conciliado']].copy()
    contable_pend = contable[~contable['Conciliado']].copy()
    
    for idx_b, reg_b in banco_pend.iterrows():
        for prov_id in contable_pend['Proveedor_ID'].dropna().unique():
            grupo = contable_pend[contable_pend['Proveedor_ID'] == prov_id]
            if len(grupo) < 1:
                continue
            
            suma = grupo['Valor'].sum()
            
            if reg_b['Valor'] != 0:
                porcentaje = abs(suma) / abs(reg_b['Valor']) * 100
            else:
                continue
            
            if 80 <= porcentaje < 100:
                diferencia = abs(reg_b['Valor']) - abs(suma)
                
                sim = calcular_similitud(
                    reg_b['Texto_Busqueda'],
                    ' '.join(grupo['Texto_Busqueda'].tolist())
                )
                
                if sim > 0.1:
                    casos_especiales.append({
                        'Tipo': 'PARCIAL',
                        'Descripcion': f'{len(grupo)} partidas suman {porcentaje:.1f}% del banco',
                        'Banco_Fecha': reg_b['Fecha'],
                        'Banco_Concepto': reg_b['Concepto'],
                        'Banco_Valor': reg_b['Valor'],
                        'Banco_Descripcion': reg_b['Descripción'],
                        'Contable_Registros': len(grupo),
                        'Contable_Suma': suma,
                        'Diferencia_Pendiente': diferencia,
                        'Porcentaje_Conciliado': round(porcentaje, 1),
                        'Similitud': round(sim, 3)
                    })
                    break
        
        mes_banco = reg_b['Fecha'].month if pd.notna(reg_b['Fecha']) else None
        
        if mes_banco:
            otros_meses = contable_pend[
                contable_pend['Fecha'].dt.month != mes_banco
            ]
            
            for _, reg_c in otros_meses.iterrows():
                if abs(reg_c['Valor'] - reg_b['Valor']) < TOLERANCIA_VALOR_EXACTA:
                    sim = calcular_similitud(reg_b['Texto_Busqueda'], reg_c['Texto_Busqueda'])
                    
                    if sim > 0.2:
                        casos_especiales.append({
                            'Tipo': 'PERIODO_DIFERENTE',
                            'Descripcion': f'Monto exacto pero mes diferente',
                            'Banco_Fecha': reg_b['Fecha'],
                            'Banco_Concepto': reg_b['Concepto'],
                            'Banco_Valor': reg_b['Valor'],
                            'Banco_Descripcion': reg_b['Descripción'],
                            'Contable_Registros': 1,
                            'Contable_Suma': reg_c['Valor'],
                            'Diferencia_Pendiente': 0,
                            'Porcentaje_Conciliado': 100.0,
                            'Similitud': round(sim, 3)
                        })
    
    casos_unicos = []
    valores_vistos = set()
    for caso in casos_especiales:
        key = (caso['Banco_Valor'], caso['Contable_Suma'])
        if key not in valores_vistos:
            valores_vistos.add(key)
            casos_unicos.append(caso)
    
    print(f"[WARN] Casos especiales detectados: {len(casos_unicos)}")
    return casos_unicos

# ============================================================================
# [STYLE] FORMATO EXCEL PROFESIONAL
# ============================================================================

def aplicar_formato_profesional(ruta):
    """Aplica formato moderno y profesional al Excel"""
    wb = openpyxl.load_workbook(ruta)
    
    C_PRI = "1A237E"
    C_SEC = "3F51B5"
    C_EXI = "2E7D32"
    C_ERR = "C62828"
    C_ADV = "EF6C00"
    C_GRI = "FAFAFA"
    C_GRI2 = "EEEEEE"
    
    f_tit = Font(name='Segoe UI', size=22, bold=True, color="FFFFFF")
    f_sub = Font(name='Segoe UI', size=12, italic=True, color="757575")
    f_enc = Font(name='Segoe UI', size=11, bold=True, color="FFFFFF")
    f_sec = Font(name='Segoe UI', size=13, bold=True, color=C_SEC)
    f_nor = Font(name='Segoe UI', size=10)
    f_num = Font(name='Segoe UI', size=10, bold=True)
    
    r_pri = PatternFill(start_color=C_PRI, end_color=C_PRI, fill_type="solid")
    r_sec = PatternFill(start_color=C_SEC, end_color=C_SEC, fill_type="solid")
    r_exi = PatternFill(start_color=C_EXI, end_color=C_EXI, fill_type="solid")
    r_err = PatternFill(start_color=C_ERR, end_color=C_ERR, fill_type="solid")
    r_adv = PatternFill(start_color=C_ADV, end_color=C_ADV, fill_type="solid")
    r_gri = PatternFill(start_color=C_GRI, end_color=C_GRI, fill_type="solid")
    r_gri2 = PatternFill(start_color=C_GRI2, end_color=C_GRI2, fill_type="solid")
    
    tipos_r = {
        '1': PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid"),
        '2': PatternFill(start_color="BBDEFB", end_color="BBDEFB", fill_type="solid"),
        '3': PatternFill(start_color="FFE082", end_color="FFE082", fill_type="solid"),
        '4': PatternFill(start_color="FFCCBC", end_color="FFCCBC", fill_type="solid"),
        '5': PatternFill(start_color="E1BEE7", end_color="E1BEE7", fill_type="solid"),
        '6': PatternFill(start_color="B2DFDB", end_color="B2DFDB", fill_type="solid"),
        '7': PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid"),
        '8': PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid"),  # Amarillo claro
    }
    
    borde = Border(
        left=Side(style='thin', color="BDBDBD"),
        right=Side(style='thin', color="BDBDBD"),
        top=Side(style='thin', color="BDBDBD"),
        bottom=Side(style='thin', color="BDBDBD")
    )
    
    if 'RESUMEN' in wb.sheetnames:
        ws = wb['RESUMEN']
        ws.insert_rows(1, 3)
        
        ws.merge_cells('A1:B1')
        ws['A1'].value = "[STATS] CONCILIACIÓN BANCARIA v6.5 DEFINITIVO"
        ws['A1'].font = f_tit
        ws['A1'].fill = r_pri
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 50
        
        ws.merge_cells('A2:B2')
        ws['A2'].value = "GREEN PARK - BANRESERBAS"
        ws['A2'].font = Font(name='Segoe UI', size=14, bold=True, color="424242")
        ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[2].height = 28
        
        ws.merge_cells('A3:B3')
        ws['A3'].value = f"[DATE] Generado: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
        ws['A3'].font = f_sub
        ws['A3'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[3].height = 24
        
        for r in range(4, ws.max_row + 1):
            for c in [1, 2]:
                cell = ws.cell(r, c)
                cell.border = borde
                cell.font = f_nor
                
                if c == 1 and cell.value and isinstance(cell.value, str):
                    if '═══' in str(cell.value):
                        cell.font = f_sec
                        cell.fill = r_gri2
                    elif '───' in str(cell.value):
                        cell.font = Font(name='Segoe UI', size=11, bold=True, color=C_SEC)
                        cell.fill = r_gri
                
                if c == 2 and isinstance(cell.value, (int, float)):
                    cell.font = f_num
                    cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                    if cell.value < 0:
                        cell.font = Font(name='Segoe UI', size=10, bold=True, color=C_ERR)
                    elif cell.value > 0:
                        cell.font = Font(name='Segoe UI', size=10, bold=True, color=C_EXI)
        
        ws.column_dimensions['A'].width = 62
        ws.column_dimensions['B'].width = 35
    
    if 'CONCILIADOS' in wb.sheetnames:
        ws = wb['CONCILIADOS']
        if ws.max_row > 1:
            ws.insert_rows(1)
            max_c = ws.max_column
            
            ws.merge_cells(f'A1:{get_column_letter(max_c)}1')
            ws['A1'].value = "[OK] PARTIDAS CONCILIADAS"
            ws['A1'].font = f_tit
            ws['A1'].fill = r_exi
            ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
            ws.row_dimensions[1].height = 45
            
            for c in range(1, max_c + 1):
                cell = ws.cell(2, c)
                cell.fill = r_sec
                cell.font = f_enc
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = borde
            ws.row_dimensions[2].height = 38
            
            for r in range(3, ws.max_row + 1):
                tipo_val = str(ws.cell(r, 2).value) if ws.cell(r, 2).value else ""
                relleno = None
                
                for t_num, t_fill in tipos_r.items():
                    if tipo_val.startswith(f'{t_num}.'):
                        relleno = t_fill
                        break
                
                if not relleno and r % 2 == 0:
                    relleno = r_gri
                
                for c in range(1, max_c + 1):
                    cell = ws.cell(r, c)
                    if relleno:
                        cell.fill = relleno
                    cell.border = borde
                    cell.font = f_nor
                    cell.alignment = Alignment(vertical='center', wrap_text=True)
                    
                    hdr = ws.cell(2, c)
                    if hdr.value:
                        if 'Fecha' in str(hdr.value) and isinstance(cell.value, datetime):
                            cell.number_format = 'DD/MM/YYYY'
                        elif ('Valor' in str(hdr.value) or 'Diferencia' in str(hdr.value)):
                            if isinstance(cell.value, (int, float)):
                                cell.number_format = '#,##0.00'
                                color = C_ERR if cell.value < 0 else C_EXI
                                cell.font = Font(name='Segoe UI', size=10, bold=True, color=color)
            
            for c in range(1, max_c + 1):
                col_letter = get_column_letter(c)
                max_len = 0
                for r in range(1, min(ws.max_row + 1, 100)):
                    val = ws.cell(r, c).value
                    if val:
                        max_len = max(max_len, len(str(val)))
                ws.column_dimensions[col_letter].width = min(max_len + 4, 55)
    
    if 'NO_CONCILIADOS' in wb.sheetnames:
        ws = wb['NO_CONCILIADOS']
        if ws.max_row > 1:
            ws.insert_rows(1)
            max_c = ws.max_column
            
            ws.merge_cells(f'A1:{get_column_letter(max_c)}1')
            ws['A1'].value = "[WARN] PARTIDAS PENDIENTES DE CONCILIACIÓN"
            ws['A1'].font = f_tit
            ws['A1'].fill = r_err
            ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
            ws.row_dimensions[1].height = 45
            
            for c in range(1, max_c + 1):
                cell = ws.cell(2, c)
                cell.font = f_enc
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = borde
                if cell.value and 'Banco' in str(cell.value):
                    cell.fill = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
                elif cell.value and 'Contable' in str(cell.value):
                    cell.fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
                else:
                    cell.fill = r_sec
            ws.row_dimensions[2].height = 38
            
            for r in range(3, ws.max_row + 1):
                relleno_fila = r_gri if r % 2 == 0 else None
                for c in range(1, max_c + 1):
                    cell = ws.cell(r, c)
                    if relleno_fila:
                        cell.fill = relleno_fila
                    cell.border = borde
                    cell.font = f_nor
                    cell.alignment = Alignment(vertical='center', wrap_text=True)
                    
                    hdr = ws.cell(2, c)
                    if hdr.value:
                        if 'Fecha' in str(hdr.value) and isinstance(cell.value, datetime):
                            cell.number_format = 'DD/MM/YYYY'
                        elif 'Valor' in str(hdr.value):
                            if isinstance(cell.value, (int, float)):
                                cell.number_format = '#,##0.00'
                                color = C_ERR if cell.value < 0 else C_EXI
                                cell.font = Font(name='Segoe UI', size=10, bold=True, color=color)
            
            for c in range(1, max_c + 1):
                col_letter = get_column_letter(c)
                max_len = 0
                for r in range(1, min(ws.max_row + 1, 100)):
                    val = ws.cell(r, c).value
                    if val:
                        max_len = max(max_len, len(str(val)))
                ws.column_dimensions[col_letter].width = min(max_len + 4, 55)
    
    if 'CASOS_ESPECIALES' in wb.sheetnames:
        ws = wb['CASOS_ESPECIALES']
        if ws.max_row > 1:
            ws.insert_rows(1)
            max_c = ws.max_column
            
            ws.merge_cells(f'A1:{get_column_letter(max_c)}1')
            ws['A1'].value = "[SEARCH] CASOS ESPECIALES - Requieren Revisión Manual"
            ws['A1'].font = f_tit
            ws['A1'].fill = r_adv
            ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
            ws.row_dimensions[1].height = 45
            
            for c in range(1, max_c + 1):
                cell = ws.cell(2, c)
                cell.fill = r_sec
                cell.font = f_enc
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = borde
            ws.row_dimensions[2].height = 38
            
            for r in range(3, ws.max_row + 1):
                for c in range(1, max_c + 1):
                    cell = ws.cell(r, c)
                    cell.border = borde
                    cell.font = f_nor
                    cell.fill = PatternFill(start_color="FFF8E1", end_color="FFF8E1", fill_type="solid")
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = '#,##0.00'
                        cell.font = f_num
            
            for c in range(1, max_c + 1):
                col_letter = get_column_letter(c)
                max_len = 0
                for r in range(1, min(ws.max_row + 1, 100)):
                    val = ws.cell(r, c).value
                    if val:
                        max_len = max(max_len, len(str(val)))
                ws.column_dimensions[col_letter].width = min(max_len + 4, 55)
    
    wb.save(ruta)
    print("✨ Formato profesional aplicado exitosamente")

# ============================================================================
# 🚀 FUNCIÓN PRINCIPAL
# ============================================================================

def ejecutar_conciliacion(arg_dir=None, arg_bank=None, arg_ledger=None, arg_currency=None, **kwargs):
    """Función principal de conciliación - VERSIÓN MULTI-PARA-C#"""
    import time
    tiempo_inicio = time.time()
    
    # Sobrescribir parámetros globales si se pasan por kwargs
    global TOLERANCIA_VALOR_EXACTA, TOLERANCIA_VALOR_AGRUPACION, TOLERANCIA_PORCENTAJE_PARCIAL
    global VENTANA_DIAS_EXACTA, VENTANA_DIAS_AGRUPACION, VENTANA_DIAS_FLEXIBLE, VENTANA_DIAS_COMISIONES
    global UMBRAL_SIMILITUD_BAJA, UMBRAL_SIMILITUD_MEDIA, UMBRAL_SIMILITUD_ALTA
    global PERMITIR_SOLO_MONTO, USAR_FECHAS_PARA_DESAMBIGUAR, DETECTAR_CASOS_ESPECIALES
    global APLICAR_FORMATO_PROFESIONAL, EJECUTAR_SEGUNDA_PASADA, EJECUTAR_BUSQUEDA_EXHAUSTIVA
    global COMISION_TRANSFERENCIA_USD, DETECTAR_COMISIONES
    global MAX_PARTIDAS_AGRUPACION, MAX_COMBINACIONES_POR_BUSQUEDA
    global UMBRAL_PARTIDAS_EXHAUSTIVA, MAX_COMBINACIONES_EXHAUSTIVA

    if 'tol_exacta' in kwargs and kwargs['tol_exacta'] is not None: TOLERANCIA_VALOR_EXACTA = float(kwargs['tol_exacta'])
    if 'tol_agrup' in kwargs and kwargs['tol_agrup'] is not None: TOLERANCIA_VALOR_AGRUPACION = float(kwargs['tol_agrup'])
    if 'tol_parcial' in kwargs and kwargs['tol_parcial'] is not None: TOLERANCIA_PORCENTAJE_PARCIAL = float(kwargs['tol_parcial'])
    if 'ven_exacta' in kwargs and kwargs['ven_exacta'] is not None: VENTANA_DIAS_EXACTA = int(kwargs['ven_exacta'])
    if 'ven_agrup' in kwargs and kwargs['ven_agrup'] is not None: VENTANA_DIAS_AGRUPACION = int(kwargs['ven_agrup'])
    if 'ven_flex' in kwargs and kwargs['ven_flex'] is not None: VENTANA_DIAS_FLEXIBLE = int(kwargs['ven_flex'])
    if 'ven_comis' in kwargs and kwargs['ven_comis'] is not None: VENTANA_DIAS_COMISIONES = int(kwargs['ven_comis'])
    if 'umb_baja' in kwargs and kwargs['umb_baja'] is not None: UMBRAL_SIMILITUD_BAJA = float(kwargs['umb_baja'])
    if 'umb_media' in kwargs and kwargs['umb_media'] is not None: UMBRAL_SIMILITUD_MEDIA = float(kwargs['umb_media'])
    if 'umb_alta' in kwargs and kwargs['umb_alta'] is not None: UMBRAL_SIMILITUD_ALTA = float(kwargs['umb_alta'])
    if 'solo_monto' in kwargs and kwargs['solo_monto'] is not None: PERMITIR_SOLO_MONTO = kwargs['solo_monto']
    if 'usar_fechas' in kwargs and kwargs['usar_fechas'] is not None: USAR_FECHAS_PARA_DESAMBIGUAR = kwargs['usar_fechas']
    if 'especiales' in kwargs and kwargs['especiales'] is not None: DETECTAR_CASOS_ESPECIALES = kwargs['especiales']
    if 'profesional' in kwargs and kwargs['profesional'] is not None: APLICAR_FORMATO_PROFESIONAL = kwargs['profesional']
    if 'segunda_pasada' in kwargs and kwargs['segunda_pasada'] is not None: EJECUTAR_SEGUNDA_PASADA = kwargs['segunda_pasada']
    if 'exhaustiva' in kwargs and kwargs['exhaustiva'] is not None: EJECUTAR_BUSQUEDA_EXHAUSTIVA = kwargs['exhaustiva']
    if 'comision_usd' in kwargs and kwargs['comision_usd'] is not None: COMISION_TRANSFERENCIA_USD = float(kwargs['comision_usd'])
    if 'det_comis' in kwargs and kwargs['det_comis'] is not None: DETECTAR_COMISIONES = kwargs['det_comis']
    if 'max_partidas' in kwargs and kwargs['max_partidas'] is not None: MAX_PARTIDAS_AGRUPACION = int(kwargs['max_partidas'])
    if 'max_comb' in kwargs and kwargs['max_comb'] is not None: MAX_COMBINACIONES_POR_BUSQUEDA = int(kwargs['max_comb'])
    if 'umb_exh' in kwargs and kwargs['umb_exh'] is not None: UMBRAL_PARTIDAS_EXHAUSTIVA = int(kwargs['umb_exh'])
    if 'max_exh' in kwargs and kwargs['max_exh'] is not None: MAX_COMBINACIONES_EXHAUSTIVA = int(kwargs['max_exh'])
    
    print("\n" + "="*70)
    print("  [SYSTEM] Motor: v6.5 + Integracion C# ERP")
    print("="*70)
    print(f"\nFecha: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")
    
    # [NEW] NUEVO: Lógica de selección de fideicomiso / carpetas
    if arg_dir:
        print(f"🚀 Ejecutando en modo HEADLESS (pasan argumentos)")
        carpeta_caso = arg_dir
    else:
        # Selección interactiva (fallback)
        carpeta_caso = seleccionar_fideicomiso()
    
    if not carpeta_caso:
        print("\n[ERROR] No se seleccionó ningún fideicomiso. Saliendo...")
        return
    
    # Configurar rutas para el caso seleccionado
    configurar_rutas_caso(carpeta_caso)
    
    # Cambiar al directorio del caso
    os.chdir(CARPETA_TRABAJO)
    
    print("\n[CONFIG] PARÁMETROS:")
    print("─"*70)
    print(f"  [MONEY] Tolerancia exacta: ${TOLERANCIA_VALOR_EXACTA}")
    print(f"  [MONEY] Tolerancia agrupación: ${TOLERANCIA_VALOR_AGRUPACION}")
    print(f"  [DATE] Ventanas: ±{VENTANA_DIAS_EXACTA}d / ±{VENTANA_DIAS_AGRUPACION}d / ±{VENTANA_DIAS_FLEXIBLE}d")
    print(f"  [PERF] Límites: {MAX_PARTIDAS_AGRUPACION} partidas / {MAX_COMBINACIONES_POR_BUSQUEDA:,} comb")
    print(f"  [TOOL] Segunda pasada: {'Sí' if EJECUTAR_SEGUNDA_PASADA else 'No'}")
    print(f"  [TOOL] Búsqueda exhaustiva: {'Sí' if EJECUTAR_BUSQUEDA_EXHAUSTIVA else 'No'}")
    
    print("\n[BANK] BANCOS:")
    print("─"*70)
    print(f"   BanReservas")
    print(f"   Popular")
    print(f"   BHD")
    print(f"   Santa Cruz")

    print("\n" + "="*70)
    print("[DIR] CARGANDO DATOS")
    print("="*70)
    
    # Búsqueda de archivos
    if arg_bank and arg_ledger:
        archivo_banco = arg_bank
        archivo_contable = arg_ledger
        
        # Detectar banco por nombre de archivo
        nombre_file_banco = os.path.basename(archivo_banco)
        codigo_banco, nombre_banco = detectar_banco_en_nombre_archivo(nombre_file_banco)
        if not codigo_banco:
            codigo_banco = "BANRESERVAS" # Fallback
            nombre_banco = "BanReservas"
    else:
        archivo_banco, archivo_contable, codigo_banco, nombre_banco = buscar_archivos_en_carpeta()
    
    if not archivo_banco or not archivo_contable:
        print("\n[ERROR] No se pudieron encontrar los archivos necesarios")
        return
    
    # Carga de Banco
    try:
        banco = cargar_banco(archivo_banco, nombre_banco, codigo_banco)
    except Exception as e:
        print(f"[ERROR] Error al cargar banco: {e}")
        return
        
    # Carga de Contable
    if arg_currency:
        usa_dolares = (arg_currency.upper() == 'USD')
    else:
        while True:
            op = input("¿Está utilizando dólares? (s/n): ").strip().lower()
            if op == 's':   
                usa_dolares = True
            elif op == 'n':
                usa_dolares = False
            if op in ['s', 'n']:
                break
            else:
                print(f"[ERROR] Entrada inválida.")
            
    try:
        contable = cargar_contable(archivo_contable, usa_dolares, "CONTABLE")
    except Exception as e:
        print(f"[ERROR] Error al cargar contable: {e}")
        return
    
    print("\n" + "="*70)
    print("🔄 EJECUTANDO MOTOR v6.5")
    print("="*70)
    
    conciliaciones = []
    
    t1 = conciliacion_por_monto_exacto(banco, contable, conciliaciones)
    t1_5 = conciliacion_con_comisiones(banco, contable, conciliaciones)
    t1_6 = conciliacion_comisiones_agrupadas(banco, contable, conciliaciones)
    t2 = conciliacion_n_a_1_inteligente(banco, contable, conciliaciones)
    t3 = conciliacion_1_a_n_inteligente(banco, contable, conciliaciones)
    t4 = conciliacion_n_a_m_inteligente(banco, contable, conciliaciones)
    t5 = conciliacion_impuestos(banco, contable, conciliaciones)
    t6 = segunda_pasada_inteligente(banco, contable, conciliaciones)
    t7 = busqueda_exhaustiva_final(banco, contable, conciliaciones)
    t8 = conciliacion_caja_chica(banco, contable, conciliaciones)
    
    casos_especiales = detectar_casos_especiales(banco, contable)
    
    if conciliaciones:
        df_conc = pd.DataFrame(conciliaciones).sort_values(['Tipo', 'ID_Conciliacion'])
    else:
        df_conc = pd.DataFrame()
    
    if casos_especiales:
        df_especiales = pd.DataFrame(casos_especiales)
    else:
        df_especiales = pd.DataFrame()
    
    no_b = banco[~banco['Conciliado']][['Fecha', 'Concepto', 'Valor', 'Descripción']].reset_index(drop=True)
    no_c = contable[~contable['Conciliado']][['Fecha', 'Concepto', 'Valor', 'Descripción']].reset_index(drop=True)
    
    max_r = max(len(no_b), len(no_c)) if len(no_b) > 0 or len(no_c) > 0 else 1
    
    df_no_conc = pd.DataFrame({
        'Fecha_Banco': list(no_b['Fecha']) + [None] * (max_r - len(no_b)),
        'Concepto_Banco': list(no_b['Concepto']) + [''] * (max_r - len(no_b)),
        'Valor_Banco': list(no_b['Valor']) + [None] * (max_r - len(no_b)),
        'Descripcion_Banco': list(no_b['Descripción']) + [''] * (max_r - len(no_b)),
        'Fecha_Contable': list(no_c['Fecha']) + [None] * (max_r - len(no_c)),
        'Concepto_Contable': list(no_c['Concepto']) + [''] * (max_r - len(no_c)),
        'Valor_Contable': list(no_c['Valor']) + [None] * (max_r - len(no_c)),
        'Descripcion_Contable': list(no_c['Descripción']) + [''] * (max_r - len(no_c))
    })
    
    total_grupos = len(df_conc['ID_Conciliacion'].unique()) if len(df_conc) > 0 else 0
    total_estrategias = t1 + t1_5 + t1_6 + t2 + t3 + t4 + t5 + t6 + t7 + t8
    
    tiempo_total = time.time() - tiempo_inicio
    
    resumen = pd.DataFrame({
        'Concepto': [
            '══════════════════════════════════════════════════════════════',
            'CONCILIACIÓN BANCARIA v6.5 DEFINITIVO',
            '══════════════════════════════════════════════════════════════',
            '',
            '─── INFORMACIÓN GENERAL ───',
            'Total registros Banco',
            'Total registros Contable',
            'Suma total Banco',
            'Suma total Contable',
            '',
            '─── CONCILIACIONES POR ESTRATEGIA ───',
            '1️⃣  Monto Exacto (1:1)',
            '1.5 Transferencias + Comisión',
            '1.6 Comisiones Agrupadas [v6.1]',
            '2️⃣  Agrupaciones N→1',
            '3️⃣  Agrupaciones 1→N',
            '4️⃣  Agrupaciones N↔M',
            '5️⃣  Impuestos DGII',
            '6️⃣  Segunda Pasada',
            '7️⃣  Búsqueda Exhaustiva [v6.1]',
            '8️⃣  Reposición Caja Chica',
            '📦  TOTAL GRUPOS CONCILIADOS',
            '',
            '─── REGISTROS PROCESADOS ───',
            'Registros Banco conciliados',
            'Registros Contable conciliados',
            'Registros Banco pendientes',
            'Registros Contable pendientes',
            '',
            '─── MONTOS ───',
            'Suma conciliados Banco',
            'Suma conciliados Contable',
            'Suma pendientes Banco',
            'Suma pendientes Contable',
            '',
            '─── INDICADORES ───',
            '[STATS] % Conciliación Banco',
            '[STATS] % Conciliación Contable',
            '[MONEY] DIFERENCIA MONETARIA TOTAL',
            '[WARN]  Casos especiales detectados',
            '⏱️  Tiempo de ejecución (seg)',
        ],
        'Valor': [
            '', '', '', '',
            '',
            len(banco),
            len(contable),
            round(banco['Valor'].sum(), 2),
            round(contable['Valor'].sum(), 2),
            '',
            '',
            t1, t1_5, t1_6, t2, t3, t4, t5, t6, t7, t8,
            total_grupos,
            '',
            '',
            banco['Conciliado'].sum(),
            contable['Conciliado'].sum(),
            len(no_b),
            len(no_c),
            '',
            '',
            round(banco[banco['Conciliado']]['Valor'].sum(), 2),
            round(contable[contable['Conciliado']]['Valor'].sum(), 2),
            round(no_b['Valor'].sum(), 2) if len(no_b) > 0 else 0,
            round(no_c['Valor'].sum(), 2) if len(no_c) > 0 else 0,
            '',
            '',
            round(banco['Conciliado'].sum() / len(banco) * 100, 2),
            round(contable['Conciliado'].sum() / len(contable) * 100, 2),
            round(banco['Valor'].sum() - contable['Valor'].sum(), 2),
            len(casos_especiales),
            round(tiempo_total, 1),
        ]
    })
    
    print("\n" + "="*70)
    print("[SAVE] EXPORTANDO RESULTADOS")
    print("="*70)

    # ─────────────────────────────────────────────
    # Generar nombre dinámico del archivo de salida
    # ─────────────────────────────────────────────
    nombre_archivo = generar_nombre_conciliacion(codigo_banco, banco)
    RUTA_BASE_SALIDA = str(CARPETA_RESULTADOS / nombre_archivo)

    # Mantener lógica de nombre único
    RUTA_SALIDA = obtener_ruta_unica(RUTA_BASE_SALIDA)

    
    with pd.ExcelWriter(RUTA_SALIDA, engine='openpyxl') as writer:
        resumen.to_excel(writer, sheet_name='RESUMEN', index=False)
        
        if len(df_conc) > 0:
            df_conc.to_excel(writer, sheet_name='CONCILIADOS', index=False)
        else:
            pd.DataFrame({'Mensaje': ['No hay conciliaciones']}).to_excel(
                writer, sheet_name='CONCILIADOS', index=False
            )
        
        df_no_conc.to_excel(writer, sheet_name='NO_CONCILIADOS', index=False)
        
        if len(df_especiales) > 0:
            df_especiales.to_excel(writer, sheet_name='CASOS_ESPECIALES', index=False)
    
    print("[OK] Datos exportados correctamente")
    
    if APLICAR_FORMATO_PROFESIONAL:
        print("[STYLE] Aplicando formato profesional...")
        try:
            aplicar_formato_profesional(RUTA_SALIDA)
        except Exception as e:
            print(f"[WARN] Error al aplicar formato: {e}")
    
    perc_b = banco['Conciliado'].sum() / len(banco) * 100
    perc_c = contable['Conciliado'].sum() / len(contable) * 100
    
    print("\n" + "="*70)
    print("       [OK] CONCILIACIÓN COMPLETADA EXITOSAMENTE")
    print("="*70)
    print(f"\n📁 Archivo generado: {RUTA_SALIDA}\n")
    
    print("[STATS] RESUMEN EJECUTIVO:")
    print("─"*70)
    print(f"📥 Datos Originales:")
    print(f"   • Banco:     {len(banco):,} registros (${banco['Valor'].sum():,.2f})")
    print(f"   • Contable:  {len(contable):,} registros (${contable['Valor'].sum():,.2f})")
    
    print(f"\n[OK] Conciliaciones por Estrategia:")
    print(f"   1️⃣  Monto Exacto:             {t1:,}")
    print(f"   1.5 Transf + Comisión:        {t1_5:,}")
    print(f"   1.6 Comisiones Agrupadas:     {t1_6:,}")
    print(f"   2️⃣  N→1:                      {t2:,}")
    print(f"   3️⃣  1→N:                      {t3:,}")
    print(f"   4️⃣  N↔M:                      {t4:,}")
    print(f"   5️⃣  Impuestos DGII:           {t5:,}")
    print(f"   6️⃣  Segunda Pasada:           {t6:,}")
    print(f"   7️⃣  Búsqueda Exhaustiva:      {t7:,}")
    print(f"   8️⃣  Reposición Caja Chica:    {t8:,}")
    print(f"   {'─'*32}")
    print(f"   📦 TOTAL GRUPOS:              {total_grupos:,}")
    
    print(f"\n📈 Resultados:")
    print(f"   • Banco:     {banco['Conciliado'].sum():,} conciliados ({perc_b:.1f}%)")
    print(f"   • Contable:  {contable['Conciliado'].sum():,} conciliados ({perc_c:.1f}%)")
    print(f"   • Pendientes: Banco {len(no_b):,} | Contable {len(no_c):,}")
    
    diferencia = banco['Valor'].sum() - contable['Valor'].sum()
    print(f"\n[MONEY] DIFERENCIA TOTAL: ${diferencia:,.2f}")
    
    if len(casos_especiales) > 0:
        print(f"\n[WARN] Casos Especiales: {len(casos_especiales)}")
        print(f"   → Revisa la hoja 'CASOS_ESPECIALES' para más detalles")
    
    print(f"\n⏱️ Tiempo de Ejecución: {tiempo_total:.1f} segundos")
    
    print(f"\n[TARGET] Evaluación:")
    if perc_b >= 99 and perc_c >= 99:
        print("   [*][*][*][*][*] PERFECTA (≥99%)")
    elif perc_b >= 95 and perc_c >= 95:
        print("   [*][*][*][*] EXCELENTE (≥95%)")
    elif perc_b >= 85 and perc_c >= 85:
        print("   [*][*][*] MUY BUENA (≥85%)")
    elif perc_b >= 75 and perc_c >= 75:
        print("   [*][*] BUENA (≥75%)")
    elif perc_b >= 60 and perc_c >= 60:
        print("   [*] REGULAR (≥60%)")
    else:
        print("   [WARN] NECESITA REVISIÓN (<60%)")
    
    print("\n" + "="*70)
    print("[INFO] TIPS:")
    print("─"*70)
    print("  1. Revisa RESUMEN para estadísticas generales")
    print("  2. Valida CONCILIADOS para verificar agrupaciones")
    print("  3. Analiza NO_CONCILIADOS para identificar diferencias")
    if len(casos_especiales) > 0:
        print("  4. Revisa CASOS_ESPECIALES para coincidencias parciales")
    print("="*70 + "\n")
    pass

def menu_principal():
    while True:
        print("\n" + "="*50)
        print("         [BANK] CONCILIACIÓN BANCARIA v0.9.6")
        print("                 --- MAIN MENU ---")
        print("="*50)
        print("1) Realizar Conciliación")
        print("0) Acabar Programa")
        print("="*50)

        opcion = input("Seleccione una opción: ").strip()

        if opcion == "1":
            ejecutar_conciliacion()

            input("\n🔁 Presione ENTER para volver al menú...")

        elif opcion == "0":
            print("\n👋 Programa finalizado. Hasta luego.")
            break

        else:
            print("\n[ERROR] Opción inválida. Intente de nuevo.")

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Conciliación Bancaria Headless")
    parser.add_argument("--dir", help="Directorio de trabajo")
    parser.add_argument("--bank", help="Archivo de banco")
    parser.add_argument("--ledger", help="Archivo de libro contable")
    parser.add_argument("--currency", choices=["DOP", "USD"], help="Moneda")
    
    # Parámetros adicionales
    parser.add_argument("--tol_exacta", type=float)
    parser.add_argument("--tol_agrup", type=float)
    parser.add_argument("--tol_parcial", type=float)
    parser.add_argument("--ven_exacta", type=int)
    parser.add_argument("--ven_agrup", type=int)
    parser.add_argument("--ven_flex", type=int)
    parser.add_argument("--ven_comis", type=int)
    parser.add_argument("--umb_baja", type=float)
    parser.add_argument("--umb_media", type=float)
    parser.add_argument("--umb_alta", type=float)
    parser.add_argument("--solo_monto", type=str)
    parser.add_argument("--usar_fechas", type=str)
    parser.add_argument("--especiales", type=str)
    parser.add_argument("--profesional", type=str)
    parser.add_argument("--segunda_pasada", type=str)
    parser.add_argument("--exhaustiva", type=str)
    parser.add_argument("--comision_usd", type=float)
    parser.add_argument("--det_comis", type=str)
    parser.add_argument("--max_partidas", type=int)
    parser.add_argument("--max_comb", type=int)
    parser.add_argument("--umb_exh", type=int)
    parser.add_argument("--max_exh", type=int)
    
    # Parse arguments
    args, unknown = parser.parse_known_args()
    
    def str_to_bool(v):
        if v is None: return None
        return v.lower() in ("yes", "true", "t", "1")

    if args.dir and args.bank and args.ledger:
        # Modo Headless
        kwargs = {
            'tol_exacta': args.tol_exacta,
            'tol_agrup': args.tol_agrup,
            'tol_parcial': args.tol_parcial,
            'ven_exacta': args.ven_exacta,
            'ven_agrup': args.ven_agrup,
            'ven_flex': args.ven_flex,
            'ven_comis': args.ven_comis,
            'umb_baja': args.umb_baja,
            'umb_media': args.umb_media,
            'umb_alta': args.umb_alta,
            'solo_monto': str_to_bool(args.solo_monto),
            'usar_fechas': str_to_bool(args.usar_fechas),
            'especiales': str_to_bool(args.especiales),
            'profesional': str_to_bool(args.profesional),
            'segunda_pasada': str_to_bool(args.segunda_pasada),
            'exhaustiva': str_to_bool(args.exhaustiva),
            'comision_usd': args.comision_usd,
            'det_comis': str_to_bool(args.det_comis),
            'max_partidas': args.max_partidas,
            'max_comb': args.max_comb,
            'umb_exh': args.umb_exh,
            'max_exh': args.max_exh
        }
        ejecutar_conciliacion(args.dir, args.bank, args.ledger, args.currency, **kwargs)
    else:
        # Modo Interactivo
        menu_principal()