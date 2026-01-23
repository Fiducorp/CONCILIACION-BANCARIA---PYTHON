"""
============================================================================
CONCILIACI√ìN BANCARIA INTELIGENTE v6.5 DEFINITIVO
============================================================================
Proyecto: GREEN PARK - BANRESERVAS

MOTOR DEFINITIVO = v0.6 (completo) + v0.6.1 (estrategias 1.6 y 7) + OPTIMIZACIONES

‚úÖ TODAS LAS ESTRATEGIAS (7 estrategias completas):
   1.   Monto Exacto (1:1)
   1.5  Transferencias + Comisi√≥n ($7)
   1.6  Comisiones Agrupadas Multi-fecha [v0.6.1]
   2.   Agrupaciones N‚Üí1 (4 m√©todos)
   3.   Agrupaciones 1‚ÜíN
   4.   Agrupaciones N‚ÜîM
   5.   Impuestos DGII (0.15%)
   6.   Segunda Pasada Flexible
   7.   B√∫squeda Exhaustiva Final [v0.6.1]

‚úÖ OPTIMIZACIONES v0.6.5:
   - Lectura r√°pida: solo filas con Fecha Y Valor v√°lidos
   - L√≠mites por estrategia para evitar timeouts
   - Alias TC/LEGAL mejorados: TC LEGAL ‚Üî TC Corporativa ‚Üî Legalizaciones ‚Üî IPI
   - Descripci√≥n vac√≠a funcional (crea columna si no existe)
   - Score combinado para desempate (Concepto_Norm del v5 restaurado)

üéØ OBJETIVO: < 1 minuto de ejecuci√≥n con m√°xima conciliaci√≥n
============================================================================
"""

import pandas as pd
import csv
import numpy as np
from datetime import datetime, timedelta
import os
import glob
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import re
from difflib import SequenceMatcher
import unicodedata
from itertools import combinations
import warnings
warnings.filterwarnings('ignore')

# ============================================================================
# ‚öôÔ∏è PAR√ÅMETROS AJUSTABLES
# ============================================================================

# üí∞ TOLERANCIAS DE VALOR
TOLERANCIA_VALOR_EXACTA = 0.01
TOLERANCIA_VALOR_AGRUPACION = 1.00
TOLERANCIA_PORCENTAJE_PARCIAL = 0.02

# üìÖ VENTANAS DE TIEMPO
VENTANA_DIAS_EXACTA = 10
VENTANA_DIAS_AGRUPACION = 20
VENTANA_DIAS_FLEXIBLE = 30
VENTANA_DIAS_COMISIONES = 45

# üéØ UMBRALES DE SIMILITUD
UMBRAL_SIMILITUD_BAJA = 0.05
UMBRAL_SIMILITUD_MEDIA = 0.20
UMBRAL_SIMILITUD_ALTA = 0.40

# üîß CONFIGURACI√ìN AVANZADA
PERMITIR_SOLO_MONTO = True
USAR_FECHAS_PARA_DESAMBIGUAR = True
DETECTAR_CASOS_ESPECIALES = True
APLICAR_FORMATO_PROFESIONAL = True
EJECUTAR_SEGUNDA_PASADA = True
EJECUTAR_BUSQUEDA_EXHAUSTIVA = True

# üíµ COMISIONES BANCARIAS
COMISION_TRANSFERENCIA_USD = 7.00
DETECTAR_COMISIONES = True

# ‚ö° L√çMITES DE RENDIMIENTO (OPTIMIZACI√ìN)
MAX_PARTIDAS_AGRUPACION = 30        # Reducido de 100 a 15 para velocidad
MAX_COMBINACIONES_POR_BUSQUEDA = 10000  # L√≠mite por cada b√∫squeda individual
UMBRAL_PARTIDAS_EXHAUSTIVA = 25     # Aumentado de 20 a 25
MAX_COMBINACIONES_EXHAUSTIVA = 100000

# ============================================================================
# ‚öôÔ∏è CONFIGURACI√ìN DE RUTAS Y ARCHIVOS RESULTADO
# ============================================================================

import calendar
def generar_nombre_conciliacion(codigo_banco, df_banco):
    """
    Genera nombre tipo:
    Conciliacion_POPULAR_Marzo_2024.xlsx
    """
    # Tomar la primera fecha v√°lida del banco
    fecha_ref = df_banco['Fecha'].dropna().iloc[0]

    mes_nombre = calendar.month_name[fecha_ref.month]
    anio = fecha_ref.year

    # Capitalizar mes (March -> Marzo si luego quieres traducir)
    mes_nombre = mes_nombre.capitalize()

    return f"Conciliacion_{codigo_banco}_{mes_nombre}_{anio}.xlsx"

def obtener_ruta_unica(ruta_base):
    """
    Genera un nombre de archivo √∫nico.
    Si el archivo existe, agrega (1), (2), etc. como en descargas est√°ndar.
    Ejemplo: Conciliacion_Resultados(1).xlsx, Conciliacion_Resultados(2).xlsx
    """
    if not Path(ruta_base).exists():
        return ruta_base
    
    # Dividir ruta base en directorio, nombre y extensi√≥n
    p = Path(ruta_base)
    directorio = p.parent
    nombre_sin_ext = p.stem
    extension = p.suffix
    
    contador = 1
    while True:
        nueva_ruta = directorio / f"{nombre_sin_ext}({contador}){extension}"
        if not nueva_ruta.exists():
            return str(nueva_ruta)
        contador += 1

# ============================================================================
# üè¶ PASO 1: SISTEMA DE RECONOCIMIENTO AUTOM√ÅTICO DE BANCOS
# ============================================================================

# Obtener la carpeta donde est√° el script
CARPETA_TRABAJO = os.path.dirname(os.path.abspath(__file__))
CARPETA_BANCOS = os.path.join(CARPETA_TRABAJO, 'Archivos Banco')
CARPETA_CONTABLE = os.path.join(CARPETA_TRABAJO, 'Archivos Libro Contable')

# Verificar que las carpetas existen
for carpeta in [CARPETA_BANCOS, CARPETA_CONTABLE]:
    if not os.path.isdir(carpeta):
        raise FileNotFoundError(f"No existe la carpeta requerida: {carpeta}")

# Lista de bancos dominicanos soportados
BANCOS_SOPORTADOS = {
    'POPULAR': ['POPULAR', 'BANCO POPULAR', 'BPD', 'BP'],
    'BANRESERVAS': ['BANRESERVAS', 'BANCO DE RESERVAS', 'RESERVAS', 'BDR'],
    'BHD': ['BHD', 'BANCO BHD', 'BHD LEON', 'BHDLEON'],
    'BANESCO': ['BANESCO', 'BANCO BANESCO'],
    'SCOTIABANK': ['SCOTIABANK', 'SCOTIA', 'BANCO SCOTIA'],
    'PROMERICA': ['PROMERICA', 'BANCO PROMERICA'],
    'SANTA_CRUZ': ['SANTA CRUZ', 'SANTACRUZ', 'BANCO SANTA CRUZ', 'BSC'],
    'ADEMI': ['ADEMI', 'BANCO ADEMI'],
    'LOPEZ_DE_HARO': ['LOPEZ DE HARO', 'LOPEZDEHARO', 'BANCO LOPEZ'],
    'BELLBANK': ['BELLBANK', 'BELL BANK'],
    'APAP': ['APAP', 'BANCO APAP', 'ASOCIACION POPULAR DE AHORROS Y PRESTAMOS', 'ASOCIACION POPULAR']
}

def normalizar_nombre_banco(texto):
    """Normaliza el nombre para b√∫squeda (sin acentos, may√∫sculas, sin espacios ni s√≠mbolos)"""
    if not texto:
        return ""
    texto = quitar_acentos(str(texto)).upper()
    texto = re.sub(r'[^A-Z0-9]', '', texto)  # Solo letras y n√∫meros
    return texto

def detectar_banco_en_nombre_archivo(nombre_archivo):
    """
    Detecta qu√© banco es bas√°ndose en el nombre del archivo
    Retorna: (codigo_banco, nombre_legible) o (None, None)
    """
    nombre_norm = normalizar_nombre_banco(nombre_archivo)
    
    for codigo_banco, variantes in BANCOS_SOPORTADOS.items():
        for variante in variantes:
            variante_norm = normalizar_nombre_banco(variante)
            if variante_norm in nombre_norm:
                return codigo_banco, variante
    
    return None, None

def buscar_archivos_en_carpeta():
    """
    Busca archivos de banco en 'Archivos Banco'
    y libro contable en 'Archivos Libro Contable'
    """

    print("\n" + "="*70)
    print("üîç PASO 1: BUSCANDO ARCHIVOS")
    print("="*70)

    print(f"üè¶ Carpeta Bancos:   {CARPETA_BANCOS}")
    print(f"üìñ Carpeta Contable: {CARPETA_CONTABLE}\n")

    # ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ
    # Buscar archivos de BANCO
    # ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ
    archivos_banco = []
    for extension in ['*.xlsx', '*.xls', '*.xlsm', '*.csv']:
        for ruta_archivo in glob.glob(os.path.join(CARPETA_BANCOS, extension)):
            nombre_archivo = os.path.basename(ruta_archivo)
            nombre_sin_ext = os.path.splitext(nombre_archivo)[0]

            banco_det, nombre_det = detectar_banco_en_nombre_archivo(nombre_sin_ext)
            if banco_det:
                archivos_banco.append((ruta_archivo, banco_det, nombre_det, nombre_archivo))
                print(f"‚úÖ BANCO identificado: {nombre_archivo}")
                print(f"   üè¶ Banco: {banco_det} ({nombre_det})")

    # ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ
    # Buscar archivo CONTABLE
    # ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ
    archivos_contable = []
    for extension in ['*.xlsx', '*.xls', '*.xlsm', '*.csv']:
        for ruta_archivo in glob.glob(os.path.join(CARPETA_CONTABLE, extension)):
            nombre_archivo = os.path.basename(ruta_archivo)
            archivos_contable.append((ruta_archivo, 'CONTABLE', 'Libro Contable', nombre_archivo))
            print(f"‚úÖ CONTABLE identificado: {nombre_archivo}")

    print("\n" + "‚îÄ"*70)

    # ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ
    # Validaciones
    # ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ
    if not archivos_banco:
        print("‚ùå ERROR: No se encontraron archivos de banco")
        return None, None, None, None

    if not archivos_contable:
        print("‚ùå ERROR: No se encontr√≥ archivo del libro contable")
        return None, None, None, None

    # ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ
    # Selecci√≥n BANCO
    # ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ
    if len(archivos_banco) > 1:
        print("\n‚ùó M√öLTIPLES ARCHIVOS DE BANCO:\n")
        for i, (_, codigo, nombre, archivo) in enumerate(archivos_banco, 1):
            print(f"   {i}) {archivo} ‚Äî {codigo} ({nombre})")

        while True:
            entrada = input(f"\nSelecciona banco (1-{len(archivos_banco)}): ").strip()
            if entrada.isdigit() and 1 <= int(entrada) <= len(archivos_banco):
                archivo_banco, codigo_banco, nombre_banco, _ = archivos_banco[int(entrada) - 1]
                break
            print("‚ùå Selecci√≥n inv√°lida")
    else:
        archivo_banco, codigo_banco, nombre_banco, _ = archivos_banco[0]

    # ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ
    # Selecci√≥n CONTABLE
    # ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ
    if len(archivos_contable) > 1:
        print("\n‚ùó M√öLTIPLES ARCHIVOS CONTABLES:\n")
        for i, (_, _, _, archivo) in enumerate(archivos_contable, 1):
            print(f"   {i}) {archivo}")

        while True:
            entrada = input(f"\nSelecciona libro contable (1-{len(archivos_contable)}): ").strip()
            if entrada.isdigit() and 1 <= int(entrada) <= len(archivos_contable):
                archivo_contable, _, _, _ = archivos_contable[int(entrada) - 1]
                break
            print("‚ùå Selecci√≥n inv√°lida")
    else:
        archivo_contable, _, _, _ = archivos_contable[0]

    # ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ
    # Resultado final
    # ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ
    print("\n‚úÖ Archivos validados correctamente")
    print(f"   üè¶ Banco:    {os.path.basename(archivo_banco)}")
    print(f"   üìñ Contable: {os.path.basename(archivo_contable)}")

    return archivo_banco, archivo_contable, codigo_banco, nombre_banco

# ============================================================================
# üìù ALIAS DE EMPRESAS Y CONCEPTOS (AMPLIADO PARA TC/LEGAL)
# ============================================================================

ALIAS_EMPRESAS = {
    # RG-RB / CONCRESOL
    'RGRB': ['RG RB', 'RGRB', 'RG-RB', 'CONCRESOL', 'CONCRESOL SRL'],
    'CONCRESOL': ['RG RB', 'RGRB', 'RG-RB', 'CONCRESOL', 'CONCRESOL SRL'],
    
    # TC / TARJETA CORPORATIVA / LEGALIZACIONES [NUEVO v6.5]
    'TC_CORPORATIVA': [
        'TC LEGAL', 'TC CORPORATIVA', 'TARJETA CORPORATIVA', 'TC',
        'LEGAL', 'LEGALIZACIONES', 'LEGALIZACION', 'IPI',
        'CORTE TC', 'CORTE TC LEGAL', 'PAGO TC', 'REEMBOLSO TC',
        'REEMBOLSO CORTE TC', 'REEMBOLSO TC LEGAL', 'DEBITO PAGO',
        'DEBITO PAGO RD', 'PAGO RD TC', 'TC CORPORATIVA'
    ],
}

# Patrones espec√≠ficos
PATRONES_TC = [
    'TC LEGAL', 'TC CORPORATIVA', 'TARJETA CORPORATIVA', 'CORTE TC',
    'REEMBOLSO TC', 'DEBITO PAGO', 'LEGALIZACIONES', 'IPI',
    'PAGO RD TC', 'REEMBOLSO CORTE'
]

PATRONES_IMPUESTO = ['IMP', 'IMPUESTO', 'DGII', '0.15', '0.15%', 'ITF', 'RETENCION']

PATRONES_COMISION = [
    'COM', 'COMISION', 'COM.', 'CARGO', 'COSTO', 'GASTO BANCARIO',
    'EXI', 'COMISION BANCARIA', 'COMISIONES', 'LBTR',
    'TRANS APROBA', 'DEBITO CUENTA', 'GASTO', 'SERVICIO'
]

PATRONES_TRANSFERENCIA = ['TRANSFER', 'TRANSF', 'ACH', 'CR TRANSFERENCIA', 'REALIT LLC','LLC','REALIT']

PALABRAS_COMUNES = {
    'DE', 'DEL', 'LA', 'EL', 'LOS', 'LAS', 'CON', 'SIN', 'PARA', 'POR', 'QUE',
    'COMO', 'DESDE', 'HASTA', 'ENTRE', 'ACH', 'TRANSFERENCIA', 'PAGO', 'TRANSF',
    'SAS', 'LTDA', 'SA', 'SRL', 'CIA', 'RECAUDO', 'INMOB', 'FIDEICOMISO',
    'DEPOSITO', 'BANCO', 'CTA', 'CTAS', 'FAC', 'FACT', 'FACTURA', 'NUM',
    'NUMERO', 'NO', 'Y', 'A', 'EN', 'AL', 'O', 'U', 'E', 'USD', 'DOP', 'RD'
}

# ============================================================================
# üìù PATRONES PARA CAJA CHICA (A√ëADIR A LA SECCI√ìN DE ALIAS)
# ============================================================================

PATRONES_CAJA_CHICA = [
    'CAJA CHICA', 'CAJACHICA', 'REPOSICION', 'REPOSICION DE CAJA',
    'REP CAJA', 'REPO CAJA', 'REEMBOLSO CAJA', 'CAJA MENOR'
]

# ============================================================================
# üîß FUNCIONES DE NORMALIZACI√ìN
# ============================================================================

def quitar_acentos(texto):
    if pd.isna(texto) or texto is None:
        return ""
    texto = str(texto).strip()
    if not texto:
        return ""
    texto_nfd = unicodedata.normalize('NFD', texto)
    return ''.join(c for c in texto_nfd if unicodedata.category(c) != 'Mn')

def normalizar_texto(texto):
    if pd.isna(texto) or texto is None:
        return ""
    texto = quitar_acentos(str(texto))
    texto = texto.upper()
    texto = re.sub(r'[^A-Z0-9\s]', ' ', texto)
    texto = re.sub(r'\s+', ' ', texto)
    return texto.strip()

def es_patron_tc(texto):
    """Detecta si el texto corresponde a TC/Legalizaciones"""
    texto_norm = normalizar_texto(texto)
    return any(patron in texto_norm for patron in PATRONES_TC)

def normalizar_nombre_empresa(texto):
    """Normaliza nombre de empresa + detecta TC/LEGAL"""
    texto_norm = normalizar_texto(texto)
    if not texto_norm:
        return ""
    
    # PRIMERO: Verificar si es TC/LEGAL (prioridad alta)
    if es_patron_tc(texto):
        return 'TC_CORPORATIVA'
    
    # Eliminar sufijos legales
    sufijos = ['SRL', 'SA', 'SAS', 'LTDA', 'CIA', 'INC', 'LLC', 'CORP']
    for sufijo in sufijos:
        texto_norm = re.sub(rf'\b{sufijo}\b', '', texto_norm)
    
    texto_norm = re.sub(r'\s+', ' ', texto_norm).strip()
    texto_compacto = texto_norm.replace(' ', '')
    
    # Buscar en alias
    for clave, aliases in ALIAS_EMPRESAS.items():
        for alias in aliases:
            alias_compacto = normalizar_texto(alias).replace(' ', '')
            if alias_compacto and texto_compacto:
                if alias_compacto in texto_compacto or texto_compacto in alias_compacto:
                    return clave
    
    return texto_norm

def extraer_palabras_clave(texto):
    texto_norm = normalizar_texto(texto)
    if not texto_norm:
        return []
    palabras = [p for p in texto_norm.split() if len(p) >= 3 and p not in PALABRAS_COMUNES]
    numeros = re.findall(r'\d{5,}', texto_norm)
    return list(set(palabras + numeros))

def extraer_identificador_proveedor(texto):
    texto_norm = normalizar_texto(texto)
    if not texto_norm:
        return None
    match = re.search(r'^\s*(\d{8,})', texto_norm)
    return match.group(1) if match else None

def calcular_similitud(texto1, texto2):
    # Comparar nombres de empresa primero
    emp1 = normalizar_nombre_empresa(texto1)
    emp2 = normalizar_nombre_empresa(texto2)
    
    if emp1 and emp2 and emp1 == emp2:
        return 1.0
    
    palabras1 = set(extraer_palabras_clave(texto1))
    palabras2 = set(extraer_palabras_clave(texto2))
    
    if not palabras1 or not palabras2:
        return 0.0
    
    coincidencias = len(palabras1 & palabras2)
    
    # Coincidencias difusas
    for p1 in palabras1:
        for p2 in palabras2:
            if len(p1) >= 4 and len(p2) >= 4:
                if SequenceMatcher(None, p1, p2).ratio() > 0.85:
                    coincidencias += 0.5
                    break
    
    max_palabras = max(len(palabras1), len(palabras2))
    return coincidencias / max_palabras if max_palabras > 0 else 0.0

def es_patron_impuesto(texto):
    texto_norm = normalizar_texto(texto)
    return any(patron in texto_norm for patron in PATRONES_IMPUESTO)

def es_patron_comision(texto):
    texto_norm = normalizar_texto(texto)
    return any(patron in texto_norm for patron in PATRONES_COMISION)

def es_patron_transferencia(texto):
    texto_norm = normalizar_texto(texto)
    return any(patron in texto_norm for patron in PATRONES_TRANSFERENCIA)

# ============================================================================
# üè¶ PASO 2: SISTEMA DE LIMPIEZA ESPEC√çFICO POR BANCO
# ============================================================================

def limpiar_banco_popular(df_original):
    """
    Limpieza espec√≠fica para POPULAR en formato EXCEL
    
    Estructura esperada (sin headers):
    Fecha de Posteo; Descripcion Corta (CONCEPTO); MONTO; USELESS; USELESS; USELESS; Descripcion
    
    Las primeras 10 filas deben ignorarse.
    """
    print("\n   Aplicando limpieza: POPULAR")
    df = df_original.copy()
    df = df.reset_index(drop=True)
    # Eliminar filas completamente vac√≠as
    df = df.dropna(how='all')

    # Buscar fila de cabecera en las primeras N filas
    header_row = None
    max_search = min(25, len(df))
    for i in range(max_search):
        vals = ' '.join([str(x) for x in df.iloc[i].values if pd.notna(x)])
        up = quitar_acentos(vals).upper()
        if 'FECHA POSTEO' in up and ("DESCRIPCION CORTA" in up or 'DESCRIPCION' in up or 'MONTO' in up or 'MONTO TRANSACCION' in up or 'TRANSACCION' in up):
            header_row = i
            break

    # Fallback: buscar solo 'FECHA'
    if header_row is None:
        for i in range(max_search):
            vals = ' '.join([str(x) for x in df.iloc[i].values if pd.notna(x)])
            up = quitar_acentos(vals).upper()
            if 'FECHA' in up:
                header_row = i
                break

    if header_row is None:
        # No se detect√≥ cabecera; devolver df con filas vac√≠as eliminadas
        return df

    # Aplicar la fila encontrada como cabecera y devolver solo las filas de datos
    raw_header = df.iloc[header_row].astype(str).tolist()
    df = df.iloc[header_row + 1 :].reset_index(drop=True)
    df.columns = raw_header

    # ===== Detectar segunda cabecera (inicio de D√©bitos) =====
    header_norm = [quitar_acentos(str(c)).upper() for c in raw_header]

    idx_separador = None

    for i in range(len(df)):
        fila_norm = [
            quitar_acentos(str(x)).upper()
            for x in df.iloc[i].values
            if pd.notna(x)
        ]

        # Si la fila contiene TODAS las columnas de cabecera ‚Üí es cabecera repetida
        if all(any(h in v for v in fila_norm) for h in header_norm):
            idx_separador = i
            break

    if idx_separador is not None:
        print(f"  üîÄ Segunda cabecera detectada en fila: {idx_separador}")
    else:
        print("  ‚ö†Ô∏è No se detect√≥ segunda cabecera (D√©bitos)")

    # Normalizar nombres de columnas m√≠nimamente (sin cambiar valores)
    rename_map = {}
    for col in list(df.columns):
        key = quitar_acentos(str(col)).upper()
        if 'FECHA POSTEO' in key:
            rename_map[col] = 'Fecha'
        elif 'DESCRIPCION CORTA' in key:
            rename_map[col] = 'Concepto'
        elif any(k in key for k in ['DETALLE', 'OBSERV', 'DESCRIPCI√ìN']):
            rename_map[col] = 'Descripci√≥n'
        elif any(k in key for k in ['VALOR', 'MONTO', 'IMPORTE', 'AMOUNT']):
            rename_map[col] = 'Valor'

    if rename_map:
        df = df.rename(columns=rename_map)

    # Convertir Valor
    df['Valor'] = pd.to_numeric(df['Valor'], errors='coerce')

    # Aplicar signo: D√©bitos negativos
    if idx_separador is not None:
        df.loc[idx_separador + 1 :, 'Valor'] *= -1

    # Quitar filas de totales u otros r√≥tulos (l√≠neas que contienen 'TOTAL' u 'NETO')
    mask_tot = df.apply(lambda r: any('TOTAL' in str(x).upper() or 'NETO' in str(x).upper() for x in r.values), axis=1)
    if mask_tot.any():
        df = df[~mask_tot].reset_index(drop=True)

    # Eliminar filas vac√≠as residuales
    df = df.dropna(how='all').reset_index(drop=True)

    return df

def detectar_delimitador_popular(ruta_archivo):
    with open(ruta_archivo, 'r', encoding='latin1') as f:
        for line in f:
            up = quitar_acentos(line).upper()
            if 'FECHA POSTEO' in up and 'MONTO' in up:
                return ';' if line.count(';') > line.count(',') else ','
    raise ValueError("No se pudo detectar el delimitador")

def detectar_headers_en_df(df):
    """
    Devuelve una lista de √≠ndices donde aparece la cabecera
    (Cr√©ditos, D√©bitos, etc.)
    """
    headers = []
    for i in range(len(df)):
        txt = ' '.join(str(x) for x in df.iloc[i].values if pd.notna(x))
        up = quitar_acentos(txt).upper()
        if 'FECHA' in up and 'MONTO' in up:
            headers.append(i)
    return headers

from io import StringIO
def read_dirty_csv_popular(ruta_archivo, sep):
    """
    Reads a bank CSV embedded in garbage text.
    Keeps ONLY lines that contain the delimiter.
    """
    good_lines = []

    with open(ruta_archivo, 'r', encoding='latin1') as f:
        for line in f:
            if sep in line:
                good_lines.append(line)

    if not good_lines:
        raise ValueError("No se encontraron l√≠neas con delimitador")

    return pd.read_csv(
        StringIO(''.join(good_lines)),
        sep=sep,
        header=None,
        engine='python',
        dtype=str
    )

def limpiar_banco_popular_csv(ruta_archivo):

    print("\n   Aplicando limpieza: POPULAR (CSV)")

    try:
        # ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ
        # 1. Detectar delimitador
        # ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ
        sep = detectar_delimitador_popular(ruta_archivo)
        print(f"  üîç Delimitador detectado: '{sep}'")

        # ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ
        # 2. Leer archivo CSV ignorando basura
        # ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ
        df = read_dirty_csv_popular(ruta_archivo, sep)


        print(f"  üìä CSV cargado: {df.shape[0]} filas √ó {df.shape[1]} columnas")

        # ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ
        # 3. Detectar cabeceras dentro del DF
        # ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ
        headers = detectar_headers_en_df(df)

        if len(headers) == 0:
            raise ValueError("No se detect√≥ ninguna tabla")

        idx_creditos = headers[0]
        idx_debitos = headers[1] if len(headers) > 1 else None

        print(f"  üîç Inicio Cr√©ditos: fila {idx_creditos}")
        if idx_debitos is not None:
            print(f"  üîç Inicio D√©bitos: fila {idx_debitos}")

        # ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ
        # 4. Recortar DF a partir de Cr√©ditos
        # ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ
        df = df.iloc[idx_creditos + 1:].reset_index(drop=True)

        df.columns = [f'col_{i}' for i in range(df.shape[1])]

        # ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ
        # 5. Construir DF limpio
        # ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ
        df_limpio = pd.DataFrame({
            'Fecha': df['col_0'],
            'Concepto': df['col_1'],
            'Valor': df['col_2'],
            'Descripci√≥n': df['col_6'],
        })

        # ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ
        # 6. Detectar cabecera D√©bitos (ya alineada)
        # ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ
        idx_debitos_df = None
        for i in range(len(df)):
            txt = ' '.join(str(x) for x in df.iloc[i].values if pd.notna(x))
            up = quitar_acentos(txt).upper()
            if 'FECHA' in up and 'MONTO' in up:
                idx_debitos_df = i
                break

        # ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ
        # 7. Tipos
        # ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ
        df_limpio['Fecha'] = pd.to_datetime(
            df_limpio['Fecha'].str.strip(),
            errors='coerce',
            dayfirst=True
        )

        df_limpio['Valor'] = pd.to_numeric(
            df_limpio['Valor'].str.strip(),
            errors='coerce'
        )

        # ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ
        # 8. Aplicar signo D√©bitos
        # ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ
        if idx_debitos_df is not None:
            print(f"  üîÄ Aplicando D√©bitos desde fila {idx_debitos_df}")
            df_limpio.loc[idx_debitos_df + 1:, 'Valor'] *= -1
            df_limpio = df_limpio.drop(index=idx_debitos_df).reset_index(drop=True)

        # ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ
        # 9. Limpieza final
        # ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ
        df_limpio = df_limpio.dropna(subset=['Valor']).reset_index(drop=True)

        print(f"  ‚úÖ Limpieza POPULAR CSV completada: {df_limpio.shape[0]} filas")
        return df_limpio

    except Exception as e:
        print(f"  ‚ùå Error al limpiar CSV de POPULAR: {e}")
        raise

def limpiar_banreservas(df_original):
    """Limpieza ligera para BANRESERVAS.

    Objetivo: eliminar filas arriba de la tabla de datos (texto libre, t√≠tulos,
    metadatos) para que pandas pueda leer correctamente la tabla. NO debe
    transformar o combinar columnas (p.ej. Debito/Credito) ‚Äî eso lo hace el
    loader universal posteriormente.
    """
    print("\n   Aplicando limpieza: BANRESERVAS (solo remover encabezados no-datos)")
    df = df_original.copy()
    df = df.reset_index(drop=True)
    # Eliminar filas completamente vac√≠as
    df = df.dropna(how='all')

    # Buscar fila de cabecera en las primeras N filas
    header_row = None
    max_search = min(25, len(df))
    for i in range(max_search):
        vals = ' '.join([str(x) for x in df.iloc[i].values if pd.notna(x)])
        up = quitar_acentos(vals).upper()
        if 'FECHA' in up and ("TRANSACC" in up or 'CONCEP' in up or 'NO DE TRANSACCION' in up or 'NO.' in up or 'NO DE TRANSAC' in up):
            header_row = i
            break

    # Fallback: buscar solo 'FECHA'
    if header_row is None:
        for i in range(max_search):
            vals = ' '.join([str(x) for x in df.iloc[i].values if pd.notna(x)])
            up = quitar_acentos(vals).upper()
            if 'FECHA' in up:
                header_row = i
                break

    if header_row is None:
        # No se detect√≥ cabecera; devolver df con filas vac√≠as eliminadas
        return df

    # Aplicar la fila encontrada como cabecera y devolver solo las filas de datos
    raw_header = df.iloc[header_row].astype(str).tolist()
    df = df.iloc[header_row + 1 :].reset_index(drop=True)
    df.columns = raw_header

    # Normalizar nombres de columnas m√≠nimamente (sin cambiar valores)
    rename_map = {}
    for col in list(df.columns):
        key = quitar_acentos(str(col)).upper()
        if 'FECHA' in key:
            rename_map[col] = 'Fecha'
        elif 'CONCEP' in key:
            rename_map[col] = 'Concepto'
        elif 'DEB' in key and 'DEBITO' in key or 'DEBITO' in key:
            rename_map[col] = 'Debito'
        elif 'CRED' in key or 'CREDITO' in key:
            rename_map[col] = 'Credito'
        elif 'REFER' in key:
            rename_map[col] = 'Referencia'
        elif any(k in key for k in ['DESCRIP', 'DETALLE', 'OBSERV', 'DESCRIPCI√ìN']):
            rename_map[col] = 'Descripci√≥n'
        elif any(k in key for k in ['VALOR', 'MONTO', 'IMPORTE', 'AMOUNT']):
            rename_map[col] = 'Valor'

    if rename_map:
        df = df.rename(columns=rename_map)

    # Quitar filas de totales u otros r√≥tulos (l√≠neas que contienen 'TOTAL' u 'NETO')
    mask_tot = df.apply(lambda r: any('TOTAL' in str(x).upper() or 'NETO' in str(x).upper() for x in r.values), axis=1)
    if mask_tot.any():
        df = df[~mask_tot].reset_index(drop=True)

    # Eliminar filas vac√≠as residuales
    df = df.dropna(how='all').reset_index(drop=True)

    return df

def limpiar_bhd(df_original):
    """Limpieza espec√≠fica para BHD (Excel)"""
    print("\n   Aplicando limpieza: BHD")
    df = df_original.copy()
    df = df.reset_index(drop=True)
    df = df.dropna(how='all')
    return df

def limpiar_bhd_csv(ruta_archivo):
    """
    Limpieza espec√≠fica para BHD en formato CSV
    
    Estructura esperada (sin headers):
    Fecha, USELESS, USELESS, USELESS, Concepto, Debito, Credito, USELESS, Balance, Hour
    
    La primera fila es un resumen y debe ignorarse.
    """
    print("\n   Aplicando limpieza: BHD (CSV)")
    
    try:
        # Leer CSV sin headers (no_column_names)
        df = pd.read_csv(ruta_archivo, header=None)
        print(f"  üìä CSV cargado: {df.shape[0]} filas √ó {df.shape[1]} columnas")
        
        # Asignar nombres temporales a las columnas (0-indexed)
        df.columns = [f'col_{i}' for i in range(df.shape[1])]
        
        # Saltar la primera fila (resumen/encabezado)
        df = df.iloc[1:].reset_index(drop=True)
        print(f"  ‚úÖ Salteada primera fila (resumen), quedaron: {df.shape[0]} filas")
        
        # Mapear columnas: Fecha=0, Concepto=4, Debito=5, Credito=6
        # Las dem√°s columnas (1,2,3,7,8,9) son in√∫tiles
        df_limpio = pd.DataFrame({
            'Fecha': df['col_0'],
            'Concepto': df['col_4'],
            'Descripci√≥n': '',  # Campo vac√≠o por defecto
            'Debito': df['col_5'],
            'Credito': df['col_6']
        })
        
        # Limpiar datos
        df_limpio = df_limpio.reset_index(drop=True)
        df_limpio = df_limpio.dropna(how='all')
        
        # Convertir a tipos de datos correctos
        df_limpio['Fecha'] = pd.to_datetime(df_limpio['Fecha'], errors='coerce')
        df_limpio['Debito'] = pd.to_numeric(df_limpio['Debito'], errors='coerce')
        df_limpio['Credito'] = pd.to_numeric(df_limpio['Credito'], errors='coerce')
        
        print(f"  ‚úÖ Limpieza BHD CSV completada: {df_limpio.shape[0]} filas")
        return df_limpio
        
    except Exception as e:
        print(f"  ‚ùå Error al limpiar CSV de BHD: {e}")
        raise

def limpiar_santa_cruz(df_original):
    """Limpieza ligera para SANTA CRUZ.

    Objetivo: eliminar filas arriba de la tabla de datos (texto libre, t√≠tulos,
    metadatos) para que pandas pueda leer correctamente la tabla. NO debe
    transformar o combinar columnas (p.ej. Debito/Credito) ‚Äî eso lo hace el
    loader universal posteriormente.
    
    Mapeo de columnas Santa Cruz:
    - Fecha de Posteo ‚Üí IGNORADA
    - Fecha Efectiva ‚Üí Fecha (IMPORTANTE)
    - No. Cheque ‚Üí IGNORADA
    - No. Referencia ‚Üí Concepto (ID de transacci√≥n)
    - Descripcion ‚Üí Descripci√≥n
    - Retiros ‚Üí Debito
    - Despositos ‚Üí Credito
    - Balance ‚Üí IGNORADA
    """
    print("\n   Aplicando limpieza: SANTA CRUZ (solo remover encabezados no-datos)")
    df = df_original.copy()
    df = df.reset_index(drop=True)
    # Eliminar filas completamente vac√≠as
    df = df.dropna(how='all')

    # Buscar fila de cabecera en las primeras N filas
    header_row = None
    max_search = min(25, len(df))
    for i in range(max_search):
        vals = ' '.join([str(x) for x in df.iloc[i].values if pd.notna(x)])
        up = quitar_acentos(vals).upper()
        # Buscar palabras clave de Santa Cruz
        if any(k in up for k in ['FECHA EFECTIVA', 'FECHA DE POSTEO', 'NO. REFERENCIA', 'REFERENCIA', 'RETIROS', 'DESPOSITOS', 'DEPOSITOS']):
            header_row = i
            break

    # Fallback: buscar solo 'FECHA'
    if header_row is None:
        for i in range(max_search):
            vals = ' '.join([str(x) for x in df.iloc[i].values if pd.notna(x)])
            up = quitar_acentos(vals).upper()
            if 'FECHA' in up:
                header_row = i
                break

    if header_row is None:
        # No se detect√≥ cabecera; devolver df con filas vac√≠as eliminadas
        return df

    # Aplicar la fila encontrada como cabecera y devolver solo las filas de datos
    raw_header = df.iloc[header_row].astype(str).tolist()
    df = df.iloc[header_row + 1 :].reset_index(drop=True)
    df.columns = raw_header

    # Normalizar nombres de columnas para Santa Cruz
    rename_map = {}
    for col in list(df.columns):
        key = quitar_acentos(str(col)).upper()
        
        # Mapeo espec√≠fico para Santa Cruz
        if 'FECHA EFECTIVA' in key:
            rename_map[col] = 'Fecha'
        elif 'FECHA DE POSTEO' in key or 'FECHA POSTEO' in key:
            # Ignorar Fecha de Posteo (no usar para mapeo)
            pass
        elif 'NO. REFERENCIA' in key or 'NO REFERENCIA' in key or 'REFERENCIA' in key:
            rename_map[col] = 'Concepto'
        elif 'DESCRIPCION' in key or 'DESCRIP' in key:
            rename_map[col] = 'Descripci√≥n'
        elif 'RETIROS' in key or 'RETIRO' in key:
            rename_map[col] = 'Debito'
        elif 'DESPOSITOS' in key or 'DEPOSITOS' in key or 'DEPOSITO' in key:
            rename_map[col] = 'Credito'
        elif 'NO. CHEQUE' in key or 'NO CHEQUE' in key or 'CHEQUE' in key:
            # Ignorar No. Cheque
            pass
        elif 'BALANCE' in key or 'SALDO' in key:
            # Ignorar Balance/Saldo
            pass

    if rename_map:
        df = df.rename(columns=rename_map)

    # Quitar filas de totales u otros r√≥tulos (l√≠neas que contienen 'TOTAL' u 'NETO')
    mask_tot = df.apply(lambda r: any('TOTAL' in str(x).upper() or 'NETO' in str(x).upper() for x in r.values), axis=1)
    if mask_tot.any():
        df = df[~mask_tot].reset_index(drop=True)

    # Eliminar filas vac√≠as residuales
    df = df.dropna(how='all').reset_index(drop=True)

    return df

def limpiar_apap(df_original):
    """Limpieza ligera para APAP.

    Objetivo: eliminar filas arriba de la tabla de datos (texto libre, t√≠tulos,
    metadatos) para que pandas pueda leer correctamente la tabla. NO debe
    transformar o combinar columnas (p.ej. Debito/Credito) ‚Äî eso lo hace el
    loader universal posteriormente.
    """
    print("\n   Aplicando limpieza: APAP (solo remover encabezados no-datos)")
    df = df_original.copy()
    df = df.reset_index(drop=True)
    # Eliminar filas completamente vac√≠as
    df = df.dropna(how='all')

    # Buscar fila de cabecera en las primeras N filas
    header_row = None
    max_search = min(25, len(df))
    for i in range(max_search):
        vals = ' '.join([str(x) for x in df.iloc[i].values if pd.notna(x)])
        up = quitar_acentos(vals).upper()
        if 'FECHA' in up and ("REFER" in up or 'CONCEP' in up or 'NO DE REFERENCIA' in up or 'NO.' in up or 'DESC' in up):
            header_row = i
            break

    # Fallback: buscar solo 'FECHA'
    if header_row is None:
        for i in range(max_search):
            vals = ' '.join([str(x) for x in df.iloc[i].values if pd.notna(x)])
            up = quitar_acentos(vals).upper()
            if 'FECHA' in up:
                header_row = i
                break

    if header_row is None:
        # No se detect√≥ cabecera; devolver df con filas vac√≠as eliminadas
        return df

    # Aplicar la fila encontrada como cabecera y devolver solo las filas de datos
    raw_header = df.iloc[header_row].astype(str).tolist()
    df = df.iloc[header_row + 1 :].reset_index(drop=True)
    df.columns = raw_header

    # Normalizar nombres de columnas m√≠nimamente (sin cambiar valores)
    rename_map = {}
    for col in list(df.columns):
        key = quitar_acentos(str(col)).upper()
        if 'FECHA DE ENTRADA' in key:
            rename_map[col] = 'Fecha'
        elif 'NO. DE REFERENCIA' in key:
            rename_map[col] = 'Concepto'
        elif any(k in key for k in ['DESCRIP', 'DETALLE', 'OBSERV', 'DESCRIPCI√ìN']):
            rename_map[col] = 'Descripci√≥n'
        elif any(k in key for k in ['VALOR', 'MONTO', 'IMPORTE', 'AMOUNT']):
            rename_map[col] = 'Valor'

    if rename_map:
        df = df.rename(columns=rename_map)

    # Quitar filas de totales u otros r√≥tulos (l√≠neas que contienen 'TOTAL' u 'NETO')
    mask_tot = df.apply(lambda r: any('TOTAL' in str(x).upper() or 'NETO' in str(x).upper() for x in r.values), axis=1)
    if mask_tot.any():
        df = df[~mask_tot].reset_index(drop=True)

    # Eliminar filas vac√≠as residuales
    df = df.dropna(how='all').reset_index(drop=True)

    return df

def limpiar_ademi(df_original):
    """Limpieza espec√≠fica para ADEMI"""
    print("\n   Aplicando limpieza: ADEMI")
    df = df_original.copy()
    df = df.reset_index(drop=True)
    df = df.dropna(how='all')
    return df

def limpiar_lopez_de_haro(df_original):
    """Limpieza espec√≠fica para LOPEZ_DE_HARO"""
    print("\n   Aplicando limpieza: LOPEZ_DE_HARO")
    df = df_original.copy()
    df = df.reset_index(drop=True)
    df = df.dropna(how='all')
    return df

def limpiar_bellbank(df_original):
    """Limpieza espec√≠fica para BELLBANK"""
    print("\n   Aplicando limpieza: BELLBANK")
    df = df_original.copy()
    df = df.reset_index(drop=True)
    df = df.dropna(how='all')
    return df

def limpiar_banesco(df_original):
    """Limpieza espec√≠fica para BANESCO"""
    print("\n   Aplicando limpieza: BANESCO")
    df = df_original.copy()
    df = df.reset_index(drop=True)
    df = df.dropna(how='all')
    return df

def limpiar_scotiabank(df_original):
    """Limpieza espec√≠fica para SCOTIABANK"""
    print("\n   Aplicando limpieza: SCOTIABANK")
    df = df_original.copy()
    df = df.reset_index(drop=True)
    df = df.dropna(how='all')
    return df

def limpiar_promerica(df_original):
    """Limpieza espec√≠fica para PROMERICA"""
    print("\n   Aplicando limpieza: PROMERICA")
    df = df_original.copy()
    df = df.reset_index(drop=True)
    df = df.dropna(how='all')
    return df

# Mapeo de bancos a funciones de limpieza
FUNCIONES_LIMPIEZA_BANCO = {
    'POPULAR': limpiar_banco_popular,
    'BANRESERVAS': limpiar_banreservas,
    'BHD': limpiar_bhd,
    'BANESCO': limpiar_banesco,
    'SCOTIABANK': limpiar_scotiabank,
    'PROMERICA': limpiar_promerica,
    'SANTA_CRUZ': limpiar_santa_cruz,
    'ADEMI': limpiar_ademi,
    'LOPEZ_DE_HARO': limpiar_lopez_de_haro,
    'BELLBANK': limpiar_bellbank,
    'APAP': limpiar_apap
}

FUNCIONES_LIMPIEZA_BANCO_CSV = {
    'BHD': limpiar_bhd_csv,
    'POPULAR': limpiar_banco_popular_csv
}

# ============================================================================
# üè¶ PASO 2B: LIMPIEZA DEL ARCHIVO CONTABLE (UNIVERSAL)
# ============================================================================

def limpiar_archivo_contable(ruta_archivo, moneda='RD$'):
    """
    Limpia el archivo contable (formato est√°ndar para todos los casos)
    
    PAR√ÅMETROS:
    - ruta_archivo: ruta del archivo Excel
    - moneda: 'RD$' (pesos) o 'USD' (d√≥lares) - para futura implementaci√≥n

    ---CODIGO PARA LIMPIEZA---
    
    """
    print("\n  Cargando archivo CONTABLE ({} )... (limpieza ligera)".format(moneda))
    # Leer Excel y eliminar filas vac√≠as por encima de la tabla principal, sin
    # interpretar la naturaleza (Natu) ni transformar signos. Solo establecer
    # la fila de cabecera correcta para que el loader universal map√©e columnas.
    try:
        df = pd.read_excel(ruta_archivo)
    except Exception as e:
        raise ValueError(f"‚ùå Error al leer contable '{ruta_archivo}': {e}")

    df = df.reset_index(drop=True)
    df = df.dropna(how='all')

    header_row = None
    max_search = min(30, len(df))
    for i in range(max_search):
        vals = ' '.join([str(x) for x in df.iloc[i].values if pd.notna(x)])
        up = quitar_acentos(vals).upper()
        
        # Skip metadata rows that contain date range info (D.Fecha, H.Fecha)
        if 'D.FECHA' in up or 'H.FECHA' in up or 'DESDE' in up or 'HASTA' in up:
            continue
        
        # Buscar elementos t√≠picos del libro contable: F.Comp, Detalle, Nro.Doc, Periodo
        # Require at least 2 of these keywords to be more specific and avoid false positives
        keyword_matches = sum(1 for k in ['F.COMP', 'F COMP', 'DETALLE', 'NRO', 'NRO DOC', 'PERIODO'] if k in up)
        if keyword_matches >= 2:
            header_row = i
            break

    if header_row is None:
        return df

    raw_header = df.iloc[header_row].astype(str).tolist()
    df = df.iloc[header_row + 1 :].reset_index(drop=True)
    df.columns = raw_header

    # No hacemos m√°s transformaciones aqu√≠ (no tocar NATU ni signos)
    df = df.dropna(how='all').reset_index(drop=True)
    return df

# ============================================================================
# üìÇ CARGA DE DATOS OPTIMIZADA (SOLO FILAS V√ÅLIDAS)
# ============================================================================

def parse_fecha_ddmmyyyy(series):
    """
    Robust date parser:
    - Reads ANY reasonable date format
    - Fixes MM/DD/YYYY accidentally parsed as DD/MM/YYYY
    - Guarantees final Fecha is DD/MM/YYYY logic
    """

    s = series.astype(str).str.strip()

    # First: flexible parse (allows Excel, CSV, cleaned files)
    fechas = pd.to_datetime(
        s,
        errors='coerce',
        infer_datetime_format=True,
        dayfirst=False  # IMPORTANT: allow pandas to read US-style too
    )

    # Detect impossible dates under DD/MM/YYYY logic
    # If day > 12 AND month <= 12, it was flipped
    mask_flip = (
        fechas.notna() &
        (fechas.dt.day <= 12) &
        (fechas.dt.month > 12)
    )

    # Swap day/month ONLY for those rows
    fechas.loc[mask_flip] = fechas.loc[mask_flip].apply(
        lambda d: d.replace(day=d.month, month=d.day)
    )

    return fechas

# Carga de Archivos
def cargar_banco(ruta, nombre="", codigo_banco=None):
    """Carga datos - OPTIMIZACI√ìN: solo lee filas con Fecha Y Valor v√°lidos
    Si se proporciona `codigo_banco`, se aplicar√° la funci√≥n de limpieza
    espec√≠fica definida en `FUNCIONES_LIMPIEZA_BANCO` antes del mapeo general.
    """
    print(f"\n  üìÇ Cargando: {nombre.upper() if nombre else ruta}")

    # Resolver ruta relativa: si el usuario pasa solo el nombre de archivo,
    # buscarlo en el mismo directorio del script y en el cwd.
    try:
        ruta_path = Path(ruta)
    except Exception:
        ruta_path = Path(str(ruta))

    if not ruta_path.is_absolute():
        script_dir = Path(__file__).resolve().parent
        candidates = [script_dir / ruta_path, Path.cwd() / ruta_path]

        # If no suffix provided, try common extensions too
        if not ruta_path.suffix:
            exts = ['.xlsx', '.xls', '.csv']
        else:
            exts = [ruta_path.suffix]

        found = None
        for base in candidates:
            for ext in exts:
                cand = base.with_suffix(ext) if not ruta_path.suffix else base
                if cand.exists():
                    found = cand
                    break
            if found:
                break

        if found:
            ruta = str(found)
        else:
            # if the literal provided (e.g., 'BANRESERVAS.xlsx') exists in script_dir
            alt = script_dir / ruta_path.name
            if alt.exists():
                ruta = str(alt)

    try:
        # Detectar si es CSV o Excel basado en la extensi√≥n del archivo
        if ruta.lower().endswith('.csv'):
            print(f"  üìã Detectado archivo CSV")
            # Aplicar limpieza espec√≠fica si existe para CSV
            if codigo_banco and codigo_banco in FUNCIONES_LIMPIEZA_BANCO_CSV:
                print(f"  üßπ Aplicando limpieza espec√≠fica para {codigo_banco} (CSV)...")
                df = FUNCIONES_LIMPIEZA_BANCO_CSV[codigo_banco](ruta)
            else:
                # Para otros CSVs, cargar directamente
                df = pd.read_csv(ruta)
        else:
            # Leer Excel
            df = pd.read_excel(ruta)

        # ‚ö° OPTIMIZACI√ìN 1: Eliminar filas completamente vac√≠as PRIMERO
        df = df.dropna(how='all')

        # Si se indic√≥ un c√≥digo de banco y existe una funci√≥n de limpieza,
        # aplicar la limpieza espec√≠fica antes del mapeo autom√°tico.
        # (Para CSV de BHD, la limpieza ya se aplic√≥ arriba)
        if codigo_banco and codigo_banco in FUNCIONES_LIMPIEZA_BANCO and not (ruta.lower().endswith('.csv')):
            try:
                print(f"  üßπ Aplicando limpieza espec√≠fica para {codigo_banco}...")
                df = FUNCIONES_LIMPIEZA_BANCO[codigo_banco](df)
            except Exception as e:
                raise ValueError(f"‚ùå Error en limpieza espec√≠fica de {codigo_banco}: {e}")

        # Evitar errores por columnas duplicadas en el Excel (por ejemplo cabeceras repetidas).
        # Hacemos los nombres de columnas √∫nicos antes del mapeo (col -> col, col_1, col_2...)
        cols = [str(c) if c is not None else '' for c in df.columns]
        seen = {}
        unique_cols = []
        for c in cols:
            if c in seen:
                seen[c] += 1
                new_c = f"{c}_{seen[c]}"
            else:
                seen[c] = 0
                new_c = c
            unique_cols.append(new_c)
        df.columns = unique_cols

    except Exception as e:
        raise ValueError(f"‚ùå Error al cargar '{ruta}': {e}")
    
    print(f"  üìã Columnas encontradas: {list(df.columns)}")
    
    # Mapeo autom√°tico de columnas
    columnas_map = {}
    tiene_debito_credito = False
    for col in df.columns:
        col_upper = str(col).upper().strip()
        if any(x in col_upper for x in ['FECHA', 'DATE','EFECTIVA']) and 'Fecha' not in columnas_map:
            columnas_map['Fecha'] = col
        elif any(x in col_upper for x in ['CONCEPTO', 'REFERENCIA']) and 'Concepto' not in columnas_map:
            columnas_map['Concepto'] = col
        elif any(x in col_upper for x in ['D√âBITO', 'DEBITO', 'RETIRO', 'RETIROS']) and 'Debito' not in columnas_map:
            columnas_map['Debito'] = col
        elif any(x in col_upper for x in ['CR√âDITO', 'CREDITO', 'DEPOSITO', 'DEPOSITOS']) and 'Credito' not in columnas_map:
            columnas_map['Credito'] = col
        elif any(x in col_upper for x in ['VALOR', 'MONTO', 'IMPORTE']) and 'Valor' not in columnas_map:
            columnas_map['Valor'] = col
        elif any(x in col_upper for x in ['DESCRIP', 'DETALLE', 'OBSERV', 'DESCRIPCI√ìN']) and 'Descripci√≥n' not in columnas_map:
            columnas_map['Descripci√≥n'] = col

    # Fallback: detect Valor-like column with additional keywords (English variants)
    if 'Valor' not in columnas_map:
        for col in df.columns:
            cu = str(col).upper()
            if any(k in cu for k in ['VALOR', 'MONTO', 'IMPORTE', 'AMOUNT', 'AMT']):
                columnas_map['Valor'] = col
                break
    
    # Detectar si hay Debito y/o Credito
    tiene_debito_credito = ('Debito' in columnas_map) or ('Credito' in columnas_map)
    print(f"  üîç Formato detectado: {'Debito/Credito' if tiene_debito_credito else 'Valor √önico'}")
    # Si no hay Descripci√≥n, crearla vac√≠a
    if 'Descripci√≥n' not in columnas_map:
        df['Descripci√≥n'] = ''
        columnas_map['Descripci√≥n'] = 'Descripci√≥n'
        print("  ‚ö†Ô∏è Columna 'Descripci√≥n' no encontrada - creada vac√≠a")
    
    # Verificar columnas requeridas
    if tiene_debito_credito:
        # Requerir Fecha, Concepto y al menos una de Debito/Credito
        requeridas = ['Fecha', 'Concepto']
        if 'Debito' in columnas_map:
            requeridas.append('Debito')
        if 'Credito' in columnas_map:
            requeridas.append('Credito')
    else:
        requeridas = ['Fecha', 'Concepto', 'Valor']
    
    faltantes = [r for r in requeridas if r not in columnas_map]
    if faltantes:
        raise ValueError(f"‚ùå Faltan columnas requeridas: {faltantes}")
    
    # Renombrar columnas - hacerlo de forma segura para evitar reindex errors
    try:
        mapping = {v: k for k, v in columnas_map.items()}
        # Construir nueva lista de columnas aplicando mapping y asegurando unicidad
        new_cols = []
        seen = {}
        for orig in list(df.columns):
            target = mapping.get(orig, orig)
            if target in seen:
                seen[target] += 1
                target = f"{target}_{seen[target]}"
            else:
                seen[target] = 0
            new_cols.append(target)
        df.columns = new_cols
    except Exception as e:
        # Proveer informaci√≥n diagn√≥stica
        msg = (
            f"Error renombrando columnas: {e}\n"
            f"Columnas actuales: {list(df.columns)}\n"
            f"Columnas mapeadas: {columnas_map}"
        )
        raise ValueError(msg)
    
    # Convertir tipos y combinar Debito/Credito si es necesario
    df['Fecha'] = parse_fecha_ddmmyyyy(df['Fecha'])
    df['Concepto'] = df['Concepto'].fillna('').astype(str)
    df['Descripci√≥n'] = df['Descripci√≥n'].fillna('').astype(str)
    
    if tiene_debito_credito:
        # Asegurar que ambas columnas existen en el DataFrame (llenar con 0 si faltan)
        if 'Debito' not in df.columns:
            df['Debito'] = 0
        if 'Credito' not in df.columns:
            df['Credito'] = 0
        # Convertir Debito y Credito a num√©ricos
        df['Debito'] = pd.to_numeric(df['Debito'], errors='coerce').fillna(0)
        df['Credito'] = pd.to_numeric(df['Credito'], errors='coerce').fillna(0)
        # Crear Valor: Credito positivo, Debito negativo
        df['Valor'] = df['Credito'] - df['Debito']
        # Eliminar columnas originales si exist√≠an
        drop_cols = [c for c in ['Debito', 'Credito'] if c in df.columns]
        if drop_cols:
            df = df.drop(columns=drop_cols)
    else:
        df['Valor'] = pd.to_numeric(df['Valor'], errors='coerce')
    
    # ‚ö° OPTIMIZACI√ìN 2: Filtrar SOLO filas con Fecha Y Valor v√°lidos
    registros_antes = len(df)
    df = df.dropna(subset=['Fecha', 'Valor'])
    df = df[df['Valor'] != 0]  # Eliminar valores $0
    registros_despues = len(df)
    
    if registros_antes != registros_despues:
        print(f"  ‚ö†Ô∏è Filtradas {registros_antes - registros_despues} filas sin Fecha/Valor v√°lidos")
    
    # Campos de b√∫squeda
    df['Texto_Busqueda'] = (df['Concepto'].astype(str) + ' ' + df['Descripci√≥n'].astype(str)).apply(normalizar_texto)
    df['Concepto_Norm'] = df['Concepto'].apply(normalizar_texto)  # ‚Üê RESTAURADO del v5
    df['Proveedor_ID'] = df['Descripci√≥n'].apply(extraer_identificador_proveedor)
    df['Empresa_Norm'] = df['Texto_Busqueda'].apply(normalizar_nombre_empresa)
    df['Es_Impuesto'] = df['Texto_Busqueda'].apply(es_patron_impuesto)
    df['Es_Comision'] = df['Texto_Busqueda'].apply(es_patron_comision)
    df['Es_TC'] = df['Texto_Busqueda'].apply(es_patron_tc)  # ‚Üê NUEVO v6.5
    df['Es_Caja_Chica'] = df['Texto_Busqueda'].apply(es_patron_caja_chica)  # ‚Üê NUEVO v6.6
    
    # Control
    df['ID_Original'] = range(len(df))
    df['Conciliado'] = False
    
    print(f"  ‚úÖ Cargados: {len(df)} registros v√°lidos")
    return df

# Cargas de Archivos CONTABLE
def cargar_contable(ruta, usa_dolares, nombre=""):
    """Carga datos - OPTIMIZACI√ìN: solo lee filas con Fecha Y Valor v√°lidos"""
    print(f"\n  üìÇ Cargando: {nombre.upper() if nombre else ruta}")

    # Resolver ruta relativa: si el usuario pasa solo el nombre de archivo,
    # buscarlo en el mismo directorio del script y en el cwd.
    try:
        ruta_path = Path(ruta)
    except Exception:
        ruta_path = Path(str(ruta))

    if not ruta_path.is_absolute():
        script_dir = Path(__file__).resolve().parent
        candidates = [script_dir / ruta_path, Path.cwd() / ruta_path]

        # If no suffix provided, try common extensions too
        if not ruta_path.suffix:
            exts = ['.xlsx', '.xls', '.csv']
        else:
            exts = [ruta_path.suffix]

        found = None
        for base in candidates:
            for ext in exts:
                cand = base.with_suffix(ext) if not ruta_path.suffix else base
                if cand.exists():
                    found = cand
                    break
            if found:
                break

        if found:
            ruta = str(found)
        else:
            # if the literal provided (e.g., 'BANRESERVAS.xlsx') exists in script_dir
            alt = script_dir / ruta_path.name
            if alt.exists():
                ruta = str(alt)

    try:
        # Usar la limpieza ligera para el libro contable (detecta y corta
        # filas no-datos encima de la tabla principal). Esta funci√≥n lee el
        # archivo y devuelve un DataFrame ya recortado.
        moneda_str = 'USD' if usa_dolares else 'RD$'
        df = limpiar_archivo_contable(ruta, moneda_str)
    except Exception as e:
        raise ValueError(f"‚ùå Error al cargar contable '{ruta}': {e}")
    
    print(f"  üìã Columnas encontradas: {list(df.columns)}")
    
    # Mapeo autom√°tico de columnas (normaliza espacios/acentos al buscar)
    columnas_map = {}
    for col in df.columns:
        col_upper = str(col).upper().strip()
        if any(x in col_upper for x in ['F.COMP', 'FECHA']) and 'Fecha' not in columnas_map:
            columnas_map['Fecha'] = col
        elif any(x in col_upper for x in ['DETALLE', 'CONCEPTO', 'CONCEP', 'OBSERV', 'REFERENCIA', 'BENEFICIARIO', 'BENEF']) and 'Concepto' not in columnas_map:
            columnas_map['Concepto'] = col
        elif 'NATU' in col_upper and 'Natu' not in columnas_map:
            columnas_map['Natu'] = col

        # Detectar "Valor Moneda Extranjera" para USD (tiene signos ya aplicados)
        elif 'EXTRANJERA' in col_upper and 'Valor_USD' not in columnas_map:
                columnas_map['Valor_USD'] = col

        # Valor (√∫nico) - detecta nombres comunes (incluye variantes en ingl√©s)
        elif any(x in col_upper for x in ['VALOR', 'MONTO', 'IMPORTE', 'AMOUNT', 'AMT']) and 'Valor' not in columnas_map:
            columnas_map['Valor'] = col
        
        # Descripci√≥n
        elif any(x in col_upper for x in ['DESCRIP', 'DETALLE', 'OBSERV', 'REFERENCIA', 'BENEFICIARIO', 'BENEF']) and 'Descripci√≥n' not in columnas_map:
            columnas_map['Descripci√≥n'] = col
    
    # Detectar libro contable YA LIMPIO
    es_limpio = (
        'Valor' in df.columns and
        'Natu' not in df.columns and
        'Valor_USD' not in df.columns
    )

    # Detectar si hay columna de naturaleza (Natu)
    tiene_natu = 'Natu' in columnas_map

    # Si no hay Descripci√≥n, crearla vac√≠a
    if 'Descripci√≥n' not in columnas_map:
        df['Descripci√≥n'] = ''
        columnas_map['Descripci√≥n'] = 'Descripci√≥n'
        print("  ‚ö†Ô∏è Columna 'Descripci√≥n' no encontrada - creada vac√≠a")
    
    # Verificar columnas requeridas seg√∫n moneda
    if usa_dolares:
        # USD: solo necesita Valor_USD
        requeridas = ['Fecha', 'Concepto', 'Valor_USD']
    elif tiene_natu:
        # RD: necesita Valor y Natu
        requeridas = ['Fecha', 'Concepto', 'Valor', 'Natu']
    else: # Chequeo de documento limpiado
        # Sin Natu: necesita solo Valor
        requeridas = ['Fecha', 'Concepto', 'Valor']
    
    faltantes = [r for r in requeridas if r not in columnas_map]
    if faltantes:
        raise ValueError(f"‚ùå Faltan columnas requeridas: {faltantes}")
    
    # Renombrar columnas
    df = df.rename(columns={v: k for k, v in columnas_map.items()})
    
    # ‚ö° OPTIMIZACI√ìN 3: Eliminar filas completamente vac√≠as (espacios en blanco bajo encabezados)
    df = df.dropna(how='all')
    df = df.reset_index(drop=True)
    
    # üö´ FILTRO ESTADO: Ignorar filas marcadas como "A" o "X" en la columna Estado
    if 'Estado' in df.columns:
        registros_antes_estado = len(df)
        # Convertir Estado a string y eliminar filas con A o X
        df['Estado'] = df['Estado'].fillna('').astype(str).str.strip().str.upper()
        df = df[~df['Estado'].isin(['A', 'X'])]
        registros_eliminados_estado = registros_antes_estado - len(df)
        if registros_eliminados_estado > 0:
            print(f"  üö´ Eliminadas {registros_eliminados_estado} filas con Estado 'A' o 'X'")
        df = df.drop(columns=['Estado'], errors='ignore')  # Eliminar la columna Estado despu√©s de filtrar
    
    # Convertir tipos y combinar/ajustar columnas seg√∫n formato
    df['Fecha'] = parse_fecha_ddmmyyyy(df['Fecha'])
    df['Concepto'] = df['Concepto'].fillna('').astype(str)
    df['Descripci√≥n'] = df['Descripci√≥n'].fillna('').astype(str)

    # Procesar Valor seg√∫n moneda detectada
    if es_limpio:
        print("  üßº Libro contable ya limpio ‚Äî usando Valor directamente")
        df['Valor'] = pd.to_numeric(df['Valor'], errors='coerce')
    
    elif usa_dolares == True and 'Valor_USD' in df.columns:
        # USD: usar Valor Moneda Extranjera directamente (ya tiene signos correctos)
        print("  üíµ Detectada moneda USD - usando 'Valor Moneda Extranjera'")
        # Descartar columnas RD primero
        if 'Valor' in df.columns:
            df = df.drop(columns=['Valor'], errors='ignore')
        if 'Natu' in df.columns:
            df = df.drop(columns=['Natu'], errors='ignore')
        # Luego usar USD como Valor
        df['Valor'] = pd.to_numeric(df['Valor_USD'], errors='coerce')
        df = df.drop(columns=['Valor_USD'], errors='ignore')
    
    elif tiene_natu == True and 'Valor' in df.columns:
        # RD: usar Valor aplicando interpretaci√≥n de NATU
        print("  Detectada moneda RD - usando 'Valor' interpretando 'Natu'")
        df['Valor'] = pd.to_numeric(df['Valor'], errors='coerce')

        if 'Natu' in df.columns:
            df['Natu'] = df['Natu'].astype(str).str.strip().str.upper()

            # üîπ APLICAR SIGNO SEG√öN NATU
            df.loc[df['Natu'] == 'E', 'Valor'] *= -1
            df.loc[df['Natu'] == 'I', 'Valor'] = df.loc[df['Natu'] == 'I', 'Valor'].abs()

            df = df.drop(columns=['Natu'])

        # Descartar columna USD si existe
        if 'Valor_USD' in df.columns:
            df = df.drop(columns=['Valor_USD'], errors='ignore')
    
    # ‚ö° OPTIMIZACI√ìN 2: Filtrar SOLO filas con Fecha Y Valor v√°lidos
    registros_antes = len(df)
    df = df.dropna(subset=['Fecha', 'Valor'])
    df = df[df['Valor'] != 0]  # Eliminar valores $0
    registros_despues = len(df)
    
    if registros_antes != registros_despues:
        print(f"  ‚ö†Ô∏è Filtradas {registros_antes - registros_despues} filas sin Fecha/Valor v√°lidos")
    
    # Campos de b√∫squeda
    df['Texto_Busqueda'] = (df['Concepto'].astype(str) + ' ' + df['Descripci√≥n'].astype(str)).apply(normalizar_texto)
    df['Concepto_Norm'] = df['Concepto'].apply(normalizar_texto)  # ‚Üê RESTAURADO del v5
    df['Proveedor_ID'] = df['Descripci√≥n'].apply(extraer_identificador_proveedor)
    df['Empresa_Norm'] = df['Texto_Busqueda'].apply(normalizar_nombre_empresa)
    df['Es_Impuesto'] = df['Texto_Busqueda'].apply(es_patron_impuesto)
    df['Es_Comision'] = df['Texto_Busqueda'].apply(es_patron_comision)
    df['Es_TC'] = df['Texto_Busqueda'].apply(es_patron_tc)  # ‚Üê NUEVO v6.5
    
    # Control
    df['ID_Original'] = range(len(df))
    df['Conciliado'] = False
    
    print(f"  ‚úÖ Cargados: {len(df)} registros v√°lidos")
    return df

# ============================================================================
# üéØ ESTRATEGIA 1: MONTO EXACTO (1:1) - CON SCORE COMBINADO
# ============================================================================

def conciliacion_por_monto_exacto(banco, contable, conciliaciones):
    print("\n" + "="*70)
    print("üéØ ESTRATEGIA 1: MONTO EXACTO (1:1)")
    print("="*70)
    
    contador = 0
    id_conc = len(conciliaciones) + 1
    
    for idx_b, reg_b in banco[~banco['Conciliado']].iterrows():
        f_min = reg_b['Fecha'] - timedelta(days=VENTANA_DIAS_EXACTA)
        f_max = reg_b['Fecha'] + timedelta(days=VENTANA_DIAS_EXACTA)
        
        cands = contable[
            (~contable['Conciliado']) &
            (contable['Fecha'] >= f_min) &
            (contable['Fecha'] <= f_max) &
            (abs(contable['Valor'] - reg_b['Valor']) < TOLERANCIA_VALOR_EXACTA)
        ]
        
        if len(cands) == 0:
            continue
        
        if len(cands) == 1:
            reg_c = cands.iloc[0]
            sim = calcular_similitud(reg_b['Texto_Busqueda'], reg_c['Texto_Busqueda'])
            
            if PERMITIR_SOLO_MONTO or sim >= UMBRAL_SIMILITUD_BAJA:
                conciliaciones.append({
                    'ID_Conciliacion': id_conc,
                    'Tipo': '1. Monto Exacto (1:1)',
                    'Similitud': round(sim, 3),
                    'Fecha_Banco': reg_b['Fecha'],
                    'Concepto_Banco': reg_b['Concepto'],
                    'Valor_Banco': reg_b['Valor'],
                    'Descripcion_Banco': reg_b['Descripci√≥n'],
                    'Fecha_Contable': reg_c['Fecha'],
                    'Concepto_Contable': reg_c['Concepto'],
                    'Valor_Contable': reg_c['Valor'],
                    'Descripcion_Contable': reg_c['Descripci√≥n'],
                    'Diferencia': abs(reg_b['Valor'] - reg_c['Valor'])
                })
                banco.loc[idx_b, 'Conciliado'] = True
                contable.loc[reg_c.name, 'Conciliado'] = True
                contador += 1
                id_conc += 1
        else:
            # ‚≠ê SCORE COMBINADO para desempatar (restaurado del v5)
            cands = cands.copy()
            cands['Sim'] = cands['Texto_Busqueda'].apply(
                lambda x: calcular_similitud(reg_b['Texto_Busqueda'], x)
            )
            cands['Dias_Diff'] = abs((cands['Fecha'] - reg_b['Fecha']).dt.days)
            cands['Score'] = (cands['Sim'] * 0.7) + ((VENTANA_DIAS_EXACTA - cands['Dias_Diff']) / VENTANA_DIAS_EXACTA * 0.3)
            
            mejor_idx = cands['Score'].idxmax()
            mejor_sim = cands.loc[mejor_idx, 'Sim']
            
            if PERMITIR_SOLO_MONTO or mejor_sim >= UMBRAL_SIMILITUD_BAJA:
                reg_c = contable.loc[mejor_idx]
                conciliaciones.append({
                    'ID_Conciliacion': id_conc,
                    'Tipo': '1. Monto Exacto (1:1)',
                    'Similitud': round(mejor_sim, 3),
                    'Fecha_Banco': reg_b['Fecha'],
                    'Concepto_Banco': reg_b['Concepto'],
                    'Valor_Banco': reg_b['Valor'],
                    'Descripcion_Banco': reg_b['Descripci√≥n'],
                    'Fecha_Contable': reg_c['Fecha'],
                    'Concepto_Contable': reg_c['Concepto'],
                    'Valor_Contable': reg_c['Valor'],
                    'Descripcion_Contable': reg_c['Descripci√≥n'],
                    'Diferencia': abs(reg_b['Valor'] - reg_c['Valor'])
                })
                banco.loc[idx_b, 'Conciliado'] = True
                contable.loc[mejor_idx, 'Conciliado'] = True
                contador += 1
                id_conc += 1
    
    print(f"‚úì Conciliaciones 1:1: {contador}")
    return contador

# ============================================================================
# üéØ ESTRATEGIA 1.5: TRANSFERENCIAS CON COMISI√ìN ($7)
# ============================================================================

def conciliacion_con_comisiones(banco, contable, conciliaciones):
    if not DETECTAR_COMISIONES:
        return 0
    
    print("\n" + "="*70)
    print("üéØ ESTRATEGIA 1.5: TRANSFERENCIAS CON COMISI√ìN ($7)")
    print("="*70)
    
    contador = 0
    id_conc = len(conciliaciones) + 1 if conciliaciones else 1
    
    banco_no_conc = banco[~banco['Conciliado']].copy()
    
    for idx_com, reg_com in banco_no_conc[banco_no_conc['Es_Comision']].iterrows():
        if abs(abs(reg_com['Valor']) - COMISION_TRANSFERENCIA_USD) > 0.50:
            continue
        
        fecha_com = reg_com['Fecha']
        candidatos_transf = banco_no_conc[
            (~banco_no_conc['Conciliado']) &
            (banco_no_conc['Fecha'] == fecha_com) &
            (banco_no_conc['Valor'] > 0) &
            (banco_no_conc.index != idx_com)
        ]
        
        for idx_transf, reg_transf in candidatos_transf.iterrows():
            valor_neto = reg_transf['Valor'] - COMISION_TRANSFERENCIA_USD
            
            f_min = fecha_com - timedelta(days=VENTANA_DIAS_EXACTA)
            f_max = fecha_com + timedelta(days=VENTANA_DIAS_EXACTA)
            
            cands_cont = contable[
                (~contable['Conciliado']) &
                (contable['Fecha'] >= f_min) &
                (contable['Fecha'] <= f_max) &
                (abs(contable['Valor'] - valor_neto) < TOLERANCIA_VALOR_EXACTA)
            ]
            
            if len(cands_cont) >= 1:
                reg_c = cands_cont.iloc[0]
                sim = calcular_similitud(reg_transf['Texto_Busqueda'], reg_c['Texto_Busqueda'])
                
                conciliaciones.append({
                    'ID_Conciliacion': id_conc,
                    'Tipo': '1.5. Transf+Comisi√≥n (2‚Üí1)',
                    'Similitud': round(sim, 3),
                    'Fecha_Banco': reg_com['Fecha'],
                    'Concepto_Banco': reg_com['Concepto'],
                    'Valor_Banco': reg_com['Valor'],
                    'Descripcion_Banco': reg_com['Descripci√≥n'],
                    'Fecha_Contable': None,
                    'Concepto_Contable': '(Comisi√≥n bancaria)',
                    'Valor_Contable': None,
                    'Descripcion_Contable': '',
                    'Diferencia': None
                })
                
                conciliaciones.append({
                    'ID_Conciliacion': id_conc,
                    'Tipo': '1.5. Transf+Comisi√≥n (2‚Üí1)',
                    'Similitud': round(sim, 3),
                    'Fecha_Banco': reg_transf['Fecha'],
                    'Concepto_Banco': reg_transf['Concepto'],
                    'Valor_Banco': reg_transf['Valor'],
                    'Descripcion_Banco': reg_transf['Descripci√≥n'],
                    'Fecha_Contable': reg_c['Fecha'],
                    'Concepto_Contable': reg_c['Concepto'],
                    'Valor_Contable': reg_c['Valor'],
                    'Descripcion_Contable': reg_c['Descripci√≥n'],
                    'Diferencia': COMISION_TRANSFERENCIA_USD
                })
                
                banco.loc[idx_com, 'Conciliado'] = True
                banco.loc[idx_transf, 'Conciliado'] = True
                contable.loc[reg_c.name, 'Conciliado'] = True
                
                contador += 1
                id_conc += 1
                break
    
    print(f"‚úì Transferencias con comisi√≥n: {contador}")
    return contador

# ============================================================================
# üéØ ESTRATEGIA 1.6: COMISIONES AGRUPADAS MULTI-FECHA [v6.1]
# ============================================================================

def conciliacion_comisiones_agrupadas(banco, contable, conciliaciones):
    print("\n" + "="*70)
    print("üéØ ESTRATEGIA 1.6: COMISIONES AGRUPADAS (Multi-fecha)")
    print("="*70)
    
    contador = 0
    id_conc = len(conciliaciones) + 1 if conciliaciones else 1
    
    contable_com = contable[(~contable['Conciliado']) & (contable['Es_Comision'])].copy()
    
    if len(contable_com) == 0:
        print("‚äò No hay comisiones pendientes en contable")
        return 0
    
    banco_com = banco[(~banco['Conciliado']) & (banco['Es_Comision'])].copy()
    
    if len(banco_com) == 0:
        print("‚äò No hay comisiones pendientes en banco")
        return 0
    
    for idx_c, reg_c in contable_com.iterrows():
        if contable.loc[idx_c, 'Conciliado']:
            continue
        
        valor_objetivo = reg_c['Valor']
        
        banco_com_pend = banco_com[~banco_com.index.isin(banco[banco['Conciliado']].index)]
        if len(banco_com_pend) == 0:
            continue
        
        # Ventana amplia para comisiones dispersas
        mes_contable = reg_c['Fecha'].month if pd.notna(reg_c['Fecha']) else None
        if mes_contable:
            banco_com_mes = banco_com_pend[
                (banco_com_pend['Fecha'].dt.month >= mes_contable - 1) &
                (banco_com_pend['Fecha'].dt.month <= mes_contable + 1)
            ]
        else:
            banco_com_mes = banco_com_pend
        
        if len(banco_com_mes) == 0:
            continue
        
        # B√∫squeda combinatoria LIMITADA
        mejor_combo, mejor_diff = None, float('inf')
        indices = banco_com_mes.index.tolist()[:MAX_PARTIDAS_AGRUPACION]
        
        total_combinaciones = 0
        for r in range(1, len(indices) + 1):
            if mejor_diff < TOLERANCIA_VALOR_AGRUPACION:
                break
            if total_combinaciones >= MAX_COMBINACIONES_POR_BUSQUEDA:
                break
            
            for combo in combinations(indices, r):
                total_combinaciones += 1
                if total_combinaciones > MAX_COMBINACIONES_POR_BUSQUEDA:
                    break
                
                suma = banco.loc[list(combo), 'Valor'].sum()
                diff = abs(suma - valor_objetivo)
                
                if diff < mejor_diff:
                    mejor_diff = diff
                    mejor_combo = combo
                
                if diff < TOLERANCIA_VALOR_AGRUPACION:
                    break
        
        if mejor_combo and mejor_diff < TOLERANCIA_VALOR_AGRUPACION:
            grupo = banco.loc[list(mejor_combo)]
            
            for i, (idx_b, reg_b) in enumerate(grupo.sort_values(['Fecha', 'Valor']).iterrows()):
                conciliaciones.append({
                    'ID_Conciliacion': id_conc,
                    'Tipo': f'1.6. Comisiones Agrupadas ({len(grupo)}‚Üí1)',
                    'Similitud': 1.0,
                    'Fecha_Banco': reg_b['Fecha'],
                    'Concepto_Banco': reg_b['Concepto'],
                    'Valor_Banco': reg_b['Valor'],
                    'Descripcion_Banco': reg_b['Descripci√≥n'],
                    'Fecha_Contable': reg_c['Fecha'] if i == 0 else None,
                    'Concepto_Contable': reg_c['Concepto'] if i == 0 else '',
                    'Valor_Contable': reg_c['Valor'] if i == 0 else None,
                    'Descripcion_Contable': reg_c['Descripci√≥n'] if i == 0 else '',
                    'Diferencia': mejor_diff if i == 0 else None
                })
            
            banco.loc[list(mejor_combo), 'Conciliado'] = True
            contable.loc[idx_c, 'Conciliado'] = True
            contador += 1
            id_conc += 1
    
    print(f"‚úì Comisiones agrupadas: {contador}")
    return contador

# ============================================================================
# üéØ ESTRATEGIA 2: N‚Üí1 INTELIGENTE (4 M√âTODOS) - RESTAURADO DEL v5
# ============================================================================

def conciliacion_n_a_1_inteligente(banco, contable, conciliaciones):
    print("\n" + "="*70)
    print("üéØ ESTRATEGIA 2: N‚Üí1 (Varios Contable ‚Üí Uno Banco)")
    print("="*70)
    
    contador = 0
    id_conc = len(conciliaciones) + 1 if conciliaciones else 1
    
    for idx_b, reg_b in banco[~banco['Conciliado']].iterrows():
        if banco.loc[idx_b, 'Conciliado']:
            continue
            
        f_min = reg_b['Fecha'] - timedelta(days=VENTANA_DIAS_AGRUPACION)
        f_max = reg_b['Fecha'] + timedelta(days=VENTANA_DIAS_AGRUPACION)
        
        cands = contable[
            (~contable['Conciliado']) &
            (contable['Fecha'] >= f_min) &
            (contable['Fecha'] <= f_max)
        ].copy()
        
        if len(cands) < 2:
            continue
        
        encontrado = False
        
        # M√âTODO 1: Por Proveedor_ID con desambiguaci√≥n por fechas
        if USAR_FECHAS_PARA_DESAMBIGUAR and 'Proveedor_ID' in cands.columns:
            for prov_id in cands['Proveedor_ID'].dropna().unique():
                if encontrado:
                    break
                    
                grupo_prov = cands[cands['Proveedor_ID'] == prov_id]
                if len(grupo_prov) < 2:
                    continue
                
                for fecha_grupo in grupo_prov['Fecha'].unique():
                    grupo = grupo_prov[
                        (grupo_prov['Fecha'] >= fecha_grupo - timedelta(days=3)) &
                        (grupo_prov['Fecha'] <= fecha_grupo + timedelta(days=3))
                    ]
                    
                    if len(grupo) < 2:
                        continue
                    
                    suma = grupo['Valor'].sum()
                    if abs(suma - reg_b['Valor']) < TOLERANCIA_VALOR_AGRUPACION:
                        sim_media = grupo['Texto_Busqueda'].apply(
                            lambda x: calcular_similitud(reg_b['Texto_Busqueda'], x)
                        ).mean()
                        
                        for i, (_, reg_c) in enumerate(grupo.sort_values(['Fecha', 'Valor']).iterrows()):
                            conciliaciones.append({
                                'ID_Conciliacion': id_conc,
                                'Tipo': f'2. Agrupaci√≥n N‚Üí1 ({len(grupo)} reg)',
                                'Similitud': round(sim_media, 3),
                                'Fecha_Banco': reg_b['Fecha'] if i == 0 else None,
                                'Concepto_Banco': reg_b['Concepto'] if i == 0 else '',
                                'Valor_Banco': reg_b['Valor'] if i == 0 else None,
                                'Descripcion_Banco': reg_b['Descripci√≥n'] if i == 0 else '',
                                'Fecha_Contable': reg_c['Fecha'],
                                'Concepto_Contable': reg_c['Concepto'],
                                'Valor_Contable': reg_c['Valor'],
                                'Descripcion_Contable': reg_c['Descripci√≥n'],
                                'Diferencia': abs(reg_b['Valor'] - suma) if i == 0 else None
                            })
                        
                        banco.loc[idx_b, 'Conciliado'] = True
                        contable.loc[grupo.index, 'Conciliado'] = True
                        contador += 1
                        id_conc += 1
                        encontrado = True
                        break
        
        if encontrado:
            continue
        
        # ‚≠ê M√âTODO 2: Por Concepto_Norm (RESTAURADO del v5)
        for concepto in cands['Concepto_Norm'].unique():
            if encontrado or not concepto:
                break
            
            grupo = cands[cands['Concepto_Norm'] == concepto]
            if len(grupo) < 2:
                continue
            
            suma = grupo['Valor'].sum()
            if abs(suma - reg_b['Valor']) < TOLERANCIA_VALOR_AGRUPACION:
                sim_media = grupo['Texto_Busqueda'].apply(
                    lambda x: calcular_similitud(reg_b['Texto_Busqueda'], x)
                ).mean()
                
                for i, (_, reg_c) in enumerate(grupo.sort_values(['Fecha', 'Valor']).iterrows()):
                    conciliaciones.append({
                        'ID_Conciliacion': id_conc,
                        'Tipo': f'2. Agrupaci√≥n N‚Üí1 ({len(grupo)} reg)',
                        'Similitud': round(sim_media, 3),
                        'Fecha_Banco': reg_b['Fecha'] if i == 0 else None,
                        'Concepto_Banco': reg_b['Concepto'] if i == 0 else '',
                        'Valor_Banco': reg_b['Valor'] if i == 0 else None,
                        'Descripcion_Banco': reg_b['Descripci√≥n'] if i == 0 else '',
                        'Fecha_Contable': reg_c['Fecha'],
                        'Concepto_Contable': reg_c['Concepto'],
                        'Valor_Contable': reg_c['Valor'],
                        'Descripcion_Contable': reg_c['Descripci√≥n'],
                        'Diferencia': abs(reg_b['Valor'] - suma) if i == 0 else None
                    })
                
                banco.loc[idx_b, 'Conciliado'] = True
                contable.loc[grupo.index, 'Conciliado'] = True
                contador += 1
                id_conc += 1
                encontrado = True
        
        if encontrado:
            continue
        
        # M√âTODO 3: Por empresa normalizada
        for emp_norm in cands['Empresa_Norm'].unique():
            if encontrado or not emp_norm:
                break
                
            grupo = cands[cands['Empresa_Norm'] == emp_norm]
            if len(grupo) < 2:
                continue
            
            suma = grupo['Valor'].sum()
            if abs(suma - reg_b['Valor']) < TOLERANCIA_VALOR_AGRUPACION:
                sim_media = grupo['Texto_Busqueda'].apply(
                    lambda x: calcular_similitud(reg_b['Texto_Busqueda'], x)
                ).mean()
                
                for i, (_, reg_c) in enumerate(grupo.sort_values(['Fecha', 'Valor']).iterrows()):
                    conciliaciones.append({
                        'ID_Conciliacion': id_conc,
                        'Tipo': f'2. Agrupaci√≥n N‚Üí1 ({len(grupo)} reg)',
                        'Similitud': round(sim_media, 3),
                        'Fecha_Banco': reg_b['Fecha'] if i == 0 else None,
                        'Concepto_Banco': reg_b['Concepto'] if i == 0 else '',
                        'Valor_Banco': reg_b['Valor'] if i == 0 else None,
                        'Descripcion_Banco': reg_b['Descripci√≥n'] if i == 0 else '',
                        'Fecha_Contable': reg_c['Fecha'],
                        'Concepto_Contable': reg_c['Concepto'],
                        'Valor_Contable': reg_c['Valor'],
                        'Descripcion_Contable': reg_c['Descripci√≥n'],
                        'Diferencia': abs(reg_b['Valor'] - suma) if i == 0 else None
                    })
                
                banco.loc[idx_b, 'Conciliado'] = True
                contable.loc[grupo.index, 'Conciliado'] = True
                contador += 1
                id_conc += 1
                encontrado = True
        
        if encontrado:
            continue
        
        # M√âTODO 4: Por palabras clave
        cands['PalKey'] = cands['Texto_Busqueda'].apply(
            lambda x: ' '.join(sorted(extraer_palabras_clave(x)))
        )
        
        for key, grupo in cands.groupby('PalKey'):
            if len(grupo) < 2 or not key:
                continue
            
            suma = grupo['Valor'].sum()
            if abs(suma - reg_b['Valor']) < TOLERANCIA_VALOR_AGRUPACION:
                sim_media = grupo['Texto_Busqueda'].apply(
                    lambda x: calcular_similitud(reg_b['Texto_Busqueda'], x)
                ).mean()
                
                if sim_media >= UMBRAL_SIMILITUD_BAJA:
                    for i, (_, reg_c) in enumerate(grupo.sort_values(['Fecha', 'Valor']).iterrows()):
                        conciliaciones.append({
                            'ID_Conciliacion': id_conc,
                            'Tipo': f'2. Agrupaci√≥n N‚Üí1 ({len(grupo)} reg)',
                            'Similitud': round(sim_media, 3),
                            'Fecha_Banco': reg_b['Fecha'] if i == 0 else None,
                            'Concepto_Banco': reg_b['Concepto'] if i == 0 else '',
                            'Valor_Banco': reg_b['Valor'] if i == 0 else None,
                            'Descripcion_Banco': reg_b['Descripci√≥n'] if i == 0 else '',
                            'Fecha_Contable': reg_c['Fecha'],
                            'Concepto_Contable': reg_c['Concepto'],
                            'Valor_Contable': reg_c['Valor'],
                            'Descripcion_Contable': reg_c['Descripci√≥n'],
                            'Diferencia': abs(reg_b['Valor'] - suma) if i == 0 else None
                        })
                    
                    banco.loc[idx_b, 'Conciliado'] = True
                    contable.loc[grupo.index, 'Conciliado'] = True
                    contador += 1
                    id_conc += 1
                    break
    
    print(f"‚úì Agrupaciones N‚Üí1: {contador}")
    return contador

# ============================================================================
# üéØ ESTRATEGIA 3: 1‚ÜíN
# ============================================================================

def conciliacion_1_a_n_inteligente(banco, contable, conciliaciones):
    print("\n" + "="*70)
    print("üéØ ESTRATEGIA 3: 1‚ÜíN (Uno Contable ‚Üí Varios Banco)")
    print("="*70)
    
    contador = 0
    id_conc = len(conciliaciones) + 1 if conciliaciones else 1
    
    for idx_c, reg_c in contable[~contable['Conciliado']].iterrows():
        if contable.loc[idx_c, 'Conciliado']:
            continue
            
        f_min = reg_c['Fecha'] - timedelta(days=VENTANA_DIAS_AGRUPACION)
        f_max = reg_c['Fecha'] + timedelta(days=VENTANA_DIAS_AGRUPACION)
        
        cands = banco[
            (~banco['Conciliado']) &
            (banco['Fecha'] >= f_min) &
            (banco['Fecha'] <= f_max)
        ].copy()
        
        if len(cands) < 2:
            continue
        
        cands['PalKey'] = cands['Texto_Busqueda'].apply(
            lambda x: ' '.join(sorted(extraer_palabras_clave(x)))
        )
        
        for key, grupo in cands.groupby('PalKey'):
            if len(grupo) < 2 or not key:
                continue
            
            suma = grupo['Valor'].sum()
            if abs(suma - reg_c['Valor']) < TOLERANCIA_VALOR_AGRUPACION:
                sim_media = grupo['Texto_Busqueda'].apply(
                    lambda x: calcular_similitud(reg_c['Texto_Busqueda'], x)
                ).mean()
                
                if sim_media >= UMBRAL_SIMILITUD_BAJA:
                    for i, (_, reg_b) in enumerate(grupo.sort_values(['Fecha', 'Valor']).iterrows()):
                        conciliaciones.append({
                            'ID_Conciliacion': id_conc,
                            'Tipo': f'3. Agrupaci√≥n 1‚ÜíN ({len(grupo)} reg)',
                            'Similitud': round(sim_media, 3),
                            'Fecha_Banco': reg_b['Fecha'],
                            'Concepto_Banco': reg_b['Concepto'],
                            'Valor_Banco': reg_b['Valor'],
                            'Descripcion_Banco': reg_b['Descripci√≥n'],
                            'Fecha_Contable': reg_c['Fecha'] if i == 0 else None,
                            'Concepto_Contable': reg_c['Concepto'] if i == 0 else '',
                            'Valor_Contable': reg_c['Valor'] if i == 0 else None,
                            'Descripcion_Contable': reg_c['Descripci√≥n'] if i == 0 else '',
                            'Diferencia': abs(suma - reg_c['Valor']) if i == 0 else None
                        })
                    
                    contable.loc[idx_c, 'Conciliado'] = True
                    banco.loc[grupo.index, 'Conciliado'] = True
                    contador += 1
                    id_conc += 1
                    break
    
    print(f"‚úì Agrupaciones 1‚ÜíN: {contador}")
    return contador

# ============================================================================
# üéØ ESTRATEGIA 4: N‚ÜîM
# ============================================================================

def conciliacion_n_a_m_inteligente(banco, contable, conciliaciones):
    print("\n" + "="*70)
    print("üéØ ESTRATEGIA 4: N‚ÜîM (Varios ‚Üî Varios)")
    print("="*70)
    
    contador = 0
    id_conc = len(conciliaciones) + 1 if conciliaciones else 1
    
    cont_temp = contable[~contable['Conciliado']].copy()
    cont_temp['PalKey'] = cont_temp['Texto_Busqueda'].apply(
        lambda x: ' '.join(sorted(extraer_palabras_clave(x)))
    )
    grupos_c = cont_temp.groupby('PalKey').filter(lambda x: len(x) >= 2)
    
    for key_c, grupo_c in grupos_c.groupby('PalKey'):
        if len(grupo_c) < 2 or not key_c:
            continue
        
        grupo_c = grupo_c[~grupo_c.index.isin(contable[contable['Conciliado']].index)]
        if len(grupo_c) < 2:
            continue
        
        suma_c = grupo_c['Valor'].sum()
        f_media = grupo_c['Fecha'].mean()
        f_min = f_media - timedelta(days=VENTANA_DIAS_AGRUPACION)
        f_max = f_media + timedelta(days=VENTANA_DIAS_AGRUPACION)
        
        cands_b = banco[
            (~banco['Conciliado']) &
            (banco['Fecha'] >= f_min) &
            (banco['Fecha'] <= f_max)
        ].copy()
        
        if len(cands_b) < 2:
            continue
        
        cands_b['PalKey'] = cands_b['Texto_Busqueda'].apply(
            lambda x: ' '.join(sorted(extraer_palabras_clave(x)))
        )
        
        for key_b, grupo_b in cands_b.groupby('PalKey'):
            if len(grupo_b) < 2 or not key_b:
                continue
            
            suma_b = grupo_b['Valor'].sum()
            
            if abs(suma_b - suma_c) < TOLERANCIA_VALOR_AGRUPACION:
                sim_total = sum(
                    calcular_similitud(rb['Texto_Busqueda'], rc['Texto_Busqueda'])
                    for _, rb in grupo_b.iterrows()
                    for _, rc in grupo_c.iterrows()
                )
                sim_media = sim_total / (len(grupo_b) * len(grupo_c))
                
                if sim_media >= UMBRAL_SIMILITUD_BAJA:
                    regs_b = grupo_b.sort_values(['Fecha', 'Valor'])
                    regs_c = grupo_c.sort_values(['Fecha', 'Valor'])
                    max_regs = max(len(regs_b), len(regs_c))
                    
                    for i in range(max_regs):
                        rb = regs_b.iloc[i] if i < len(regs_b) else None
                        rc = regs_c.iloc[i] if i < len(regs_c) else None
                        
                        conciliaciones.append({
                            'ID_Conciliacion': id_conc,
                            'Tipo': f'4. Agrupaci√≥n N‚ÜîM ({len(regs_b)}‚Üî{len(regs_c)})',
                            'Similitud': round(sim_media, 3),
                            'Fecha_Banco': rb['Fecha'] if rb is not None else None,
                            'Concepto_Banco': rb['Concepto'] if rb is not None else '',
                            'Valor_Banco': rb['Valor'] if rb is not None else None,
                            'Descripcion_Banco': rb['Descripci√≥n'] if rb is not None else '',
                            'Fecha_Contable': rc['Fecha'] if rc is not None else None,
                            'Concepto_Contable': rc['Concepto'] if rc is not None else '',
                            'Valor_Contable': rc['Valor'] if rc is not None else None,
                            'Descripcion_Contable': rc['Descripci√≥n'] if rc is not None else '',
                            'Diferencia': abs(suma_b - suma_c) if i == 0 else None
                        })
                    
                    banco.loc[regs_b.index, 'Conciliado'] = True
                    contable.loc[regs_c.index, 'Conciliado'] = True
                    contador += 1
                    id_conc += 1
                    break
    
    print(f"‚úì Agrupaciones N‚ÜîM: {contador}")
    return contador

# ============================================================================
# üéØ ESTRATEGIA 5: IMPUESTOS DGII
# ============================================================================

def conciliacion_impuestos(banco, contable, conciliaciones):
    print("\n" + "="*70)
    print("üéØ ESTRATEGIA 5: IMPUESTOS DGII (0.15%)")
    print("="*70)
    
    contador = 0
    id_conc = len(conciliaciones) + 1 if conciliaciones else 1
    
    banco_imp = banco[(~banco['Conciliado']) & (banco['Es_Impuesto'])].copy()
    
    if len(banco_imp) == 0:
        print("‚äò No hay impuestos pendientes en banco")
        return 0
    
    contable_imp = contable[
        (~contable['Conciliado']) &
        (contable['Es_Impuesto'])
    ].copy()
    
    if len(contable_imp) == 0:
        print("‚äò No hay impuestos pendientes en contable")
        return 0
    
    banco_imp['Mes'] = banco_imp['Fecha'].dt.to_period('M')
    
    for mes, grupo_banco in banco_imp.groupby('Mes'):
        suma_banco = abs(grupo_banco['Valor'].sum())
        
        for idx_c, reg_c in contable_imp.iterrows():
            if contable.loc[idx_c, 'Conciliado']:
                continue
            
            if abs(abs(reg_c['Valor']) - suma_banco) < TOLERANCIA_VALOR_AGRUPACION:
                for i, (idx_b, reg_b) in enumerate(grupo_banco.iterrows()):
                    conciliaciones.append({
                        'ID_Conciliacion': id_conc,
                        'Tipo': f'5. Impuestos DGII ({len(grupo_banco)}‚Üí1)',
                        'Similitud': 1.0,
                        'Fecha_Banco': reg_b['Fecha'],
                        'Concepto_Banco': reg_b['Concepto'],
                        'Valor_Banco': reg_b['Valor'],
                        'Descripcion_Banco': reg_b['Descripci√≥n'],
                        'Fecha_Contable': reg_c['Fecha'] if i == 0 else None,
                        'Concepto_Contable': reg_c['Concepto'] if i == 0 else '',
                        'Valor_Contable': reg_c['Valor'] if i == 0 else None,
                        'Descripcion_Contable': reg_c['Descripci√≥n'] if i == 0 else '',
                        'Diferencia': abs(suma_banco - abs(reg_c['Valor'])) if i == 0 else None
                    })
                
                banco.loc[grupo_banco.index, 'Conciliado'] = True
                contable.loc[idx_c, 'Conciliado'] = True
                contador += 1
                id_conc += 1
                break
    
    print(f"‚úì Impuestos conciliados: {contador}")
    return contador

# ============================================================================
# üéØ ESTRATEGIA 6: SEGUNDA PASADA FLEXIBLE - OPTIMIZADA
# ============================================================================

def segunda_pasada_inteligente(banco, contable, conciliaciones):
    if not EJECUTAR_SEGUNDA_PASADA:
        return 0
    
    print("\n" + "="*70)
    print("üéØ ESTRATEGIA 6: SEGUNDA PASADA (B√∫squeda Flexible)")
    print("="*70)
    
    contador = 0
    id_conc = len(conciliaciones) + 1 if conciliaciones else 1
    
    for idx_b, reg_b in banco[~banco['Conciliado']].iterrows():
        if banco.loc[idx_b, 'Conciliado']:
            continue
        
        f_min = reg_b['Fecha'] - timedelta(days=VENTANA_DIAS_FLEXIBLE)
        f_max = reg_b['Fecha'] + timedelta(days=VENTANA_DIAS_FLEXIBLE)
        
        cands = contable[
            (~contable['Conciliado']) &
            (contable['Fecha'] >= f_min) &
            (contable['Fecha'] <= f_max)
        ].copy()
        
        if len(cands) == 0:
            continue
        
        valor_objetivo = reg_b['Valor']
        mejor_combo, mejor_diff = None, float('inf')
        
        # ‚ö° OPTIMIZACI√ìN: Limitar partidas y combinaciones
        max_items = min(len(cands), MAX_PARTIDAS_AGRUPACION)
        cands_limitados = cands.head(max_items)
        
        total_combinaciones = 0
        for r in range(2, min(len(cands_limitados) + 1, 8)):
            if mejor_diff < TOLERANCIA_VALOR_AGRUPACION:
                break
            if total_combinaciones >= MAX_COMBINACIONES_POR_BUSQUEDA:
                break
            
            for combo in combinations(cands_limitados.index, r):
                total_combinaciones += 1
                if total_combinaciones > MAX_COMBINACIONES_POR_BUSQUEDA:
                    break
                
                suma = contable.loc[list(combo), 'Valor'].sum()
                diff = abs(suma - valor_objetivo)
                
                if diff < mejor_diff:
                    mejor_diff = diff
                    mejor_combo = combo
                
                if diff < TOLERANCIA_VALOR_AGRUPACION:
                    break
        
        if mejor_combo and mejor_diff < TOLERANCIA_VALOR_AGRUPACION:
            grupo = contable.loc[list(mejor_combo)]
            suma = grupo['Valor'].sum()
            
            sim_media = grupo['Texto_Busqueda'].apply(
                lambda x: calcular_similitud(reg_b['Texto_Busqueda'], x)
            ).mean()
            
            for i, (_, reg_c) in enumerate(grupo.sort_values(['Fecha', 'Valor']).iterrows()):
                conciliaciones.append({
                    'ID_Conciliacion': id_conc,
                    'Tipo': f'6. Segunda Pasada ({len(grupo)} reg)',
                    'Similitud': round(sim_media, 3),
                    'Fecha_Banco': reg_b['Fecha'] if i == 0 else None,
                    'Concepto_Banco': reg_b['Concepto'] if i == 0 else '',
                    'Valor_Banco': reg_b['Valor'] if i == 0 else None,
                    'Descripcion_Banco': reg_b['Descripci√≥n'] if i == 0 else '',
                    'Fecha_Contable': reg_c['Fecha'],
                    'Concepto_Contable': reg_c['Concepto'],
                    'Valor_Contable': reg_c['Valor'],
                    'Descripcion_Contable': reg_c['Descripci√≥n'],
                    'Diferencia': mejor_diff if i == 0 else None
                })
            
            banco.loc[idx_b, 'Conciliado'] = True
            contable.loc[list(mejor_combo), 'Conciliado'] = True
            contador += 1
            id_conc += 1
    
    print(f"‚úì Segunda pasada: {contador}")
    return contador

# ============================================================================
# üéØ ESTRATEGIA 7: B√öSQUEDA EXHAUSTIVA FINAL [v6.1] - OPTIMIZADA
# ============================================================================

def busqueda_exhaustiva_final(banco, contable, conciliaciones):
    if not EJECUTAR_BUSQUEDA_EXHAUSTIVA:
        return 0
    
    print("\n" + "="*70)
    print("üéØ ESTRATEGIA 7: B√öSQUEDA EXHAUSTIVA (Sin restricci√≥n fechas)")
    print("="*70)
    
    contador = 0
    id_conc = len(conciliaciones) + 1 if conciliaciones else 1
    
    banco_pend = banco[~banco['Conciliado']].copy()
    contable_pend = contable[~contable['Conciliado']].copy()
    
    n_banco_pend = len(banco_pend)
    n_contable_pend = len(contable_pend)
    
    print(f"  üìä Pendientes: Banco={n_banco_pend}, Contable={n_contable_pend}")
    
    if n_banco_pend > UMBRAL_PARTIDAS_EXHAUSTIVA or n_contable_pend > UMBRAL_PARTIDAS_EXHAUSTIVA:
        print(f"  ‚äò Demasiadas partidas (umbral={UMBRAL_PARTIDAS_EXHAUSTIVA})")
        return 0
    
    if n_banco_pend == 0 or n_contable_pend == 0:
        print("  ‚äò No hay partidas en ambos lados")
        return 0
    
    # CASO 1: Buscar N contable ‚Üí 1 banco
    if n_contable_pend <= 10:
        print(f"\n  üîç Buscando: Banco ‚Üí Contable...")
        for idx_c, reg_c in contable_pend.iterrows():
            if contable.loc[idx_c, 'Conciliado']:
                continue
            
            valor_objetivo = reg_c['Valor']
            banco_disponible = banco[~banco['Conciliado']]
            
            if len(banco_disponible) == 0:
                continue
            
            mejor_combo, mejor_diff = None, float('inf')
            indices = banco_disponible.index.tolist()[:MAX_PARTIDAS_AGRUPACION]
            
            total_comb = 0
            for r in range(1, len(indices) + 1):
                if mejor_diff < TOLERANCIA_VALOR_AGRUPACION or total_comb > MAX_COMBINACIONES_EXHAUSTIVA:
                    break
                
                for combo in combinations(indices, r):
                    total_comb += 1
                    if total_comb > MAX_COMBINACIONES_EXHAUSTIVA:
                        break
                    
                    suma = banco.loc[list(combo), 'Valor'].sum()
                    diff = abs(suma - valor_objetivo)
                    
                    if diff < mejor_diff:
                        mejor_diff = diff
                        mejor_combo = combo
                    
                    if diff < TOLERANCIA_VALOR_AGRUPACION:
                        break
            
            if mejor_combo and mejor_diff < TOLERANCIA_VALOR_AGRUPACION:
                grupo = banco.loc[list(mejor_combo)]
                print(f"    ‚úì {len(grupo)} banco ‚Üí 1 contable (diff=${mejor_diff:.2f})")
                
                for i, (idx_b, reg_b) in enumerate(grupo.sort_values(['Fecha', 'Valor']).iterrows()):
                    conciliaciones.append({
                        'ID_Conciliacion': id_conc,
                        'Tipo': f'7. Exhaustiva ({len(grupo)}‚Üí1)',
                        'Similitud': 0.0,
                        'Fecha_Banco': reg_b['Fecha'],
                        'Concepto_Banco': reg_b['Concepto'],
                        'Valor_Banco': reg_b['Valor'],
                        'Descripcion_Banco': reg_b['Descripci√≥n'],
                        'Fecha_Contable': reg_c['Fecha'] if i == 0 else None,
                        'Concepto_Contable': reg_c['Concepto'] if i == 0 else '',
                        'Valor_Contable': reg_c['Valor'] if i == 0 else None,
                        'Descripcion_Contable': reg_c['Descripci√≥n'] if i == 0 else '',
                        'Diferencia': mejor_diff if i == 0 else None
                    })
                
                banco.loc[list(mejor_combo), 'Conciliado'] = True
                contable.loc[idx_c, 'Conciliado'] = True
                contador += 1
                id_conc += 1
    
    # CASO 2: Buscar N banco ‚Üí 1 contable
    banco_pend = banco[~banco['Conciliado']].copy()
    if len(banco_pend) <= 10 and len(banco_pend) > 0:
        print(f"\n  üîç Buscando: Contable ‚Üí Banco...")
        for idx_b, reg_b in banco_pend.iterrows():
            if banco.loc[idx_b, 'Conciliado']:
                continue
            
            valor_objetivo = reg_b['Valor']
            contable_disponible = contable[~contable['Conciliado']]
            
            if len(contable_disponible) == 0:
                continue
            
            mejor_combo, mejor_diff = None, float('inf')
            indices = contable_disponible.index.tolist()[:MAX_PARTIDAS_AGRUPACION]
            
            total_comb = 0
            for r in range(1, len(indices) + 1):
                if mejor_diff < TOLERANCIA_VALOR_AGRUPACION or total_comb > MAX_COMBINACIONES_EXHAUSTIVA:
                    break
                
                for combo in combinations(indices, r):
                    total_comb += 1
                    if total_comb > MAX_COMBINACIONES_EXHAUSTIVA:
                        break
                    
                    suma = contable.loc[list(combo), 'Valor'].sum()
                    diff = abs(suma - valor_objetivo)
                    
                    if diff < mejor_diff:
                        mejor_diff = diff
                        mejor_combo = combo
                    
                    if diff < TOLERANCIA_VALOR_AGRUPACION:
                        break
            
            if mejor_combo and mejor_diff < TOLERANCIA_VALOR_AGRUPACION:
                grupo = contable.loc[list(mejor_combo)]
                print(f"    ‚úì 1 banco ‚Üí {len(grupo)} contable (diff=${mejor_diff:.2f})")
                
                for i, (_, reg_c) in enumerate(grupo.sort_values(['Fecha', 'Valor']).iterrows()):
                    conciliaciones.append({
                        'ID_Conciliacion': id_conc,
                        'Tipo': f'7. Exhaustiva (1‚Üí{len(grupo)})',
                        'Similitud': 0.0,
                        'Fecha_Banco': reg_b['Fecha'] if i == 0 else None,
                        'Concepto_Banco': reg_b['Concepto'] if i == 0 else '',
                        'Valor_Banco': reg_b['Valor'] if i == 0 else None,
                        'Descripcion_Banco': reg_b['Descripci√≥n'] if i == 0 else '',
                        'Fecha_Contable': reg_c['Fecha'],
                        'Concepto_Contable': reg_c['Concepto'],
                        'Valor_Contable': reg_c['Valor'],
                        'Descripcion_Contable': reg_c['Descripci√≥n'],
                        'Diferencia': mejor_diff if i == 0 else None
                    })
                
                banco.loc[idx_b, 'Conciliado'] = True
                contable.loc[list(mejor_combo), 'Conciliado'] = True
                contador += 1
                id_conc += 1
    
    print(f"\n‚úì B√∫squeda exhaustiva: {contador}")
    return contador

# ============================================================================
# üîß FUNCIONES AUXILIARES PARA CAJA CHICA (A√ëADIR DESPU√âS DE LAS FUNCIONES DE NORMALIZACI√ìN)
# ============================================================================

def es_patron_caja_chica(texto):
    """Detecta si el texto corresponde a reposici√≥n de caja chica"""
    texto_norm = normalizar_texto(texto)
    return any(patron in texto_norm for patron in PATRONES_CAJA_CHICA)

def extraer_nombre_persona(texto):
    """
    Extrae el nombre completo de una persona del texto
    Busca patrones como: "TRANSFERENCIA A JUAN PEREZ" o "MIGUEL ANGEL SANTIAGO"
    """
    texto_norm = normalizar_texto(texto)
    if not texto_norm:
        return None
    
    # Patr√≥n 1: "TRANSFERENCIA A [NOMBRE]"
    match = re.search(r'(?:TRANSFERENCIA|TRANSF|PAGO|DEBITO)\s+(?:A|PARA|DE)\s+([A-Z\s]{6,})', texto_norm)
    if match:
        nombre = match.group(1).strip()
        # Eliminar tokens que no son parte del nombre (ej. REPOSICION, CAJA, CHICA)
        ruido_tokens = set()
        for pat in PATRONES_CAJA_CHICA:
            for w in pat.split():
                if len(w) >= 3:
                    ruido_tokens.add(w)

        palabras = [p for p in nombre.split() if len(p) >= 3 and p not in PALABRAS_COMUNES and p not in ruido_tokens]
        if len(palabras) >= 2:
            return ' '.join(palabras)
    
    # Patr√≥n 2: Buscar nombres completos (2-4 palabras capitalizadas consecutivas)
    palabras = texto_norm.split()
    nombres_encontrados = []
    
    for i in range(len(palabras) - 1):
        # Buscar secuencias de 2-4 palabras que parezcan nombres
        for longitud in [4, 3, 2]:  # Empezar por las m√°s largas
            if i + longitud <= len(palabras):
                candidato = palabras[i:i+longitud]
                # Filtrar palabras comunes
                candidato_limpio = [p for p in candidato if len(p) >= 3 and p not in PALABRAS_COMUNES]
                
                if len(candidato_limpio) >= 2:
                    # Tambi√©n eliminar tokens de ruido de caja chica
                    ruido_tokens = set()
                    for pat in PATRONES_CAJA_CHICA:
                        for w in pat.split():
                            if len(w) >= 3:
                                ruido_tokens.add(w)

                    candidato_final = [p for p in candidato_limpio if p not in ruido_tokens]
                    if len(candidato_final) >= 2:
                        nombre_completo = ' '.join(candidato_final)
                        # Validar que parezca un nombre (longitud razonable)
                        if 10 <= len(nombre_completo) <= 50:
                            nombres_encontrados.append(nombre_completo)
    
    # Retornar el nombre m√°s largo encontrado (generalmente es el m√°s completo)
    if nombres_encontrados:
        return max(nombres_encontrados, key=len)
    
    return None

def similitud_nombres(nombre1, nombre2):
    """
    Calcula similitud entre dos nombres de persona
    Retorna score de 0.0 a 1.0
    """
    if not nombre1 or not nombre2:
        return 0.0
    
    n1 = normalizar_texto(nombre1)
    n2 = normalizar_texto(nombre2)
    
    palabras1 = set(n1.split())
    palabras2 = set(n2.split())
    
    # Si las palabras principales coinciden, es muy probable que sea la misma persona
    coincidencias = len(palabras1 & palabras2)
    max_palabras = max(len(palabras1), len(palabras2))
    
    if max_palabras == 0:
        return 0.0
    
    return coincidencias / max_palabras

# ============================================================================
# üéØ ESTRATEGIA 8: REPOSICI√ìN DE CAJA CHICA [NUEVA v6.6]
# ============================================================================

def conciliacion_caja_chica(banco, contable, conciliaciones):
    """
    Estrategia especializada para reposiciones de caja chica
    
    L√ìGICA:
    1. Detecta registros de banco con patr√≥n "REPOSICI√ìN DE CAJA CHICA"
    2. Extrae el nombre de la persona del banco
    3. Busca TODAS las partidas contables con ese mismo nombre
    4. Agrupa por suma de montos (los conceptos pueden ser diversos)
    5. Concilia si la suma coincide con la reposici√≥n del banco
    """
    print("\n" + "="*70)
    print("üéØ ESTRATEGIA 8: REPOSICI√ìN DE CAJA CHICA")
    print("="*70)
    
    contador = 0
    id_conc = len(conciliaciones) + 1 if conciliaciones else 1
    
    # Buscar registros de caja chica en banco
    banco_caja_chica = banco[
        (~banco['Conciliado']) & 
        (banco['Texto_Busqueda'].apply(es_patron_caja_chica))
    ].copy()
    
    if len(banco_caja_chica) == 0:
        print("‚äò No hay reposiciones de caja chica pendientes en banco")
        return 0
    
    print(f"  üìã Encontradas {len(banco_caja_chica)} reposiciones de caja chica en banco")
    
    for idx_b, reg_b in banco_caja_chica.iterrows():
        if banco.loc[idx_b, 'Conciliado']:
            continue
        
        # PASO 1: Extraer nombre de persona del banco
        nombre_banco = extraer_nombre_persona(reg_b['Texto_Busqueda'])
        
        if not nombre_banco:
            print(f"  ‚ö†Ô∏è  No se pudo extraer nombre de persona de: '{reg_b['Concepto']}'")
            continue
        
        print(f"\n  üîç Buscando gastos de: {nombre_banco}")
        print(f"      Banco: {reg_b['Concepto']} - ${reg_b['Valor']:,.2f}")
        
        # PASO 2: Buscar TODAS las partidas contables con ese nombre
        # Ventana de tiempo amplia (generalmente las reposiciones cubren varios d√≠as)
        f_min = reg_b['Fecha'] - timedelta(days=VENTANA_DIAS_AGRUPACION)
        f_max = reg_b['Fecha'] + timedelta(days=VENTANA_DIAS_AGRUPACION)
        
        # Excluir partidas que claramente son comisiones o impuestos;
        # adem√°s filtrar por presencia del nombre extra√≠do en el Texto_Busqueda
        candidatos = contable[
            (~contable['Conciliado']) &
            (~contable.get('Es_Comision', False)) &
            (~contable.get('Es_Impuesto', False)) &
            (contable['Fecha'] >= f_min) &
            (contable['Fecha'] <= f_max) &
            (contable['Texto_Busqueda'].str.contains(nombre_banco))
        ].copy()
        
        if len(candidatos) == 0:
            continue
        
        # Buscar partidas que contengan el nombre de la persona
        partidas_persona = []
        # Diagn√≥stico: comparar con todas las partidas del periodo que contienen el nombre (sin excluir flags)
        todas_partidas_nombre = []
        for idx_c_all, reg_c_all in contable[(contable['Fecha'] >= f_min) & (contable['Fecha'] <= f_max)].iterrows():
            # Incluir cualquier partida del periodo que contenga el nombre en Texto_Busqueda (no filtrar por flags aqu√≠)
            if nombre_banco and nombre_banco in reg_c_all['Texto_Busqueda']:
                todas_partidas_nombre.append({'index': idx_c_all, 'registro': reg_c_all, 'Es_Comision': reg_c_all.get('Es_Comision', False), 'Es_Impuesto': reg_c_all.get('Es_Impuesto', False)})

        if len(todas_partidas_nombre) != 0:
            suma_todas = sum(r['registro']['Valor'] for r in todas_partidas_nombre)
            excluidas = [r for r in todas_partidas_nombre if r['Es_Comision'] or r['Es_Impuesto']]
            suma_excluidas = sum(r['registro']['Valor'] for r in excluidas) if excluidas else 0
            print(f"      (DEBUG) Partidas totales con nombre en ventana: {len(todas_partidas_nombre)}; suma={suma_todas:,.2f}; excluidas por flag: {len(excluidas)} suma_excluidas={suma_excluidas:,.2f}")
            # Mostrar detalles de las partidas totales (√≠ndice, valor, concepto, flags)
            for r in todas_partidas_nombre:
                reg = r['registro']
                print(f"        (DBG-TOT) idx={r['index']} valor={reg['Valor']:,.2f} flag_com={r['Es_Comision']} flag_imp={r['Es_Impuesto']} concept='{reg['Concepto'][:50]}' desc='{reg['Descripci√≥n'][:40]}'")
            if excluidas:
                print("        (DBG-TOT) Excluidas:")
                for e in excluidas:
                    reg = e['registro']
                    print(f"          - idx={e['index']} valor={reg['Valor']:,.2f} concept='{reg['Concepto'][:50]}' desc='{reg['Descripci√≥n'][:40]}'")
        for idx_c, reg_c in candidatos.iterrows():
            # Asegurar que no se consideren partidas marcadas como comisi√≥n/impuesto
            if reg_c.get('Es_Comision', False) or reg_c.get('Es_Impuesto', False):
                continue

            # Si el nombre extra√≠do del banco aparece literalmente en Texto_Busqueda, asumir match directo
            if nombre_banco and nombre_banco in reg_c['Texto_Busqueda']:
                partidas_persona.append({
                    'index': idx_c,
                    'registro': reg_c,
                    'similitud_nombre': 1.0
                })
                continue

            nombre_contable = extraer_nombre_persona(reg_c['Texto_Busqueda'])

            if nombre_contable:
                sim = similitud_nombres(nombre_banco, nombre_contable)

                # Si hay alta similitud en nombres, es la misma persona
                if sim >= 0.6:  # Al menos 60% de coincidencia
                    partidas_persona.append({
                        'index': idx_c,
                        'registro': reg_c,
                        'similitud_nombre': sim
                    })

        # Diagn√≥stico: imprimir suma de partidas_persona seleccionadas
        if partidas_persona:
            suma_seleccionadas = sum(p['registro']['Valor'] for p in partidas_persona)
            print(f"      (DEBUG) Partidas seleccionadas (sin flags): {len(partidas_persona)}; suma={suma_seleccionadas:,.2f}")
        
        if len(partidas_persona) == 0:
            print(f"      ‚ùå No se encontraron gastos de '{nombre_banco}' en contable")
            continue
        
        # PASO 3: Agrupar todas las partidas de esa persona
        indices_grupo = [p['index'] for p in partidas_persona]
        grupo = contable.loc[indices_grupo]
        
        suma_contable = grupo['Valor'].sum()
        diferencia = abs(reg_b['Valor'] - suma_contable)
        
        print(f"      ‚úì Encontradas {len(grupo)} partidas:")
        for _, reg in grupo.head(5).iterrows():  # Mostrar solo las primeras 5
            print(f"        ‚Ä¢ {reg['Concepto'][:40]:40} ${reg['Valor']:>10,.2f}")
        if len(grupo) > 5:
            print(f"        ... y {len(grupo) - 5} partidas m√°s")
        print(f"      üìä Suma contable: ${suma_contable:,.2f}")
        print(f"      üìä Diferencia:    ${diferencia:,.2f}")
        
        # PASO 4: Conciliar si la suma coincide
        if diferencia < TOLERANCIA_VALOR_AGRUPACION:
            print(f"      ‚úÖ CONCILIADO - Diferencia aceptable: ${diferencia:.2f}")
            
            # Registrar conciliaci√≥n
            for i, (idx_c, reg_c) in enumerate(grupo.sort_values(['Fecha', 'Valor']).iterrows()):
                conciliaciones.append({
                    'ID_Conciliacion': id_conc,
                    'Tipo': f'8. Caja Chica ({len(grupo)}‚Üí1)',
                    'Similitud': round(partidas_persona[0]['similitud_nombre'], 3),
                    'Fecha_Banco': reg_b['Fecha'] if i == 0 else None,
                    'Concepto_Banco': reg_b['Concepto'] if i == 0 else '',
                    'Valor_Banco': reg_b['Valor'] if i == 0 else None,
                    'Descripcion_Banco': reg_b['Descripci√≥n'] if i == 0 else '',
                    'Fecha_Contable': reg_c['Fecha'],
                    'Concepto_Contable': reg_c['Concepto'],
                    'Valor_Contable': reg_c['Valor'],
                    'Descripcion_Contable': reg_c['Descripci√≥n'],
                    'Diferencia': diferencia if i == 0 else None
                })
            
            # Marcar como conciliadas
            banco.loc[idx_b, 'Conciliado'] = True
            contable.loc[indices_grupo, 'Conciliado'] = True
            
            contador += 1
            id_conc += 1
        else:
            # Intentar conciliar sumando TODAS las reposiciones de banco para la misma persona
            # (caso: varias transferencias que en conjunto cubren los gastos)
            similares_banco = []
            for idx_bb, reg_bb in banco_caja_chica.iterrows():
                if banco.loc[idx_bb, 'Conciliado']:
                    continue
                nombre_bb = extraer_nombre_persona(reg_bb['Texto_Busqueda'])
                if nombre_bb and similitud_nombres(nombre_banco, nombre_bb) >= 0.6:
                    similares_banco.append({'index': idx_bb, 'registro': reg_bb})

            if similares_banco:
                suma_bancos = sum(b['registro']['Valor'] for b in similares_banco)
                diff_grupal = abs(suma_bancos - suma_contable)
                print(f"      (DEBUG) Suma reposiciones banco para {nombre_banco}: {suma_bancos:,.2f}; diff_grupal={diff_grupal:,.2f}")
                if diff_grupal < TOLERANCIA_VALOR_AGRUPACION:
                    print(f"      ‚úÖ CONCILIADO COMO GRUPO - {len(similares_banco)} reposiciones suman {suma_bancos:,.2f}")
                    # Registrar conciliaciones: map each banco y cada contable registro
                    for i_b, binfo in enumerate(sorted(similares_banco, key=lambda x: x['registro']['Fecha'])):
                        for i_c, (idx_c, reg_c) in enumerate(grupo.sort_values(['Fecha', 'Valor']).iterrows()):
                            conciliaciones.append({
                                'ID_Conciliacion': id_conc,
                                'Tipo': f'8. Caja Chica (Grupo {len(similares_banco)}‚Üí{len(grupo)})',
                                'Similitud': round(partidas_persona[0]['similitud_nombre'], 3) if partidas_persona else 0.0,
                                'Fecha_Banco': binfo['registro']['Fecha'] if i_c == 0 and i_b == 0 else None,
                                'Concepto_Banco': binfo['registro']['Concepto'] if i_c == 0 and i_b == 0 else '',
                                'Valor_Banco': binfo['registro']['Valor'] if i_c == 0 and i_b == 0 else None,
                                'Descripcion_Banco': binfo['registro']['Descripci√≥n'] if i_c == 0 and i_b == 0 else '',
                                'Fecha_Contable': reg_c['Fecha'] if i_b == 0 else None,
                                'Concepto_Contable': reg_c['Concepto'] if i_b == 0 else '',
                                'Valor_Contable': reg_c['Valor'] if i_b == 0 else None,
                                'Descripcion_Contable': reg_c['Descripci√≥n'] if i_b == 0 else '',
                                'Diferencia': diff_grupal if i_b == 0 and i_c == 0 else None
                            })

                    # Marcar como conciliadas
                    banco.loc[[b['index'] for b in similares_banco], 'Conciliado'] = True
                    contable.loc[indices_grupo, 'Conciliado'] = True
                    contador += 1
                    id_conc += 1
                else:
                    print(f"      ‚ö†Ô∏è  Diferencia muy alta: ${diferencia:.2f} > ${TOLERANCIA_VALOR_AGRUPACION:.2f}")
            else:
                print(f"      ‚ö†Ô∏è  Diferencia muy alta: ${diferencia:.2f} > ${TOLERANCIA_VALOR_AGRUPACION:.2f}")
    
    print(f"\n‚úì Reposiciones de caja chica conciliadas: {contador}")
    return contador

# ============================================================================
# üîç DETECCI√ìN DE CASOS ESPECIALES
# ============================================================================

def detectar_casos_especiales(banco, contable):
    print("\n" + "="*70)
    print("üîç DETECTANDO CASOS ESPECIALES")
    print("="*70)
    
    casos_especiales = []
    
    if not DETECTAR_CASOS_ESPECIALES:
        print("‚äò Detecci√≥n desactivada")
        return casos_especiales
    
    banco_pend = banco[~banco['Conciliado']].copy()
    contable_pend = contable[~contable['Conciliado']].copy()
    
    for idx_b, reg_b in banco_pend.iterrows():
        for prov_id in contable_pend['Proveedor_ID'].dropna().unique():
            grupo = contable_pend[contable_pend['Proveedor_ID'] == prov_id]
            if len(grupo) < 1:
                continue
            
            suma = grupo['Valor'].sum()
            
            if reg_b['Valor'] != 0:
                porcentaje = abs(suma) / abs(reg_b['Valor']) * 100
            else:
                continue
            
            if 80 <= porcentaje < 100:
                diferencia = abs(reg_b['Valor']) - abs(suma)
                
                sim = calcular_similitud(
                    reg_b['Texto_Busqueda'],
                    ' '.join(grupo['Texto_Busqueda'].tolist())
                )
                
                if sim > 0.1:
                    casos_especiales.append({
                        'Tipo': 'PARCIAL',
                        'Descripcion': f'{len(grupo)} partidas suman {porcentaje:.1f}% del banco',
                        'Banco_Fecha': reg_b['Fecha'],
                        'Banco_Concepto': reg_b['Concepto'],
                        'Banco_Valor': reg_b['Valor'],
                        'Banco_Descripcion': reg_b['Descripci√≥n'],
                        'Contable_Registros': len(grupo),
                        'Contable_Suma': suma,
                        'Diferencia_Pendiente': diferencia,
                        'Porcentaje_Conciliado': round(porcentaje, 1),
                        'Similitud': round(sim, 3)
                    })
                    break
        
        mes_banco = reg_b['Fecha'].month if pd.notna(reg_b['Fecha']) else None
        
        if mes_banco:
            otros_meses = contable_pend[
                contable_pend['Fecha'].dt.month != mes_banco
            ]
            
            for _, reg_c in otros_meses.iterrows():
                if abs(reg_c['Valor'] - reg_b['Valor']) < TOLERANCIA_VALOR_EXACTA:
                    sim = calcular_similitud(reg_b['Texto_Busqueda'], reg_c['Texto_Busqueda'])
                    
                    if sim > 0.2:
                        casos_especiales.append({
                            'Tipo': 'PERIODO_DIFERENTE',
                            'Descripcion': f'Monto exacto pero mes diferente',
                            'Banco_Fecha': reg_b['Fecha'],
                            'Banco_Concepto': reg_b['Concepto'],
                            'Banco_Valor': reg_b['Valor'],
                            'Banco_Descripcion': reg_b['Descripci√≥n'],
                            'Contable_Registros': 1,
                            'Contable_Suma': reg_c['Valor'],
                            'Diferencia_Pendiente': 0,
                            'Porcentaje_Conciliado': 100.0,
                            'Similitud': round(sim, 3)
                        })
    
    casos_unicos = []
    valores_vistos = set()
    for caso in casos_especiales:
        key = (caso['Banco_Valor'], caso['Contable_Suma'])
        if key not in valores_vistos:
            valores_vistos.add(key)
            casos_unicos.append(caso)
    
    print(f"‚ö†Ô∏è Casos especiales detectados: {len(casos_unicos)}")
    return casos_unicos

# ============================================================================
# üé® FORMATO EXCEL PROFESIONAL
# ============================================================================

def aplicar_formato_profesional(ruta):
    """Aplica formato moderno y profesional al Excel"""
    wb = openpyxl.load_workbook(ruta)
    
    C_PRI = "1A237E"
    C_SEC = "3F51B5"
    C_EXI = "2E7D32"
    C_ERR = "C62828"
    C_ADV = "EF6C00"
    C_GRI = "FAFAFA"
    C_GRI2 = "EEEEEE"
    
    f_tit = Font(name='Segoe UI', size=22, bold=True, color="FFFFFF")
    f_sub = Font(name='Segoe UI', size=12, italic=True, color="757575")
    f_enc = Font(name='Segoe UI', size=11, bold=True, color="FFFFFF")
    f_sec = Font(name='Segoe UI', size=13, bold=True, color=C_SEC)
    f_nor = Font(name='Segoe UI', size=10)
    f_num = Font(name='Segoe UI', size=10, bold=True)
    
    r_pri = PatternFill(start_color=C_PRI, end_color=C_PRI, fill_type="solid")
    r_sec = PatternFill(start_color=C_SEC, end_color=C_SEC, fill_type="solid")
    r_exi = PatternFill(start_color=C_EXI, end_color=C_EXI, fill_type="solid")
    r_err = PatternFill(start_color=C_ERR, end_color=C_ERR, fill_type="solid")
    r_adv = PatternFill(start_color=C_ADV, end_color=C_ADV, fill_type="solid")
    r_gri = PatternFill(start_color=C_GRI, end_color=C_GRI, fill_type="solid")
    r_gri2 = PatternFill(start_color=C_GRI2, end_color=C_GRI2, fill_type="solid")
    
    tipos_r = {
        '1': PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid"),
        '2': PatternFill(start_color="BBDEFB", end_color="BBDEFB", fill_type="solid"),
        '3': PatternFill(start_color="FFE082", end_color="FFE082", fill_type="solid"),
        '4': PatternFill(start_color="FFCCBC", end_color="FFCCBC", fill_type="solid"),
        '5': PatternFill(start_color="E1BEE7", end_color="E1BEE7", fill_type="solid"),
        '6': PatternFill(start_color="B2DFDB", end_color="B2DFDB", fill_type="solid"),
        '7': PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid"),
        '8': PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid"),  # Amarillo claro
    }
    
    borde = Border(
        left=Side(style='thin', color="BDBDBD"),
        right=Side(style='thin', color="BDBDBD"),
        top=Side(style='thin', color="BDBDBD"),
        bottom=Side(style='thin', color="BDBDBD")
    )
    
    if 'RESUMEN' in wb.sheetnames:
        ws = wb['RESUMEN']
        ws.insert_rows(1, 3)
        
        ws.merge_cells('A1:B1')
        ws['A1'].value = "üìä CONCILIACI√ìN BANCARIA v6.5 DEFINITIVO"
        ws['A1'].font = f_tit
        ws['A1'].fill = r_pri
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 50
        
        ws.merge_cells('A2:B2')
        ws['A2'].value = "GREEN PARK - BANRESERBAS"
        ws['A2'].font = Font(name='Segoe UI', size=14, bold=True, color="424242")
        ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[2].height = 28
        
        ws.merge_cells('A3:B3')
        ws['A3'].value = f"üìÖ Generado: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
        ws['A3'].font = f_sub
        ws['A3'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[3].height = 24
        
        for r in range(4, ws.max_row + 1):
            for c in [1, 2]:
                cell = ws.cell(r, c)
                cell.border = borde
                cell.font = f_nor
                
                if c == 1 and cell.value and isinstance(cell.value, str):
                    if '‚ïê‚ïê‚ïê' in str(cell.value):
                        cell.font = f_sec
                        cell.fill = r_gri2
                    elif '‚îÄ‚îÄ‚îÄ' in str(cell.value):
                        cell.font = Font(name='Segoe UI', size=11, bold=True, color=C_SEC)
                        cell.fill = r_gri
                
                if c == 2 and isinstance(cell.value, (int, float)):
                    cell.font = f_num
                    cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                    if cell.value < 0:
                        cell.font = Font(name='Segoe UI', size=10, bold=True, color=C_ERR)
                    elif cell.value > 0:
                        cell.font = Font(name='Segoe UI', size=10, bold=True, color=C_EXI)
        
        ws.column_dimensions['A'].width = 62
        ws.column_dimensions['B'].width = 35
    
    if 'CONCILIADOS' in wb.sheetnames:
        ws = wb['CONCILIADOS']
        if ws.max_row > 1:
            ws.insert_rows(1)
            max_c = ws.max_column
            
            ws.merge_cells(f'A1:{get_column_letter(max_c)}1')
            ws['A1'].value = "‚úÖ PARTIDAS CONCILIADAS"
            ws['A1'].font = f_tit
            ws['A1'].fill = r_exi
            ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
            ws.row_dimensions[1].height = 45
            
            for c in range(1, max_c + 1):
                cell = ws.cell(2, c)
                cell.fill = r_sec
                cell.font = f_enc
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = borde
            ws.row_dimensions[2].height = 38
            
            for r in range(3, ws.max_row + 1):
                tipo_val = str(ws.cell(r, 2).value) if ws.cell(r, 2).value else ""
                relleno = None
                
                for t_num, t_fill in tipos_r.items():
                    if tipo_val.startswith(f'{t_num}.'):
                        relleno = t_fill
                        break
                
                if not relleno and r % 2 == 0:
                    relleno = r_gri
                
                for c in range(1, max_c + 1):
                    cell = ws.cell(r, c)
                    if relleno:
                        cell.fill = relleno
                    cell.border = borde
                    cell.font = f_nor
                    cell.alignment = Alignment(vertical='center', wrap_text=True)
                    
                    hdr = ws.cell(2, c)
                    if hdr.value:
                        if 'Fecha' in str(hdr.value) and isinstance(cell.value, datetime):
                            cell.number_format = 'DD/MM/YYYY'
                        elif ('Valor' in str(hdr.value) or 'Diferencia' in str(hdr.value)):
                            if isinstance(cell.value, (int, float)):
                                cell.number_format = '#,##0.00'
                                color = C_ERR if cell.value < 0 else C_EXI
                                cell.font = Font(name='Segoe UI', size=10, bold=True, color=color)
            
            for c in range(1, max_c + 1):
                col_letter = get_column_letter(c)
                max_len = 0
                for r in range(1, min(ws.max_row + 1, 100)):
                    val = ws.cell(r, c).value
                    if val:
                        max_len = max(max_len, len(str(val)))
                ws.column_dimensions[col_letter].width = min(max_len + 4, 55)
    
    if 'NO_CONCILIADOS' in wb.sheetnames:
        ws = wb['NO_CONCILIADOS']
        if ws.max_row > 1:
            ws.insert_rows(1)
            max_c = ws.max_column
            
            ws.merge_cells(f'A1:{get_column_letter(max_c)}1')
            ws['A1'].value = "‚ö†Ô∏è PARTIDAS PENDIENTES DE CONCILIACI√ìN"
            ws['A1'].font = f_tit
            ws['A1'].fill = r_err
            ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
            ws.row_dimensions[1].height = 45
            
            for c in range(1, max_c + 1):
                cell = ws.cell(2, c)
                cell.font = f_enc
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = borde
                if cell.value and 'Banco' in str(cell.value):
                    cell.fill = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
                elif cell.value and 'Contable' in str(cell.value):
                    cell.fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
                else:
                    cell.fill = r_sec
            ws.row_dimensions[2].height = 38
            
            for r in range(3, ws.max_row + 1):
                relleno_fila = r_gri if r % 2 == 0 else None
                for c in range(1, max_c + 1):
                    cell = ws.cell(r, c)
                    if relleno_fila:
                        cell.fill = relleno_fila
                    cell.border = borde
                    cell.font = f_nor
                    cell.alignment = Alignment(vertical='center', wrap_text=True)
                    
                    hdr = ws.cell(2, c)
                    if hdr.value:
                        if 'Fecha' in str(hdr.value) and isinstance(cell.value, datetime):
                            cell.number_format = 'DD/MM/YYYY'
                        elif 'Valor' in str(hdr.value):
                            if isinstance(cell.value, (int, float)):
                                cell.number_format = '#,##0.00'
                                color = C_ERR if cell.value < 0 else C_EXI
                                cell.font = Font(name='Segoe UI', size=10, bold=True, color=color)
            
            for c in range(1, max_c + 1):
                col_letter = get_column_letter(c)
                max_len = 0
                for r in range(1, min(ws.max_row + 1, 100)):
                    val = ws.cell(r, c).value
                    if val:
                        max_len = max(max_len, len(str(val)))
                ws.column_dimensions[col_letter].width = min(max_len + 4, 55)
    
    if 'CASOS_ESPECIALES' in wb.sheetnames:
        ws = wb['CASOS_ESPECIALES']
        if ws.max_row > 1:
            ws.insert_rows(1)
            max_c = ws.max_column
            
            ws.merge_cells(f'A1:{get_column_letter(max_c)}1')
            ws['A1'].value = "üîç CASOS ESPECIALES - Requieren Revisi√≥n Manual"
            ws['A1'].font = f_tit
            ws['A1'].fill = r_adv
            ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
            ws.row_dimensions[1].height = 45
            
            for c in range(1, max_c + 1):
                cell = ws.cell(2, c)
                cell.fill = r_sec
                cell.font = f_enc
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = borde
            ws.row_dimensions[2].height = 38
            
            for r in range(3, ws.max_row + 1):
                for c in range(1, max_c + 1):
                    cell = ws.cell(r, c)
                    cell.border = borde
                    cell.font = f_nor
                    cell.fill = PatternFill(start_color="FFF8E1", end_color="FFF8E1", fill_type="solid")
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = '#,##0.00'
                        cell.font = f_num
            
            for c in range(1, max_c + 1):
                col_letter = get_column_letter(c)
                max_len = 0
                for r in range(1, min(ws.max_row + 1, 100)):
                    val = ws.cell(r, c).value
                    if val:
                        max_len = max(max_len, len(str(val)))
                ws.column_dimensions[col_letter].width = min(max_len + 4, 55)
    
    wb.save(ruta)
    print("‚ú® Formato profesional aplicado exitosamente")

# ============================================================================
# üöÄ FUNCI√ìN PRINCIPAL
# ============================================================================

def main():

    """Funci√≥n principal de conciliaci√≥n"""
    import time
    tiempo_inicio = time.time()
    
    # Establecer el directorio de trabajo al directorio del script
    os.chdir(CARPETA_TRABAJO)
    print(f"üìÅ Directorio de trabajo: {CARPETA_TRABAJO}")
    
    print("\n" + "="*70)
    print("  üè¶ CONCILIACI√ìN BANCARIA v6.5 DEFINITIVO")
    print("  üî• Motor: v6.0 + v6.1 + Optimizaciones")
    print("="*70)
    print(f"\nüìÖ {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")
    print(f"üìÅ Proyecto: GREEN PARK - BANRESERBAS")
    
    print("\n‚öôÔ∏è PAR√ÅMETROS:")
    print("‚îÄ"*70)
    print(f"  üí∞ Tolerancia exacta: ${TOLERANCIA_VALOR_EXACTA}")
    print(f"  üí∞ Tolerancia agrupaci√≥n: ${TOLERANCIA_VALOR_AGRUPACION}")
    print(f"  üìÖ Ventanas: ¬±{VENTANA_DIAS_EXACTA}d / ¬±{VENTANA_DIAS_AGRUPACION}d / ¬±{VENTANA_DIAS_FLEXIBLE}d")
    print(f"  ‚ö° L√≠mites: {MAX_PARTIDAS_AGRUPACION} partidas / {MAX_COMBINACIONES_POR_BUSQUEDA:,} comb")
    print(f"  üîß Segunda pasada: {'S√≠' if EJECUTAR_SEGUNDA_PASADA else 'No'}")
    print(f"  üîß B√∫squeda exhaustiva: {'S√≠' if EJECUTAR_BUSQUEDA_EXHAUSTIVA else 'No'}")
    
    print("\nüè¶ BANCOS:")
    print("‚îÄ"*70)
    print(f"   BanReservas")
    print(f"   Popular")
    print(f"   BHD")
    print(f"   Santa Cruz")

    print("\n" + "="*70)
    print("üìÇ CARGANDO DATOS")
    print("="*70)
    
    # B√∫squeda autom√°tica de archivos
    archivo_banco, archivo_contable, codigo_banco, nombre_banco = buscar_archivos_en_carpeta()
    
    if not archivo_banco or not archivo_contable:
        print("\n‚ùå No se pudieron encontrar los archivos necesarios")
        print("   Aseg√∫rate de que los archivos de banco y contable est√©n en la carpeta del script")
        return
    
    # Carga de Banco
    try:
        banco = cargar_banco(archivo_banco, nombre_banco, codigo_banco)
    except Exception as e:
        print(f"‚ùå Error al cargar banco: {e}")
        return
        
    # Carga de Contable
    while True:
        op = input("¬øEst√° utilizando d√≥lares? (s/n): ").strip().lower()
        if op == 's':   
            usa_dolares = True
        elif op == 'n':
            usa_dolares = False
        if op in ['s', 'n']:
            break
        else:
            print(f"‚ùå Entrada inv√°lida.")
            
    try:
        contable = cargar_contable(archivo_contable, usa_dolares, "CONTABLE")
    except Exception as e:
        print(f"‚ùå Error al cargar contable: {e}")
        return
    
    print("\n" + "="*70)
    print("üîÑ EJECUTANDO MOTOR v6.5")
    print("="*70)
    
    conciliaciones = []
    
    t1 = conciliacion_por_monto_exacto(banco, contable, conciliaciones)
    t1_5 = conciliacion_con_comisiones(banco, contable, conciliaciones)
    t1_6 = conciliacion_comisiones_agrupadas(banco, contable, conciliaciones)
    t2 = conciliacion_n_a_1_inteligente(banco, contable, conciliaciones)
    t3 = conciliacion_1_a_n_inteligente(banco, contable, conciliaciones)
    t4 = conciliacion_n_a_m_inteligente(banco, contable, conciliaciones)
    t5 = conciliacion_impuestos(banco, contable, conciliaciones)
    t6 = segunda_pasada_inteligente(banco, contable, conciliaciones)
    t7 = busqueda_exhaustiva_final(banco, contable, conciliaciones)
    t8 = conciliacion_caja_chica(banco, contable, conciliaciones)
    
    casos_especiales = detectar_casos_especiales(banco, contable)
    
    if conciliaciones:
        df_conc = pd.DataFrame(conciliaciones).sort_values(['Tipo', 'ID_Conciliacion'])
    else:
        df_conc = pd.DataFrame()
    
    if casos_especiales:
        df_especiales = pd.DataFrame(casos_especiales)
    else:
        df_especiales = pd.DataFrame()
    
    no_b = banco[~banco['Conciliado']][['Fecha', 'Concepto', 'Valor', 'Descripci√≥n']].reset_index(drop=True)
    no_c = contable[~contable['Conciliado']][['Fecha', 'Concepto', 'Valor', 'Descripci√≥n']].reset_index(drop=True)
    
    max_r = max(len(no_b), len(no_c)) if len(no_b) > 0 or len(no_c) > 0 else 1
    
    df_no_conc = pd.DataFrame({
        'Fecha_Banco': list(no_b['Fecha']) + [None] * (max_r - len(no_b)),
        'Concepto_Banco': list(no_b['Concepto']) + [''] * (max_r - len(no_b)),
        'Valor_Banco': list(no_b['Valor']) + [None] * (max_r - len(no_b)),
        'Descripcion_Banco': list(no_b['Descripci√≥n']) + [''] * (max_r - len(no_b)),
        'Fecha_Contable': list(no_c['Fecha']) + [None] * (max_r - len(no_c)),
        'Concepto_Contable': list(no_c['Concepto']) + [''] * (max_r - len(no_c)),
        'Valor_Contable': list(no_c['Valor']) + [None] * (max_r - len(no_c)),
        'Descripcion_Contable': list(no_c['Descripci√≥n']) + [''] * (max_r - len(no_c))
    })
    
    total_grupos = len(df_conc['ID_Conciliacion'].unique()) if len(df_conc) > 0 else 0
    total_estrategias = t1 + t1_5 + t1_6 + t2 + t3 + t4 + t5 + t6 + t7 + t8
    
    tiempo_total = time.time() - tiempo_inicio
    
    resumen = pd.DataFrame({
        'Concepto': [
            '‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê',
            'CONCILIACI√ìN BANCARIA v6.5 DEFINITIVO',
            '‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê',
            '',
            '‚îÄ‚îÄ‚îÄ INFORMACI√ìN GENERAL ‚îÄ‚îÄ‚îÄ',
            'Total registros Banco',
            'Total registros Contable',
            'Suma total Banco',
            'Suma total Contable',
            '',
            '‚îÄ‚îÄ‚îÄ CONCILIACIONES POR ESTRATEGIA ‚îÄ‚îÄ‚îÄ',
            '1Ô∏è‚É£  Monto Exacto (1:1)',
            '1.5 Transferencias + Comisi√≥n',
            '1.6 Comisiones Agrupadas [v6.1]',
            '2Ô∏è‚É£  Agrupaciones N‚Üí1',
            '3Ô∏è‚É£  Agrupaciones 1‚ÜíN',
            '4Ô∏è‚É£  Agrupaciones N‚ÜîM',
            '5Ô∏è‚É£  Impuestos DGII',
            '6Ô∏è‚É£  Segunda Pasada',
            '7Ô∏è‚É£  B√∫squeda Exhaustiva [v6.1]',
            '8Ô∏è‚É£  Reposici√≥n Caja Chica',
            'üì¶  TOTAL GRUPOS CONCILIADOS',
            '',
            '‚îÄ‚îÄ‚îÄ REGISTROS PROCESADOS ‚îÄ‚îÄ‚îÄ',
            'Registros Banco conciliados',
            'Registros Contable conciliados',
            'Registros Banco pendientes',
            'Registros Contable pendientes',
            '',
            '‚îÄ‚îÄ‚îÄ MONTOS ‚îÄ‚îÄ‚îÄ',
            'Suma conciliados Banco',
            'Suma conciliados Contable',
            'Suma pendientes Banco',
            'Suma pendientes Contable',
            '',
            '‚îÄ‚îÄ‚îÄ INDICADORES ‚îÄ‚îÄ‚îÄ',
            'üìä % Conciliaci√≥n Banco',
            'üìä % Conciliaci√≥n Contable',
            'üí∞ DIFERENCIA MONETARIA TOTAL',
            '‚ö†Ô∏è  Casos especiales detectados',
            '‚è±Ô∏è  Tiempo de ejecuci√≥n (seg)',
        ],
        'Valor': [
            '', '', '', '',
            '',
            len(banco),
            len(contable),
            round(banco['Valor'].sum(), 2),
            round(contable['Valor'].sum(), 2),
            '',
            '',
            t1, t1_5, t1_6, t2, t3, t4, t5, t6, t7, t8,
            total_grupos,
            '',
            '',
            banco['Conciliado'].sum(),
            contable['Conciliado'].sum(),
            len(no_b),
            len(no_c),
            '',
            '',
            round(banco[banco['Conciliado']]['Valor'].sum(), 2),
            round(contable[contable['Conciliado']]['Valor'].sum(), 2),
            round(no_b['Valor'].sum(), 2) if len(no_b) > 0 else 0,
            round(no_c['Valor'].sum(), 2) if len(no_c) > 0 else 0,
            '',
            '',
            round(banco['Conciliado'].sum() / len(banco) * 100, 2),
            round(contable['Conciliado'].sum() / len(contable) * 100, 2),
            round(banco['Valor'].sum() - contable['Valor'].sum(), 2),
            len(casos_especiales),
            round(tiempo_total, 1),
        ]
    })
    
    print("\n" + "="*70)
    print("üíæ EXPORTANDO RESULTADOS")
    print("="*70)

    # ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ
    # Generar nombre din√°mico del archivo de salida
    # ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ
    CARPETA_RESULTADOS = Path(__file__).resolve().parent / "Resultados"

    nombre_archivo = generar_nombre_conciliacion(codigo_banco, banco)
    RUTA_BASE_SALIDA = str(CARPETA_RESULTADOS / nombre_archivo)

    # Mantener l√≥gica de nombre √∫nico
    RUTA_SALIDA = obtener_ruta_unica(RUTA_BASE_SALIDA)

    
    with pd.ExcelWriter(RUTA_SALIDA, engine='openpyxl') as writer:
        resumen.to_excel(writer, sheet_name='RESUMEN', index=False)
        
        if len(df_conc) > 0:
            df_conc.to_excel(writer, sheet_name='CONCILIADOS', index=False)
        else:
            pd.DataFrame({'Mensaje': ['No hay conciliaciones']}).to_excel(
                writer, sheet_name='CONCILIADOS', index=False
            )
        
        df_no_conc.to_excel(writer, sheet_name='NO_CONCILIADOS', index=False)
        
        if len(df_especiales) > 0:
            df_especiales.to_excel(writer, sheet_name='CASOS_ESPECIALES', index=False)
    
    print("‚úÖ Datos exportados correctamente")
    
    if APLICAR_FORMATO_PROFESIONAL:
        print("üé® Aplicando formato profesional...")
        try:
            aplicar_formato_profesional(RUTA_SALIDA)
        except Exception as e:
            print(f"‚ö†Ô∏è Error al aplicar formato: {e}")
    
    perc_b = banco['Conciliado'].sum() / len(banco) * 100
    perc_c = contable['Conciliado'].sum() / len(contable) * 100
    
    print("\n" + "="*70)
    print("       ‚úÖ CONCILIACI√ìN COMPLETADA EXITOSAMENTE")
    print("="*70)
    print(f"\nüìÅ Archivo generado: {RUTA_SALIDA}\n")
    
    print("üìä RESUMEN EJECUTIVO:")
    print("‚îÄ"*70)
    print(f"üì• Datos Originales:")
    print(f"   ‚Ä¢ Banco:     {len(banco):,} registros (${banco['Valor'].sum():,.2f})")
    print(f"   ‚Ä¢ Contable:  {len(contable):,} registros (${contable['Valor'].sum():,.2f})")
    
    print(f"\n‚úÖ Conciliaciones por Estrategia:")
    print(f"   1Ô∏è‚É£  Monto Exacto:             {t1:,}")
    print(f"   1.5 Transf + Comisi√≥n:        {t1_5:,}")
    print(f"   1.6 Comisiones Agrupadas:     {t1_6:,}")
    print(f"   2Ô∏è‚É£  N‚Üí1:                      {t2:,}")
    print(f"   3Ô∏è‚É£  1‚ÜíN:                      {t3:,}")
    print(f"   4Ô∏è‚É£  N‚ÜîM:                      {t4:,}")
    print(f"   5Ô∏è‚É£  Impuestos DGII:           {t5:,}")
    print(f"   6Ô∏è‚É£  Segunda Pasada:           {t6:,}")
    print(f"   7Ô∏è‚É£  B√∫squeda Exhaustiva:      {t7:,}")
    print(f"   8Ô∏è‚É£  Reposici√≥n Caja Chica:    {t8:,}")
    print(f"   {'‚îÄ'*32}")
    print(f"   üì¶ TOTAL GRUPOS:              {total_grupos:,}")
    
    print(f"\nüìà Resultados:")
    print(f"   ‚Ä¢ Banco:     {banco['Conciliado'].sum():,} conciliados ({perc_b:.1f}%)")
    print(f"   ‚Ä¢ Contable:  {contable['Conciliado'].sum():,} conciliados ({perc_c:.1f}%)")
    print(f"   ‚Ä¢ Pendientes: Banco {len(no_b):,} | Contable {len(no_c):,}")
    
    diferencia = banco['Valor'].sum() - contable['Valor'].sum()
    print(f"\nüí∞ DIFERENCIA TOTAL: ${diferencia:,.2f}")
    
    if len(casos_especiales) > 0:
        print(f"\n‚ö†Ô∏è Casos Especiales: {len(casos_especiales)}")
        print(f"   ‚Üí Revisa la hoja 'CASOS_ESPECIALES' para m√°s detalles")
    
    print(f"\n‚è±Ô∏è Tiempo de Ejecuci√≥n: {tiempo_total:.1f} segundos")
    
    print(f"\nüéØ Evaluaci√≥n:")
    if perc_b >= 99 and perc_c >= 99:
        print("   ‚≠ê‚≠ê‚≠ê‚≠ê‚≠ê PERFECTA (‚â•99%)")
    elif perc_b >= 95 and perc_c >= 95:
        print("   ‚≠ê‚≠ê‚≠ê‚≠ê EXCELENTE (‚â•95%)")
    elif perc_b >= 85 and perc_c >= 85:
        print("   ‚≠ê‚≠ê‚≠ê MUY BUENA (‚â•85%)")
    elif perc_b >= 75 and perc_c >= 75:
        print("   ‚≠ê‚≠ê BUENA (‚â•75%)")
    elif perc_b >= 60 and perc_c >= 60:
        print("   ‚≠ê REGULAR (‚â•60%)")
    else:
        print("   ‚ö†Ô∏è NECESITA REVISI√ìN (<60%)")
    
    print("\n" + "="*70)
    print("üí° TIPS:")
    print("‚îÄ"*70)
    print("  1. Revisa RESUMEN para estad√≠sticas generales")
    print("  2. Valida CONCILIADOS para verificar agrupaciones")
    print("  3. Analiza NO_CONCILIADOS para identificar diferencias")
    if len(casos_especiales) > 0:
        print("  4. Revisa CASOS_ESPECIALES para coincidencias parciales")
    print("="*70 + "\n")

if __name__ == "__main__":
    main()